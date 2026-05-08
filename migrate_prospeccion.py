"""
Importa QoriCash_BASE_MAESTRA_2026.xlsx a la tabla prospectos de la DB principal.

Uso local:
  python3 migrate_prospeccion.py

En produccion (Render shell):
  python3 migrate_prospeccion.py
"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from run import app
from app.extensions import db
from app.models.prospecto import Prospecto

EXCEL_PATH = os.path.join(
    os.path.expanduser("~"), "Desktop", "QoriCash_BASE_MAESTRA_2026.xlsx"
)

HOJAS_SKIP = {"📊 RESUMEN"}

COL_MAP = {
    1:  "razon_social",
    2:  "ruc",
    3:  "tipo",
    4:  "rubro",
    5:  "departamento",
    6:  "provincia",
    7:  "nombre_contacto",
    8:  "cargo",
    9:  "email",
    10: "email_alt",
    11: "telefono",
    12: "cliente_lfc",
    13: "score",
    14: "clasificacion",
    15: "canal",
    16: "remitente",
    17: "tipo_ultimo_envio",
    18: "fecha_primer_contacto",
    19: "fecha_ultimo_contacto",
    20: "fecha_proximo_contacto",
    21: "num_contactos",
    22: "estado_email",
    23: "estado_comercial",
    24: "nivel_interes",
    25: "fuente",
    26: "notas",
}

HOJA_GRUPO = {
    "🎯 PIPELINE ACTIVO":      "PIPELINE ACTIVO",
    "✅ CLIENTES LFC":          "CLIENTES LFC",
    "🔥 PRIORITARIOS":          "PRIORITARIOS",
    "⭐ CALIFICADOS":           "CALIFICADOS",
    "📬 POR CONTACTAR":        "POR CONTACTAR",
    "📋 UNIVERSO CONTACTADOS":  "UNIVERSO CONTACTADOS",
    "🔄 SOFT BOUNCE":           "SOFT BOUNCE",
    "🚫 NO CONTACTAR":          "NO CONTACTAR",
}


def val(cell):
    v = cell.value
    return str(v).strip() if v is not None else None


def run():
    with app.app_context():
        print(f"Cargando {EXCEL_PATH} ...")
        if not os.path.exists(EXCEL_PATH):
            # En produccion intentar ruta alternativa
            alt = "/opt/render/project/src/QoriCash_BASE_MAESTRA_2026.xlsx"
            if os.path.exists(alt):
                global EXCEL_PATH
                EXCEL_PATH = alt
            else:
                print("ERROR: No se encontro el archivo Excel.")
                print("Sube el archivo al servidor o corre este script localmente apuntando a la DB de prod.")
                sys.exit(1)

        existing = Prospecto.query.count()
        if existing > 0:
            resp = input(f"Ya existen {existing} registros. Sobreescribir? (s/N): ")
            if resp.lower() != "s":
                print("Cancelado.")
                return
            Prospecto.query.delete()
            db.session.commit()
            print("Tabla limpiada.")

        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        emails_vistos = set()
        total = duplicados = 0

        for hoja_name in wb.sheetnames:
            if hoja_name in HOJAS_SKIP:
                continue
            ws    = wb[hoja_name]
            grupo = HOJA_GRUPO.get(hoja_name, hoja_name)
            rows  = list(ws.iter_rows(min_row=2))
            print(f"  [{hoja_name}] — {len(rows)} filas...")

            lote = []
            for row in rows:
                if len(row) <= 9:
                    continue
                email_val = row[9].value
                if not email_val:
                    continue
                email = str(email_val).strip().lower()
                if not email or "@" not in email:
                    continue
                if email in emails_vistos:
                    duplicados += 1
                    continue
                emails_vistos.add(email)

                kwargs = {"grupo": grupo}
                for col_idx, field in COL_MAP.items():
                    if col_idx < len(row):
                        v = val(row[col_idx])
                        if field in ("score", "num_contactos"):
                            try:
                                kwargs[field] = int(float(v)) if v else 0
                            except (ValueError, TypeError):
                                kwargs[field] = 0
                        else:
                            kwargs[field] = v or None

                lote.append(Prospecto(**kwargs))
                if len(lote) >= 500:
                    db.session.bulk_save_objects(lote)
                    db.session.commit()
                    total += len(lote)
                    lote = []

            if lote:
                db.session.bulk_save_objects(lote)
                db.session.commit()
                total += len(lote)

        wb.close()
        print(f"\nMigracion completada:")
        print(f"  Importados : {total}")
        print(f"  Duplicados : {duplicados}")
        print(f"  Total en BD: {Prospecto.query.count()}")


if __name__ == "__main__":
    run()
