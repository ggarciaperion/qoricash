"""
Rutas de Clientes para QoriCash Trading V2
"""
from flask import Blueprint, render_template, request, jsonify, send_file, current_app
from flask_login import login_required, current_user
from app.services.client_service import ClientService
from app.services.file_service import FileService
from app.services.notification_service import NotificationService
from app.utils.decorators import require_role
from app.utils.formatters import now_peru
from app.extensions import csrf
import io
import csv
from datetime import datetime
import json
import logging

logger = logging.getLogger(__name__)

clients_bp = Blueprint('clients', __name__, url_prefix='/clients')



@clients_bp.route('/')
@clients_bp.route('/list')
@login_required
@require_role('Master', 'Trader', 'Operador', 'Middle Office', 'App', 'Web')
def list_clients():
    """
    Página de listado de clientes

    Roles permitidos: Master, Trader, Operador, Middle Office, App, Web

    Filtrado:
    - Trader/App/Web: Solo ve sus propios clientes (created_by = user_id)
    - Otros roles: Ven todos los clientes
    """
    from app.models.client import Client

    if current_user.role in ['Trader', 'App', 'Web']:
        # Trader y Plataforma solo ven sus propios clientes
        clients = Client.query.filter_by(created_by=current_user.id).order_by(Client.created_at.desc()).all()
    else:
        # Otros roles ven todos los clientes
        clients = ClientService.get_all_clients()

    return render_template('clients/list.html',
                           user=current_user,
                           clients=clients)


@clients_bp.route('/api/list')
@login_required
@require_role('Master', 'Trader', 'Operador', 'Middle Office', 'App', 'Web')
def api_list():
    """
    API: Listar clientes (JSON)

    Filtrado:
    - Trader/App/Web: Solo sus propios clientes
    - Otros roles: Todos los clientes
    """
    from app.models.client import Client

    if current_user.role in ['Trader', 'App', 'Web']:
        clients = Client.query.filter_by(created_by=current_user.id).order_by(Client.created_at.desc()).all()
    else:
        clients = ClientService.get_all_clients()

    return jsonify({
        'success': True,
        'clients': [client.to_dict() for client in clients]
    })


@clients_bp.route('/api/create', methods=['POST'])
@login_required
@require_role('Master', 'Trader', 'App', 'Web')
def create_client():
    """
    API/Endpoint para crear nuevo cliente.
    Acepta JSON (application/json) o form-data multipart (desde modal), con archivos en request.files
    """
    files = request.files or {}
    # Si es JSON, request.get_json() devolverá dict; si viene form-data, usar request.form
    if request.is_json:
        data = request.get_json() or {}
    else:
        # request.form es ImmutableMultiDict; convertir a dict simple
        data = request.form.to_dict(flat=True)

    # Normalizar claves a strings y strip
    for k, v in list(data.items()):
        if isinstance(v, str):
            data[k] = v.strip()

    file_service = FileService()
    uploaded = {}

    # Subida de archivos desde el modal (si vienen): mapear a nombres de campo esperados en el servicio
    try:
        # DNI/CE
        if 'dni_front' in files:
            ok, msg, url = file_service.upload_file(files['dni_front'], 'dni', f"{data.get('dni')}_front")
            if not ok:
                return jsonify({'success': False, 'message': msg}), 400
            uploaded['dni_front_url'] = url
        if 'dni_back' in files:
            ok, msg, url = file_service.upload_file(files['dni_back'], 'dni', f"{data.get('dni')}_back")
            if not ok:
                return jsonify({'success': False, 'message': msg}), 400
            uploaded['dni_back_url'] = url

        # RUC / representante
        if 'dni_representante_front' in files:
            ok, msg, url = file_service.upload_file(files['dni_representante_front'], 'dni', f"{data.get('dni')}_rep_front")
            if not ok:
                return jsonify({'success': False, 'message': msg}), 400
            uploaded['dni_representante_front_url'] = url
        if 'dni_representante_back' in files:
            ok, msg, url = file_service.upload_file(files['dni_representante_back'], 'dni', f"{data.get('dni')}_rep_back")
            if not ok:
                return jsonify({'success': False, 'message': msg}), 400
            uploaded['dni_representante_back_url'] = url
        if 'ficha_ruc' in files:
            ok, msg, url = file_service.upload_file(files['ficha_ruc'], 'ruc', f"{data.get('dni')}_ruc")
            if not ok:
                return jsonify({'success': False, 'message': msg}), 400
            uploaded['ficha_ruc_url'] = url
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error en subida de archivos: {str(e)}'}), 500

    # Incorporar URLs subidas al payload
    data.update(uploaded)

    # Bank accounts: primer preferencia campo 'bank_accounts' (ya implementado en JS),
    # si no viene, recoger legacy indexed fields: bank_name1, bank_account_number1, etc.
    bank_accounts_raw = data.get('bank_accounts')
    if bank_accounts_raw:
        if isinstance(bank_accounts_raw, str):
            try:
                data['bank_accounts'] = json.loads(bank_accounts_raw)
            except Exception:
                return jsonify({'success': False, 'message': 'Formato inválido para bank_accounts'}), 400
    # Si request included legacy fields (bank_name1, bank_account_number1, etc.), el servicio los detectará

    success, message, client = ClientService.create_client(current_user=current_user, data=data, files=files)
    if success:
        try:
            NotificationService.notify_new_client(client, current_user)
        except Exception:
            pass
        try:
            from app.services.email_service import EmailService
            EmailService.send_new_client_registration_email(client, current_user)
        except Exception as e:
            logger.warning(f'Error al enviar email de nuevo cliente: {str(e)}')
        return jsonify({'success': True, 'message': message, 'client': client.to_dict()}), 201
    else:
        return jsonify({'success': False, 'message': message}), 400


@clients_bp.route('/api/update/<int:client_id>', methods=['PUT', 'PATCH'])
@login_required
@require_role('Master', 'Trader', 'Operador')
def update_client(client_id):
    """
    API: Actualizar cliente existente
    """
    data = request.get_json() or {}
    success, message, client = ClientService.update_client(current_user=current_user, client_id=client_id, data=data)
    if success:
        return jsonify({'success': True, 'message': message, 'client': client.to_dict()})
    else:
        return jsonify({'success': False, 'message': message}), 400


@clients_bp.route('/api/change_status/<int:client_id>', methods=['PATCH'])
@login_required
@require_role('Master', 'Middle Office')
def change_status(client_id):
    """
    API: Cambiar estado del cliente (Activo/Inactivo)
    SOLO Master y Middle Office - Operador YA NO tiene este permiso
    """
    data = request.get_json() or {}
    new_status = data.get('status')

    if not new_status:
        return jsonify({'success': False, 'message': 'El estado es requerido'}), 400

    # Obtener estado anterior antes de cambiarlo
    client_before = ClientService.get_client_by_id(client_id)
    old_status = client_before.status if client_before else None

    success, message, client = ClientService.change_client_status(current_user=current_user, client_id=client_id, new_status=new_status)
    if success:
        # Si el cliente fue activado (de Inactivo a Activo), generar contraseña y enviar email
        if old_status == 'Inactivo' and new_status == 'Activo':
            try:
                from app.services.email_service import EmailService
                from app.utils.password_generator import generate_simple_password
                from app.extensions import db

                # Generar contraseña temporal SOLO al activar la cuenta
                temporary_password = generate_simple_password(length=10)

                # Establecer contraseña en el cliente
                client.set_password(temporary_password)
                client.requires_password_change = True
                db.session.commit()

                logger.info(f'✅ Contraseña temporal generada para cliente {client.dni} al activar cuenta')

                # Enviar correo con contraseña temporal y el trader que creó al cliente
                trader = client.creator if hasattr(client, 'creator') and client.creator else current_user
                EmailService.send_client_activation_email(client, trader, temporary_password)
            except Exception as e:
                # No bloquear por errores de email
                logger.warning(f'Error al enviar email de cliente activado: {str(e)}')
        return jsonify({'success': True, 'message': message, 'client': client.to_dict()})
    else:
        return jsonify({'success': False, 'message': message}), 400


@clients_bp.route('/api/delete/<int:client_id>', methods=['DELETE'])
@login_required
@require_role('Master')
def delete_client(client_id):
    """
    API: Eliminar cliente
    """
    success, message = ClientService.delete_client(current_user=current_user, client_id=client_id)
    if success:
        return jsonify({'success': True, 'message': message})
    else:
        return jsonify({'success': False, 'message': message}), 400


@clients_bp.route('/api/upload_documents/<int:client_id>', methods=['POST'])
@login_required
@require_role('Master', 'Trader', 'App', 'Web')
def upload_documents(client_id):
    """
    API: Subir documentos del cliente
    """
    client = ClientService.get_client_by_id(client_id)
    if not client:
        return jsonify({'success': False, 'message': 'Cliente no encontrado'}), 404

    file_service = FileService()
    document_urls = {}

    try:
        if client.document_type == 'RUC':
            # Subir documentos RUC
            if 'dni_representante_front' in request.files:
                file = request.files['dni_representante_front']
                ok, msg, url = file_service.upload_file(file, 'dni', f"REP_{client.dni}")
                if ok:
                    document_urls['dni_representante_front_url'] = url
                else:
                    return jsonify({'success': False, 'message': f'Error DNI representante frontal: {msg}'}), 400

            if 'dni_representante_back' in request.files:
                file = request.files['dni_representante_back']
                ok, msg, url = file_service.upload_file(file, 'dni', f"REP_{client.dni}")
                if ok:
                    document_urls['dni_representante_back_url'] = url
                else:
                    return jsonify({'success': False, 'message': f'Error DNI representante reverso: {msg}'}), 400

            if 'ficha_ruc' in request.files:
                file = request.files['ficha_ruc']
                ok, msg, url = file_service.upload_file(file, 'ruc', f"RUC_{client.dni}")
                if ok:
                    document_urls['ficha_ruc_url'] = url
                else:
                    return jsonify({'success': False, 'message': f'Error Ficha RUC: {msg}'}), 400

        else:
            # Subir documentos DNI/CE
            if 'dni_front' in request.files:
                file = request.files['dni_front']
                ok, msg, url = file_service.upload_dni_front(file, client.dni)
                if ok:
                    document_urls['dni_front_url'] = url
                else:
                    return jsonify({'success': False, 'message': f'Error documento frontal: {msg}'}), 400

            if 'dni_back' in request.files:
                file = request.files['dni_back']
                ok, msg, url = file_service.upload_dni_back(file, client.dni)
                if ok:
                    document_urls['dni_back_url'] = url
                else:
                    return jsonify({'success': False, 'message': f'Error documento reverso: {msg}'}), 400

        # Actualizar cliente con URLs
        if document_urls:
            success, message, client = ClientService.update_client_documents(current_user=current_user, client_id=client_id, document_urls=document_urls)
            if success:
                return jsonify({'success': True, 'message': message, 'client': client.to_dict()})
            else:
                return jsonify({'success': False, 'message': message}), 400

        return jsonify({'success': False, 'message': 'No se seleccionó ningún archivo'}), 400

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al subir documentos: {str(e)}'}), 500


@clients_bp.route('/api/<int:client_id>/stats')
@login_required
@require_role('Master', 'Trader', 'Operador', 'Middle Office')
def get_stats(client_id):
    """
    API: Obtener estadísticas de un cliente
    """
    stats = ClientService.get_client_stats(client_id)

    if stats:
        return jsonify({'success': True, 'stats': stats})
    else:
        return jsonify({'success': False, 'message': 'Cliente no encontrado'}), 404


@clients_bp.route('/api/search')
@login_required
@require_role('Master', 'Trader', 'Operador', 'Middle Office', 'App', 'Web')
def search():
    """
    API: Buscar clientes

    Filtrado:
    - Trader/App/Web: Solo sus propios clientes
    - Otros roles: Todos los clientes
    """
    from app.models.client import Client
    from sqlalchemy import or_

    query = request.args.get('q', '').strip()

    if not query or len(query) < 3:
        return jsonify({'success': False, 'message': 'La búsqueda debe tener al menos 3 caracteres'}), 400

    if current_user.role in ['Trader', 'App', 'Web']:
        # Buscar solo en sus propios clientes
        search = f"%{query}%"
        clients = Client.query.filter(
            Client.created_by == current_user.id,
            or_(
                Client.dni.ilike(search),
                Client.email.ilike(search),
                Client.apellido_paterno.ilike(search),
                Client.apellido_materno.ilike(search),
                Client.nombres.ilike(search),
                Client.razon_social.ilike(search)
            )
        ).all()
    else:
        # Buscar en todos los clientes
        clients = ClientService.search_clients(query)

    return jsonify({'success': True, 'clients': [client.to_dict() for client in clients]})


@clients_bp.route('/api/approve_documents/<int:client_id>', methods=['POST'])
@login_required
@require_role('Master', 'Middle Office')
def approve_documents(client_id):
    """
    API: Aprobar documentos del cliente y reactivarlo (Middle Office)

    Este endpoint permite a Middle Office:
    - Marcar los documentos del cliente como completos
    - Resetear el contador de operaciones sin documentos
    - Reactivar el cliente automáticamente
    """
    try:
        from app.extensions import db
        from app.models.audit_log import AuditLog

        client = ClientService.get_client_by_id(client_id)
        if not client:
            return jsonify({'success': False, 'message': 'Cliente no encontrado'}), 404

        # Verificar que el cliente tiene documentos completos antes de aprobar
        if not client.check_documents_uploaded():
            # Determinar qué documentos faltan
            if client.document_type in ('DNI', 'CE'):
                missing = []
                if not client.dni_front_url:
                    missing.append('DNI frente')
                if not client.dni_back_url:
                    missing.append('DNI reverso')
                docs_faltantes = ', '.join(missing)
                return jsonify({
                    'success': False,
                    'message': f'El cliente aún no tiene documentos completos. Faltan: {docs_faltantes}'
                }), 400
            else:  # RUC
                missing = []
                if not client.dni_representante_front_url:
                    missing.append('DNI representante frente')
                if not client.dni_representante_back_url:
                    missing.append('DNI representante reverso')
                if not client.ficha_ruc_url:
                    missing.append('Ficha RUC')
                docs_faltantes = ', '.join(missing)
                return jsonify({
                    'success': False,
                    'message': f'El cliente aún no tiene documentos completos. Faltan: {docs_faltantes}'
                }), 400

        # Guardar estado anterior para auditoría
        was_inactive = client.status == 'Inactivo'
        old_reason = client.inactive_reason

        # Aprobar documentos y resetear contadores
        logger.info(f'📋 Aprobando documentos para cliente {client.dni} - {client.full_name}')
        client.complete_documents_and_reset()
        logger.info(f'✅ has_complete_documents establecido a: {client.has_complete_documents}')

        # Auditoría
        AuditLog.log_action(
            user_id=current_user.id,
            action='APPROVE_DOCUMENTS',
            entity='Client',
            entity_id=client.id,
            details=f'Documentos aprobados para {client.full_name}. Cliente reactivado automáticamente. '
                   f'Contador de operaciones sin docs reseteado. Estado anterior: {old_reason or "N/A"}'
        )

        # Recalcular perfil de riesgo automáticamente - auto_commit=False para hacer un solo commit
        from app.services.compliance_service import ComplianceService
        ComplianceService.update_client_risk_profile(client.id, current_user.id, auto_commit=False)

        db.session.commit()
        logger.info(f'💾 Cambios guardados en BD para cliente {client.dni}')

        # Enviar email de activación con contraseña temporal si corresponde
        if was_inactive:
            try:
                from app.services.email_service import EmailService
                from app.utils.password_generator import generate_simple_password

                # Generar contraseña temporal SOLO al activar la cuenta
                temporary_password = generate_simple_password(length=10)

                # Establecer contraseña en el cliente
                client.set_password(temporary_password)
                client.requires_password_change = True
                db.session.commit()

                logger.info(f'✅ Contraseña temporal generada para cliente {client.dni} al aprobar documentos')

                # Enviar email con contraseña temporal
                EmailService.send_client_activation_email(client, current_user, temporary_password)
            except Exception as e:
                logger.warning(f'Error al enviar email de activación: {str(e)}')

        # Enviar notificación Socket.IO al cliente
        logger.info(f'📡 Enviando notificación Socket.IO a cliente {client.dni}...')
        try:
            from app.services.notification_service import NotificationService
            NotificationService.notify_client_documents_approved(client)
            logger.info(f'✅ Notificación Socket.IO enviada correctamente al room client_{client.dni}')
        except Exception as e:
            logger.error(f'❌ Error al enviar notificación Socket.IO al cliente: {str(e)}')
            logger.exception(e)

        return jsonify({
            'success': True,
            'message': 'Documentos aprobados exitosamente. Cliente reactivado.',
            'client': client.to_dict()
        })

    except Exception as e:
        db.session.rollback()
        logger.exception(f'Error al aprobar documentos del cliente {client_id}')
        return jsonify({
            'success': False,
            'message': f'Error al aprobar documentos: {str(e)}'
        }), 500


@clients_bp.route('/api/pending_documents')
@login_required
@require_role('Master', 'Middle Office')
def get_pending_documents():
    """
    API: Obtener lista de clientes con documentos pendientes

    Retorna clientes que:
    - No tienen documentos completos
    - Están inactivos por falta de documentos
    - Tienen documents_pending_since establecido
    """
    from app.models.client import Client

    try:
        # Obtener clientes con documentos pendientes
        clients = Client.query.filter(
            Client.has_complete_documents == False,
            Client.documents_pending_since.isnot(None)
        ).order_by(Client.documents_pending_since.asc()).all()

        # Construir respuesta con información detallada
        clients_data = []
        for client in clients:
            client_dict = client.to_dict()

            # Agregar información adicional de documentos
            if client.document_type in ('DNI', 'CE'):
                client_dict['required_docs'] = {
                    'dni_front': bool(client.dni_front_url),
                    'dni_back': bool(client.dni_back_url)
                }
                client_dict['missing_docs'] = []
                if not client.dni_front_url:
                    client_dict['missing_docs'].append('DNI frente')
                if not client.dni_back_url:
                    client_dict['missing_docs'].append('DNI reverso')
            else:  # RUC
                client_dict['required_docs'] = {
                    'dni_representante_front': bool(client.dni_representante_front_url),
                    'dni_representante_back': bool(client.dni_representante_back_url),
                    'ficha_ruc': bool(client.ficha_ruc_url)
                }
                client_dict['missing_docs'] = []
                if not client.dni_representante_front_url:
                    client_dict['missing_docs'].append('DNI representante frente')
                if not client.dni_representante_back_url:
                    client_dict['missing_docs'].append('DNI representante reverso')
                if not client.ficha_ruc_url:
                    client_dict['missing_docs'].append('Ficha RUC')

            # Información de límites
            client_dict['limits_info'] = {
                'operations_count': client.operations_without_docs_count,
                'operations_limit': client.operations_without_docs_limit,
                'max_amount': float(client.max_amount_without_docs) if client.max_amount_without_docs else None,
                'pending_since': client.documents_pending_since.isoformat() if client.documents_pending_since else None
            }

            clients_data.append(client_dict)

        return jsonify({
            'success': True,
            'count': len(clients_data),
            'clients': clients_data
        })

    except Exception as e:
        logger.exception('Error al obtener clientes con documentos pendientes')
        return jsonify({
            'success': False,
            'message': f'Error al obtener clientes: {str(e)}'
        }), 500


@clients_bp.route('/api/export/csv')
@login_required
@require_role('Master')
def export_csv():
    """
    API: Exportar clientes a Excel con formato de tabla
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        # Obtener todos los clientes con eager loading
        from sqlalchemy.orm import joinedload
        from app.models.client import Client
        clients = Client.query.options(joinedload(Client.creator)).order_by(Client.created_at.desc()).all()

        if not clients:
            return jsonify({'success': False, 'message': 'No hay clientes para exportar'}), 404

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"

        # Definir columnas en orden específico (ahora con más detalle)
        headers = [
            'ID',
            'Tipo Documento',
            'Número Documento',
            'Nombre Completo',
            'Persona Contacto',  # Para RUC
            'Email',
            'Teléfono',
            'Dirección',
            'Distrito',
            'Provincia',
            'Departamento',
            'Usuario Registro',
            'Fecha Registro',
            'Estado',
            'Total Operaciones',
            'Operaciones Completadas',
            'Cuenta Bancaria 1',
            'Cuenta Bancaria 2',
            'Cuenta Bancaria 3',
            'Cuenta Bancaria 4',
            'Cuenta Bancaria 5',
            'Cuenta Bancaria 6'
        ]

        # Escribir encabezados con formato
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Escribir datos
        for row_num, client in enumerate(clients, 2):
            col = 1
            ws.cell(row=row_num, column=col, value=client.id); col += 1
            ws.cell(row=row_num, column=col, value=client.document_type); col += 1
            ws.cell(row=row_num, column=col, value=client.dni); col += 1
            ws.cell(row=row_num, column=col, value=client.full_name or ''); col += 1

            # Persona de contacto (solo para RUC)
            ws.cell(row=row_num, column=col, value=client.persona_contacto if client.document_type == 'RUC' else ''); col += 1

            ws.cell(row=row_num, column=col, value=client.email); col += 1
            ws.cell(row=row_num, column=col, value=client.phone or ''); col += 1

            # Dirección separada en columnas
            ws.cell(row=row_num, column=col, value=client.direccion or ''); col += 1
            ws.cell(row=row_num, column=col, value=client.distrito or ''); col += 1
            ws.cell(row=row_num, column=col, value=client.provincia or ''); col += 1
            ws.cell(row=row_num, column=col, value=client.departamento or ''); col += 1

            ws.cell(row=row_num, column=col, value=client.creator.email if client.creator else 'N/A'); col += 1
            ws.cell(row=row_num, column=col, value=client.created_at.strftime('%d/%m/%Y %H:%M') if client.created_at else ''); col += 1
            ws.cell(row=row_num, column=col, value=client.status); col += 1
            ws.cell(row=row_num, column=col, value=client.get_total_operations()); col += 1
            ws.cell(row=row_num, column=col, value=client.get_completed_operations()); col += 1

            # Cuentas bancarias (hasta 6)
            bank_accounts = client.bank_accounts or []
            for i in range(6):
                if i < len(bank_accounts):
                    account = bank_accounts[i]
                    account_str = f"{account.get('bank_name', '')} | {account.get('account_type', '')} | {account.get('currency', '')} | {account.get('account_number', '')}"
                    ws.cell(row=row_num, column=col, value=account_str)
                else:
                    ws.cell(row=row_num, column=col, value='')
                col += 1

        # Ajustar ancho de columnas
        column_widths = {
            'A': 8,   # ID
            'B': 15,  # Tipo Documento
            'C': 18,  # Número Documento
            'D': 35,  # Nombre Completo
            'E': 30,  # Persona Contacto
            'F': 30,  # Email
            'G': 15,  # Teléfono
            'H': 30,  # Dirección
            'I': 20,  # Distrito
            'J': 20,  # Provincia
            'K': 20,  # Departamento
            'L': 30,  # Usuario Registro
            'M': 18,  # Fecha Registro
            'N': 12,  # Estado
            'O': 18,  # Total Ops
            'P': 20,  # Ops Completadas
            'Q': 50,  # Cuenta 1
            'R': 50,  # Cuenta 2
            'S': 50,  # Cuenta 3
            'T': 50,  # Cuenta 4
            'U': 50,  # Cuenta 5
            'V': 50   # Cuenta 6
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Guardar en memoria
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Nombre del archivo con fecha
        filename = f"clientes_qoricash_{now_peru().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error al exportar: {str(e)}'}), 500


@clients_bp.route('/api/<int:client_id>')
@login_required
@require_role('Master', 'Trader', 'Operador', 'Middle Office', 'App', 'Web')
def get_client(client_id):
    """
    API: Obtener detalles de un cliente

    Filtrado:
    - Trader/App/Web: Solo pueden acceder a sus propios clientes
    - Otros roles: Pueden acceder a cualquier cliente
    """
    try:
        from sqlalchemy.orm import joinedload
        from app.models.client import Client

        # Cargar cliente con relaciones necesarias (eager loading)
        client = Client.query.options(
            joinedload(Client.creator)
        ).filter_by(id=client_id).first()

        if not client:
            return jsonify({'success': False, 'message': 'Cliente no encontrado'}), 404

        # Verificar que Trader/App/Web solo accedan a sus propios clientes
        if current_user.role in ['Trader', 'App', 'Web']:
            if client.created_by != current_user.id:
                return jsonify({'success': False, 'message': 'No tiene permisos para acceder a este cliente'}), 403

        return jsonify({'success': True, 'client': client.to_dict(include_stats=True)})

    except Exception as e:
        import traceback
        logger.error(f'Error en get_client({client_id}): {str(e)}')
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': 'Error interno del servidor'
        }), 500


@clients_bp.route('/api/active')
@login_required
@require_role('Master', 'Trader', 'Operador', 'Middle Office', 'App', 'Web')
def get_active():
    """
    API: Obtener solo clientes activos

    Filtrado:
    - Trader/App/Web: Solo sus propios clientes activos
    - Otros roles: Todos los clientes activos
    """
    from app.models.client import Client

    if current_user.role in ['Trader', 'App', 'Web']:
        clients = Client.query.filter_by(
            created_by=current_user.id,
            status='Activo'
        ).order_by(Client.created_at.desc()).all()
    else:
        clients = ClientService.get_active_clients()

    return jsonify({'success': True, 'clients': [client.to_dict() for client in clients]})


@clients_bp.route('/api/upload_validation_oc/<int:client_id>', methods=['POST'])
@login_required
@require_role('Master', 'Operador')
def upload_validation_oc(client_id):
    """
    API: Subir documento de validación del Oficial de Cumplimiento (OC)
    Solo para roles Master y Operador
    """
    client = ClientService.get_client_by_id(client_id)
    if not client:
        return jsonify({'success': False, 'message': 'Cliente no encontrado'}), 404

    if 'validation_oc_file' not in request.files:
        return jsonify({'success': False, 'message': 'No se envió ningún archivo'}), 400

    file = request.files['validation_oc_file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No se seleccionó ningún archivo'}), 400

    try:
        file_service = FileService()

        # Subir archivo a Cloudinary con folder específico para validación OC
        ok, msg, url = file_service.upload_file(file, 'validation_oc', f"OC_{client.dni}")

        if not ok:
            return jsonify({'success': False, 'message': f'Error al subir archivo: {msg}'}), 400

        # Actualizar cliente con URL de validación OC
        client.validation_oc_url = url
        from app.extensions import db
        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Documento de validación OC subido correctamente',
            'validation_oc_url': url
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al subir documento: {str(e)}'}), 500


@clients_bp.route('/api/traders/active')
@login_required
@csrf.exempt
def get_active_traders():
    """
    API: Obtener lista de traders activos para reasignación
    Solo Master puede acceder
    """
    # Verificar rol manualmente
    if current_user.role != 'Master':
        return jsonify({'success': False, 'message': 'No autorizado'}), 403

    try:
        from app.models.user import User
        traders = User.query.filter_by(role='Trader', status='Activo').order_by(User.username).all()

        traders_list = [{
            'id': trader.id,
            'username': trader.username,
            'email': trader.email,
            'total_clients': len(ClientService.get_clients_by_trader(trader.id))
        } for trader in traders]

        return jsonify({
            'success': True,
            'traders': traders_list
        })
    except Exception as e:
        logger.error(f'Error al obtener traders: {str(e)}')
        return jsonify({'success': False, 'message': f'Error al obtener traders: {str(e)}'}), 500


@clients_bp.route('/api/reassign/<int:client_id>', methods=['POST'])
@login_required
@csrf.exempt
def reassign_client(client_id):
    """
    API: Reasignar un cliente individual a otro trader
    Solo Master puede reasignar
    """
    # Verificar rol manualmente
    if current_user.role != 'Master':
        return jsonify({'success': False, 'message': 'No autorizado. Solo Master puede reasignar clientes'}), 403

    try:
        data = request.get_json(force=True, silent=True)
        if not data:
            return jsonify({'success': False, 'message': 'No se recibieron datos válidos'}), 400

        new_trader_id = data.get('new_trader_id')

        if not new_trader_id:
            return jsonify({'success': False, 'message': 'El ID del nuevo trader es requerido'}), 400

        success, message, client = ClientService.reassign_client(
            current_user=current_user,
            client_id=client_id,
            new_trader_id=new_trader_id
        )

        if success:
            # Enviar notificaciones
            try:
                NotificationService.notify_client_reassignment(client, current_user, new_trader_id)
            except Exception:
                logger.warning('Error al enviar notificación de reasignación')

            return jsonify({
                'success': True,
                'message': message,
                'client': client.to_dict(include_stats=True)
            })
        else:
            return jsonify({'success': False, 'message': message}), 400

    except Exception as e:
        logger.error(f'Error en reasignación: {str(e)}')
        return jsonify({'success': False, 'message': f'Error al reasignar cliente: {str(e)}'}), 500


@clients_bp.route('/api/reassign/bulk', methods=['POST'])
@login_required
@csrf.exempt
def reassign_clients_bulk():
    """
    API: Reasignar múltiples clientes a un nuevo trader
    Solo Master puede reasignar
    """
    # Verificar rol manualmente para mejor control
    if current_user.role != 'Master':
        return jsonify({'success': False, 'message': 'No autorizado. Solo Master puede reasignar clientes'}), 403

    try:
        data = request.get_json(force=True, silent=True)
        if not data:
            return jsonify({'success': False, 'message': 'No se recibieron datos válidos'}), 400

        client_ids = data.get('client_ids', [])
        new_trader_id = data.get('new_trader_id')

        if not new_trader_id:
            return jsonify({'success': False, 'message': 'El ID del nuevo trader es requerido'}), 400

        if not client_ids or not isinstance(client_ids, list):
            return jsonify({'success': False, 'message': 'Debe proporcionar una lista de clientes'}), 400

        success, message, results = ClientService.reassign_clients_bulk(
            current_user=current_user,
            client_ids=client_ids,
            new_trader_id=new_trader_id
        )

        if success or results:
            # Enviar notificación al nuevo trader
            try:
                from app.models.user import User
                new_trader = User.query.get(new_trader_id)
                if new_trader and results:
                    NotificationService.notify_bulk_client_reassignment(
                        new_trader,
                        current_user,
                        len(results.get('success', []))
                    )
            except Exception:
                logger.warning('Error al enviar notificación de reasignación masiva')

            return jsonify({
                'success': success,
                'message': message,
                'results': results
            })
        else:
            return jsonify({'success': False, 'message': message}), 400

    except Exception as e:
        logger.error(f'Error en reasignación masiva: {str(e)}')
        return jsonify({'success': False, 'message': f'Error al reasignar clientes: {str(e)}'}), 500


@clients_bp.route('/detail/<int:client_id>')
@login_required
@require_role('Master', 'Trader', 'Operador', 'Middle Office')
def client_detail(client_id):
    """
    Página de detalle completo del cliente
    Muestra información, documentos, operaciones, perfil de riesgo
    """
    client = ClientService.get_client_by_id(client_id)

    if not client:
        return render_template('errors/404.html', message='Cliente no encontrado'), 404

    # Obtener perfil de riesgo — crear si no existe (aplica a todos los canales)
    from app.models.compliance import ClientRiskProfile, RestrictiveListCheck
    from app.models.operation import Operation
    risk_profile = ClientRiskProfile.query.filter_by(client_id=client_id).first()
    if not risk_profile:
        try:
            from app.services.compliance_service import ComplianceService
            ComplianceService.update_client_risk_profile(client_id, current_user.id)
            risk_profile = ClientRiskProfile.query.filter_by(client_id=client_id).first()
        except Exception as rp_exc:
            logger.warning(f'No se pudo crear perfil de riesgo para cliente {client_id}: {rp_exc}')

    # Obtener últimas operaciones (corregido para SQLAlchemy 2.x)
    recent_operations = Operation.query.filter_by(client_id=client_id).order_by(Operation.created_at.desc()).limit(10).all()

    # Obtener verificaciones de listas restrictivas (manual + auto-screening)
    restrictive_checks = RestrictiveListCheck.query.filter_by(
        client_id=client_id,
        is_manual=True
    ).order_by(RestrictiveListCheck.checked_at.desc()).all()

    # Calcular estado de verificación de listas restrictivas
    restrictive_status = {
        'has_checks': len(restrictive_checks) > 0,
        'total_checks': len(restrictive_checks),
        'all_clean': False,
        'has_matches': False,
        'partial': False
    }

    if restrictive_checks:
        # Obtener la última verificación manual
        last_check = restrictive_checks[0]

        # Contar verificaciones por tipo en la última verificación
        list_types = ['ofac', 'onu', 'uif', 'interpol', 'denuncias', 'otras_listas']
        total_lists = len(list_types)
        verified_lists = 0
        clean_lists = 0
        matched_lists = 0

        for list_type in list_types:
            if getattr(last_check, f'{list_type}_checked', False):
                verified_lists += 1
                result = getattr(last_check, f'{list_type}_result', None)
                if result == 'Clean':
                    clean_lists += 1
                elif result == 'Match':
                    matched_lists += 1

        # Determinar estado
        if matched_lists > 0:
            restrictive_status['has_matches'] = True
        elif verified_lists == total_lists and clean_lists == total_lists:
            restrictive_status['all_clean'] = True
        elif verified_lists > 0:
            restrictive_status['partial'] = True
    else:
        # Sin verificación manual — considerar el último auto-screening resuelto
        auto_check = RestrictiveListCheck.query.filter_by(
            client_id=client_id,
            is_manual=False,
            list_type='AUTO_COMPREHENSIVE'
        ).order_by(RestrictiveListCheck.checked_at.desc()).first()

        if auto_check and auto_check.result:
            restrictive_status['has_checks'] = True
            if auto_check.result == 'Clean':
                restrictive_status['all_clean'] = True
            elif auto_check.result == 'Match':
                restrictive_status['has_matches'] = True
            else:
                restrictive_status['partial'] = True

    return render_template('clients/detail.html',
                         client=client,
                         risk_profile=risk_profile,
                         recent_operations=recent_operations,
                         restrictive_status=restrictive_status)


@clients_bp.route('/api/trader/<int:trader_id>/clients')
@login_required
@csrf.exempt
def get_trader_clients(trader_id):
    """
    API: Obtener todos los clientes de un trader específico
    Solo Master puede acceder
    """
    # Verificar rol manualmente
    if current_user.role != 'Master':
        return jsonify({'success': False, 'message': 'No autorizado'}), 403

    try:
        clients = ClientService.get_clients_by_trader(trader_id)

        return jsonify({
            'success': True,
            'clients': [client.to_dict(include_stats=True) for client in clients],
            'total': len(clients)
        })
    except Exception as e:
        logger.error(f'Error al obtener clientes del trader: {str(e)}')
        return jsonify({'success': False, 'message': f'Error al obtener clientes: {str(e)}'}), 500


@clients_bp.route('/api/ruc-lookup')
@login_required
@require_role('Master', 'Trader', 'Operador', 'Middle Office')
def ruc_lookup():
    """
    Consulta información de un RUC:
    1. Verifica si ya existe en la base de datos (evita duplicados).
    2. Si no existe, consulta apis.net.pe (SUNAT) para obtener razón social,
       estado del contribuyente y condición (habido/no habido).
    """
    import urllib.request
    import urllib.error
    from app.models.client import Client

    ruc = request.args.get('ruc', '').strip()

    # ── Validación básica ────────────────────────────────────────────────────
    if not ruc:
        return jsonify({'success': False, 'message': 'RUC requerido'}), 400

    if not ruc.isdigit() or len(ruc) != 11:
        return jsonify({'success': False, 'message': 'El RUC debe tener exactamente 11 dígitos numéricos'}), 400

    if ruc[0] not in ('1', '2'):
        return jsonify({'success': False, 'message': 'RUC inválido — debe comenzar con 1 o 2'}), 400

    # ── Verificar si ya existe en la base de datos ───────────────────────────
    existing = Client.query.filter_by(dni=ruc, document_type='RUC').first()
    if existing:
        return jsonify({
            'success': False,
            'already_exists': True,
            'message': f'Este RUC ya se encuentra registrado con la razón social: {existing.razon_social or "(sin nombre)"}',
            'client_id': existing.id,
        }), 200

    # ── Consulta externa ──────────────────────────────────────────────────────
    import json as _json
    import os

    token = (
        os.environ.get('APIS_NET_PE_TOKEN') or
        current_app.config.get('APIS_NET_PE_TOKEN', '')
    ).strip()

    try:
        # Primario: decolecta.com con token (más completo)
        # Fallback: apis.net.pe v1 libre
        if token:
            url  = f'https://api.decolecta.com/v1/sunat/ruc?numero={ruc}'
            hdrs = {
                'Accept':        'application/json',
                'User-Agent':    'QoriCash/2.0',
                'Authorization': f'Bearer {token}',
            }
        else:
            url  = f'https://api.apis.net.pe/v1/ruc?numero={ruc}'
            hdrs = {'Accept': 'application/json', 'User-Agent': 'QoriCash/2.0'}

        req = urllib.request.Request(url, headers=hdrs)
        with urllib.request.urlopen(req, timeout=8) as resp:
            data = _json.loads(resp.read().decode())

        # decolecta usa 'razon_social'; apis.net.pe usa 'nombre'
        razon_social = (
            data.get('razon_social') or data.get('nombre') or data.get('razonSocial') or ''
        ).strip().upper()

        if not razon_social:
            return jsonify({'success': False, 'message': 'RUC no encontrado en SUNAT'}), 404

        return jsonify({
            'success':               True,
            'ruc':                   ruc,
            'razon_social':          razon_social,
            'estado':                (data.get('estado') or '').strip().upper(),
            'condicion':             (data.get('condicion') or '').strip().upper(),
            'direccion':             (data.get('direccion') or '').strip(),
            'departamento':          (data.get('departamento') or '').strip(),
            'provincia':             (data.get('provincia') or '').strip(),
            'distrito':              (data.get('distrito') or '').strip(),
            'es_agente_retencion':   data.get('es_agente_retencion', False),
            'es_buen_contribuyente': data.get('es_buen_contribuyente', False),
        })

    except urllib.error.HTTPError as e:
        if e.code == 404:
            return jsonify({'success': False, 'message': 'RUC no encontrado en los registros de SUNAT'}), 404
        logger.error(f'[RUC lookup] HTTP {e.code} para RUC {ruc}')
        return jsonify({'success': False, 'message': 'Error al consultar SUNAT — intente nuevamente'}), 502
    except urllib.error.URLError:
        return jsonify({'success': False, 'message': 'No se pudo conectar al servicio de consulta — verifique la conexión'}), 503
    except Exception as e:
        logger.error(f'[RUC lookup] Error inesperado: {e}')
        return jsonify({'success': False, 'message': 'Error interno — ingrese la razón social manualmente'}), 500


@clients_bp.route('/api/dni-lookup')
@login_required
@require_role('Master', 'Trader', 'Operador', 'Middle Office')
def dni_lookup():
    """
    Consulta información de un DNI:
    1. Verifica si ya existe en la base de datos (evita duplicados).
    2. Si no existe, consulta decolecta.com (RENIEC) para obtener nombres y apellidos.
       Fallback: apis.net.pe v1 libre.
    """
    import urllib.request
    import urllib.error
    from app.models.client import Client
    import json as _json
    import os

    dni = request.args.get('dni', '').strip()

    # ── Validación básica ────────────────────────────────────────────────────
    if not dni:
        return jsonify({'success': False, 'message': 'DNI requerido'}), 400

    if not dni.isdigit() or len(dni) != 8:
        return jsonify({'success': False, 'message': 'El DNI debe tener exactamente 8 dígitos numéricos'}), 400

    # ── Verificar si ya existe en la base de datos ───────────────────────────
    existing = Client.query.filter_by(dni=dni, document_type='DNI').first()
    if existing:
        return jsonify({
            'success':      False,
            'already_exists': True,
            'message':      f'Este DNI ya está registrado: {existing.full_name or "(sin nombre)"}',
            'client_id':    existing.id,
        }), 200

    # ── Consulta externa ──────────────────────────────────────────────────────
    token = (
        os.environ.get('APIS_NET_PE_TOKEN') or
        current_app.config.get('APIS_NET_PE_TOKEN', '')
    ).strip()

    try:
        if token:
            url  = f'https://api.decolecta.com/v1/reniec/dni?numero={dni}'
            hdrs = {
                'Accept':        'application/json',
                'User-Agent':    'QoriCash/2.0',
                'Authorization': f'Bearer {token}',
            }
        else:
            url  = f'https://api.apis.net.pe/v1/dni?numero={dni}'
            hdrs = {'Accept': 'application/json', 'User-Agent': 'QoriCash/2.0'}

        req = urllib.request.Request(url, headers=hdrs)
        with urllib.request.urlopen(req, timeout=8) as resp:
            data = _json.loads(resp.read().decode())

        nombres = (data.get('nombres') or data.get('nombre') or '').strip().upper()
        ap_pat  = (data.get('apellidoPaterno') or data.get('apellido_paterno') or '').strip().upper()
        ap_mat  = (data.get('apellidoMaterno') or data.get('apellido_materno') or '').strip().upper()

        if not nombres and not ap_pat:
            # Fallback a apis.net.pe v1 si decolecta no retorna datos
            if token:
                url2  = f'https://api.apis.net.pe/v1/dni?numero={dni}'
                hdrs2 = {'Accept': 'application/json', 'User-Agent': 'QoriCash/2.0'}
                req2  = urllib.request.Request(url2, headers=hdrs2)
                with urllib.request.urlopen(req2, timeout=8) as resp2:
                    data = _json.loads(resp2.read().decode())
                nombres = (data.get('nombres') or data.get('nombre') or '').strip().upper()
                ap_pat  = (data.get('apellido_paterno') or data.get('apellidoPaterno') or '').strip().upper()
                ap_mat  = (data.get('apellido_materno') or data.get('apellidoMaterno') or '').strip().upper()

        if not nombres and not ap_pat:
            return jsonify({'success': False, 'message': 'DNI no encontrado en RENIEC'}), 404

        return jsonify({
            'success':          True,
            'dni':              dni,
            'nombres':          nombres,
            'apellido_paterno': ap_pat,
            'apellido_materno': ap_mat,
        })

    except urllib.error.HTTPError as e:
        if e.code == 404:
            return jsonify({'success': False, 'message': 'DNI no encontrado en RENIEC'}), 404
        logger.error(f'[DNI lookup] HTTP {e.code} para DNI {dni}')
        return jsonify({'success': False, 'message': 'Error al consultar RENIEC — intente nuevamente'}), 502
    except urllib.error.URLError:
        return jsonify({'success': False, 'message': 'No se pudo conectar al servicio — verifique la conexión'}), 503
    except Exception as e:
        logger.error(f'[DNI lookup] Error inesperado: {e}')
        return jsonify({'success': False, 'message': 'Error interno — ingrese los datos manualmente'}), 500


@clients_bp.route('/api/<int:client_id>/export_history')
@login_required
def export_client_history(client_id):
    """
    Exportar historial de operaciones de un cliente específico a Excel.
    Mismo formato que /operations/api/export_history, filtrado por client_id.

    Query params opcionales:
        start_date: Fecha inicio (formato: YYYY-MM-DD)
        end_date:   Fecha fin   (formato: YYYY-MM-DD)
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    from app.models.client import Client
    from app.models.operation import Operation
    from app.routes.operations import get_bank_account_info

    client = Client.query.get_or_404(client_id)

    start_date = request.args.get('start_date')
    end_date   = request.args.get('end_date')

    query = Operation.query.filter_by(client_id=client_id)

    if start_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            query = query.filter(Operation.created_at >= start)
        except ValueError:
            pass

    if end_date:
        try:
            end = datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            query = query.filter(Operation.created_at <= end)
        except ValueError:
            pass

    operations = query.order_by(Operation.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial de Operaciones"

    header_fill      = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font      = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    if current_user.role in ['Master', 'Operador']:
        headers = ['ID OP.', 'DOCUMENTO', 'CLIENTE', 'USD', 'T.C.', 'PEN',
                   'CUENTA CARGO', 'CUENTA DESTINO', 'CANAL', 'ESTADO', 'FECHA', 'USUARIO']
    else:
        headers = ['ID OP.', 'DOCUMENTO', 'CLIENTE', 'USD', 'T.C.', 'PEN',
                   'CUENTA CARGO', 'CUENTA DESTINO', 'CANAL', 'ESTADO', 'FECHA']

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row_num, op in enumerate(operations, 2):
        ws.cell(row=row_num, column=1, value=op.operation_id)
        ws.cell(row=row_num, column=2, value=op.client.dni if op.client else '-')
        ws.cell(row=row_num, column=3, value=op.client.full_name if op.client else '-')
        ws.cell(row=row_num, column=4, value=float(op.amount_usd))
        ws.cell(row=row_num, column=5, value=float(op.exchange_rate))
        ws.cell(row=row_num, column=6, value=float(op.amount_pen))
        ws.cell(row=row_num, column=7, value=get_bank_account_info(op, op.source_account))
        ws.cell(row=row_num, column=8, value=get_bank_account_info(op, op.destination_account))
        ws.cell(row=row_num, column=9, value='Web' if op.origen == 'plataforma' else 'Sistema')
        ws.cell(row=row_num, column=10, value=op.status)
        ws.cell(row=row_num, column=11, value=op.created_at.strftime('%d/%m/%Y %H:%M') if op.created_at else '-')
        if current_user.role in ['Master', 'Operador']:
            ws.cell(row=row_num, column=12, value=op.user.email if op.user else '-')

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 35
    ws.column_dimensions['H'].width = 35
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 18
    if current_user.role in ['Master', 'Operador']:
        ws.column_dimensions['L'].width = 30

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    filename = f"historial_{client.dni}_{now_peru().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )




@clients_bp.route('/api/repair/assign-app-canal', methods=['POST'])
@login_required
@require_role('Master')
def repair_assign_app_canal():
    """
    Ruta de mantenimiento (solo Master).
    Repara registration_canal en todos los clientes que no lo tienen.
    - created_by NULL → 'app'
    - creator email = web@qoricash.pe → 'web'
    - creator email = app@qoricash.pe → 'app'
    """
    from app.models.client import Client
    from app.models.user import User
    from sqlalchemy import text

    updated = []

    # Clientes sin registration_canal
    clients = Client.query.filter(Client.registration_canal == None).all()

    web_emails = {'web@qoricash.pe'}
    app_emails = {'app@qoricash.pe'}

    for c in clients:
        if c.created_by is None:
            c.registration_canal = 'app'
        elif c.creator and c.creator.email in web_emails:
            c.registration_canal = 'web'
        elif c.creator and c.creator.email in app_emails:
            c.registration_canal = 'app'
        else:
            c.registration_canal = 'system'
        updated.append({'id': c.id, 'dni': c.dni, 'canal': c.registration_canal})

    if updated:
        db.session.commit()

    return jsonify({
        'success': True,
        'message': f'{len(updated)} clientes actualizados con registration_canal',
        'updated_clients': updated
    })
