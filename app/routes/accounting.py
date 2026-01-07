"""
Rutas de Contabilidad para QoriCash Trading V2
"""
from flask import Blueprint, render_template, request, jsonify, send_file
from flask_login import login_required, current_user
from app.models import Operation, AccountingMatch, AccountingBatch
from app.services.accounting_service import AccountingService
from app.extensions import db
from app.utils.decorators import require_role
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

accounting_bp = Blueprint('accounting', __name__)


@accounting_bp.route('/')
@accounting_bp.route('/dashboard')
@login_required
@require_role('Master')
def dashboard():
    """Página principal de contabilidad"""
    return render_template('accounting/dashboard.html', user=current_user)


@accounting_bp.route('/api/available_operations')
@login_required
@require_role('Master')
def get_available_operations():
    """API: Obtener operaciones disponibles para amarrar"""
    try:
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        operation_type = request.args.get('operation_type')

        operations = AccountingService.get_available_operations(
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            operation_type=operation_type
        )

        results = []
        for op in operations:
            available = float(AccountingService.get_available_amount_for_operation(op.id))
            if available > 0:  # Solo mostrar operaciones con monto disponible
                results.append({
                    'id': op.id,
                    'operation_id': op.operation_id,
                    'operation_type': op.operation_type,
                    'amount_usd': float(op.amount_usd),
                    'exchange_rate': float(op.exchange_rate),
                    'amount_pen': float(op.amount_pen),
                    'client_name': op.client.full_name if op.client else 'N/A',
                    'completed_at': op.completed_at.isoformat() if op.completed_at else None,
                    'available_usd': available,
                    'matched_usd': float(AccountingService.get_matched_amount_for_operation(op.id))
                })

        return jsonify({'success': True, 'operations': results})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@accounting_bp.route('/api/create_match', methods=['POST'])
@login_required
@require_role('Master')
def create_match():
    """API: Crear un match entre compra y venta"""
    try:
        data = request.get_json()

        if not data:
            return jsonify({'success': False, 'error': 'No se recibieron datos'}), 400

        buy_op_id = data.get('buy_operation_id')
        sell_op_id = data.get('sell_operation_id')
        amount_usd = data.get('matched_amount_usd')
        notes = data.get('notes', '')

        # Validaciones básicas
        if not buy_op_id or not sell_op_id:
            return jsonify({'success': False, 'error': 'Debe seleccionar ambas operaciones'}), 400

        if not amount_usd or float(amount_usd) <= 0:
            return jsonify({'success': False, 'error': 'El monto debe ser mayor a cero'}), 400

        success, message, match = AccountingService.create_match(
            buy_operation_id=buy_op_id,
            sell_operation_id=sell_op_id,
            matched_amount_usd=amount_usd,
            user_id=current_user.id,
            notes=notes
        )

        if success:
            return jsonify({'success': True, 'message': message, 'match': match.to_dict()})
        else:
            return jsonify({'success': False, 'error': message}), 400

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'Error interno: {str(e)}'}), 500


@accounting_bp.route('/api/delete_match/<int:match_id>', methods=['DELETE'])
@login_required
@require_role('Master')
def delete_match(match_id):
    """API: Eliminar un match"""
    try:
        success, message = AccountingService.delete_match(match_id, current_user.id)

        if success:
            return jsonify({'success': True, 'message': message})
        else:
            return jsonify({'success': False, 'error': message}), 400

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@accounting_bp.route('/api/matches')
@login_required
@require_role('Master')
def get_matches():
    """API: Obtener todos los matches"""
    try:
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        batch_id = request.args.get('batch_id')
        status = request.args.get('status', 'Activo')

        matches = AccountingService.get_all_matches(
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            batch_id=batch_id,
            status=status
        )

        return jsonify({
            'success': True,
            'matches': [m.to_dict() for m in matches]
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@accounting_bp.route('/api/create_batch', methods=['POST'])
@login_required
@require_role('Master')
def create_batch():
    """API: Crear un batch de neteo"""
    try:
        data = request.get_json()
        match_ids = data.get('match_ids', [])
        description = data.get('description', '')
        netting_date = data.get('netting_date')

        if not netting_date:
            netting_date = datetime.now().strftime('%Y-%m-%d')

        success, message, batch = AccountingService.create_batch(
            match_ids=match_ids,
            description=description,
            netting_date=netting_date,
            user_id=current_user.id
        )

        if success:
            return jsonify({'success': True, 'message': message, 'batch': batch.to_dict(include_matches=True)})
        else:
            return jsonify({'success': False, 'error': message}), 400

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@accounting_bp.route('/api/batches')
@login_required
@require_role('Master')
def get_batches():
    """API: Obtener todos los batches"""
    try:
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        status = request.args.get('status')

        batches = AccountingService.get_all_batches(
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            status=status
        )

        return jsonify({
            'success': True,
            'batches': [b.to_dict() for b in batches]
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@accounting_bp.route('/api/batch/<int:batch_id>')
@login_required
@require_role('Master')
def get_batch(batch_id):
    """API: Obtener detalles de un batch"""
    try:
        batch = AccountingBatch.query.get(batch_id)
        if not batch:
            return jsonify({'success': False, 'error': 'Batch no encontrado'}), 404

        return jsonify({
            'success': True,
            'batch': batch.to_dict(include_matches=True)
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@accounting_bp.route('/api/close_batch/<int:batch_id>', methods=['POST'])
@login_required
@require_role('Master')
def close_batch(batch_id):
    """API: Cerrar un batch"""
    try:
        success, message = AccountingService.close_batch(batch_id, current_user.id)

        if success:
            return jsonify({'success': True, 'message': message})
        else:
            return jsonify({'success': False, 'error': message}), 400

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@accounting_bp.route('/api/profit_by_operation')
@login_required
@require_role('Master')
def get_profit_by_operation():
    """API: Obtener utilidad por operación"""
    try:
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')

        profits = AccountingService.get_profit_by_operation(
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin
        )

        return jsonify({'success': True, 'profits': profits})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@accounting_bp.route('/api/profit_by_client')
@login_required
@require_role('Master')
def get_profit_by_client():
    """API: Obtener utilidad por cliente"""
    try:
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')

        profits = AccountingService.get_profit_by_client(
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin
        )

        return jsonify({'success': True, 'profits': profits})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@accounting_bp.route('/export/libro_diario')
@login_required
@require_role('Master')
def export_libro_diario():
    """Exportar libro diario en Excel"""
    try:
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')

        batches = AccountingService.get_all_batches(
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            status='Cerrado'
        )

        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Libro Diario'

        # Estilos
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        # Encabezados
        headers = ['Fecha', 'Batch', 'Cuenta', 'Descripción', 'Debe', 'Haber']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Datos
        row = 2
        for batch in batches:
            for entry in batch.accounting_entry:
                ws.cell(row=row, column=1, value=batch.netting_date.strftime('%d/%m/%Y'))
                ws.cell(row=row, column=2, value=batch.batch_code)
                ws.cell(row=row, column=3, value=entry.get('cuenta', ''))
                ws.cell(row=row, column=4, value=entry.get('glosa', ''))
                ws.cell(row=row, column=5, value=entry.get('debe', 0))
                ws.cell(row=row, column=6, value=entry.get('haber', 0))
                row += 1

        # Ajustar anchos
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 60
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15

        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'libro_diario_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# Continúa en el siguiente mensaje debido a limitaciones de longitud...
