"""
Modulo de Prospeccion — QoriCash Trading V2
Rutas para Master y Trader.
"""
import os, json, base64, threading
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from flask import Blueprint, render_template, request, jsonify, flash, redirect, url_for, current_app
from flask_login import login_required, current_user
from sqlalchemy import or_, func
from app.extensions import db, csrf
from app.models.prospecto import Prospecto, AsignacionProspecto, ActividadProspecto, SeguimientoProspecto, ProspectoEmail
from app.models.user import User
from app.utils.decorators import require_role
from app.utils.formatters import now_peru

# Credenciales OAuth2 compartidas — se leen de variables de entorno de Render
# GMAIL_CLIENT_ID, GMAIL_CLIENT_SECRET
_GMAIL_TOKEN_URI = "https://oauth2.googleapis.com/token"
_GMAIL_SCOPES    = ["https://mail.google.com/"]

# Mapeo email → variable de entorno que contiene SOLO el refresh_token (string simple)
_GMAIL_REFRESH_ENV = {
    "ggarcia@qoricash.pe":  "GMAIL_REFRESH_TOKEN_GGARCIA",
    "gerencia@qoricash.pe": "GMAIL_REFRESH_TOKEN_GERENCIA",
    "info@qoricash.pe":     "GMAIL_REFRESH_TOKEN_INFO",       # FIX 3
    "luacosta@qoricash.pe": "GMAIL_REFRESH_TOKEN_LUACOSTA",
}

# Mapeo email → nombre completo para el cuerpo del correo
_TRADER_NOMBRES = {
    "ggarcia@qoricash.pe":  "Gian Pierre García",
    "gerencia@qoricash.pe": "QoriCash",
    "info@qoricash.pe":     "QoriCash",
    "luacosta@qoricash.pe": "Luciana Acosta",
}

# Emails que usan cargo "Presidente de Negocios"
_EMAILS_PRESIDENTE = {"ggarcia@qoricash.pe"}

# Estado de campañas masivas activas: trader_id → dict
_campanas: dict = {}
_campanas_lock  = threading.Lock()


def _get_trader_info(sender_email: str, role: str = "Trader"):
    """Retorna (nombre_completo, cargo) para el remitente."""
    es_pres         = (role == "Master" or sender_email in _EMAILS_PRESIDENTE)
    nombre_completo = _TRADER_NOMBRES.get(sender_email, sender_email.split("@")[0])
    cargo           = "Presidente de Negocios" if es_pres else "Trader Fx"
    return nombre_completo, cargo


def _get_gmail_service(sender_email):
    """Construye y retorna el servicio Gmail API para el remitente."""
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build

    env_key       = _GMAIL_REFRESH_ENV.get(sender_email.lower())
    refresh_token = os.environ.get(env_key, "").strip() if env_key else ""
    if not refresh_token:
        raise ValueError(f"Refresh token no configurado para {sender_email} (var: {env_key})")

    client_id     = os.environ.get("GMAIL_CLIENT_ID", "")
    client_secret = os.environ.get("GMAIL_CLIENT_SECRET", "")
    if not client_id or not client_secret:
        raise ValueError("GMAIL_CLIENT_ID o GMAIL_CLIENT_SECRET no configurados en Render")

    creds = Credentials(
        token=None,
        refresh_token=refresh_token,
        token_uri=_GMAIL_TOKEN_URI,
        client_id=client_id,
        client_secret=client_secret,
        scopes=_GMAIL_SCOPES,
    )
    creds.refresh(Request())
    return build("gmail", "v1", credentials=creds)


def _get_firma_gmail(sender_email):
    """Obtiene la firma HTML real de la cuenta Gmail del remitente. Retorna '' si falla."""
    try:
        service = _get_gmail_service(sender_email)
        result  = service.users().settings().sendAs().list(userId="me").execute()
        for alias in result.get("sendAs", []):
            if alias.get("isDefault"):
                return alias.get("signature", "")
        sendas = result.get("sendAs", [])
        if sendas:
            return sendas[0].get("signature", "")
    except Exception:
        pass
    return ""


def _send_via_gmail_api(sender_email, to_email, subject, html_body):
    """Envía un correo usando Gmail API con OAuth2. El email queda en Enviados del remitente."""
    service = _get_gmail_service(sender_email)

    msg = MIMEMultipart("alternative")
    msg["To"]      = to_email
    msg["From"]    = sender_email
    msg["Subject"] = subject
    msg.attach(MIMEText(html_body, "html"))

    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    service.users().messages().send(userId="me", body={"raw": raw}).execute()


def _crear_borrador_gmail(sender_email, to_email, subject, html_body):
    """Crea un borrador HTML en Gmail API. Retorna (draft_id, message_id)."""
    service = _get_gmail_service(sender_email)

    msg = MIMEMultipart("alternative")
    msg["To"]      = to_email
    msg["From"]    = sender_email
    msg["Subject"] = subject
    msg.attach(MIMEText(html_body, "html"))

    raw   = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    draft = service.users().drafts().create(
        userId="me",
        body={"message": {"raw": raw}},
    ).execute()
    draft_id   = draft.get("id", "")
    message_id = draft.get("message", {}).get("id", "")
    return draft_id, message_id

prospeccion_bp = Blueprint("prospeccion", __name__, url_prefix="/prospeccion")

DIAS_VIGENCIA = 45  # días de vigencia de un prospecto asignado a un trader

# Estados canónicos (post-normalización)
_FASE_ESTADOS = {
    "sin_contactar": [],
    "contactado":    ["contactado", "presentado", "seguimiento", "P1"],
    "interesado":    ["interesado", "precio_enviado", "P2"],
    "negociando":    ["negociando", "negociacion", "P3"],
    "clientes":      ["cliente", "P4"],
    "no_contactar":  ["no_contactar"],
    "no_interesado": ["no_interesado"],
    "inactivo":      ["inactivo"],
}

# Mapa legacy → canónico para normalizar en escritura
_ESTADO_NORM = {
    "P1": "contactado", "P2": "interesado", "P3": "negociando", "P4": "cliente",
    "seguimiento": "contactado", "negociacion": "negociando",
    "presentado": "contactado", "precio_enviado": "interesado",
}

ESTADOS_VALIDOS = {
    "sin_contactar", "contactado", "interesado", "negociando",
    "cliente", "no_contactar", "no_interesado", "inactivo",
}

GRUPOS_ORDEN = [
    "PIPELINE ACTIVO", "CLIENTES LFC", "PRIORITARIOS",
    "CALIFICADOS", "POR CONTACTAR", "UNIVERSO CONTACTADOS",
    "SOFT BOUNCE", "NO CONTACTAR",
]


def _base_query():
    """Retorna query base de prospectos."""
    return Prospecto.query


def _log_actividad(prospecto_id, user_id, tipo, descripcion,
                   canal=None, bandeja=None, resultado=None, nuevo_estado=None):
    """Helper: crea una ActividadProspecto y actualiza fecha_ultimo_contacto."""
    act = ActividadProspecto(
        prospecto_id=prospecto_id,
        user_id=user_id,
        tipo=tipo,
        canal=canal or tipo,
        bandeja=bandeja,
        descripcion=descripcion,
        resultado=resultado,
        nuevo_estado=nuevo_estado,
    )
    db.session.add(act)
    # Actualizar fecha_ultimo_contacto en el prospecto
    Prospecto.query.filter_by(id=prospecto_id).update({
        "fecha_ultimo_contacto": now_peru().strftime("%Y-%m-%d %H:%M")
    }, synchronize_session=False)


# ── Dashboard CRM ─────────────────────────────────────────────────────────────

@prospeccion_bp.route("/")
@login_required
@require_role("Master")
def dashboard():
    """Dashboard CRM con KPIs, embudo, timeline y seguimientos pendientes."""
    from datetime import date as _date
    hoy      = now_peru()
    hace7    = hoy - timedelta(days=7)
    hace30   = hoy - timedelta(days=30)
    manana   = hoy + timedelta(days=1)
    en48h    = hoy + timedelta(hours=48)

    total    = Prospecto.query.count()

    def _count_estado(vals):
        if not vals:
            return Prospecto.query.filter(or_(
                Prospecto.estado_comercial == None,
                Prospecto.estado_comercial == "",
                Prospecto.estado_comercial == "sin_contactar",
            )).count()
        return Prospecto.query.filter(Prospecto.estado_comercial.in_(vals)).count()

    kpis = {
        "total":         total,
        "sin_contactar": _count_estado([]),
        "contactado":    _count_estado(_FASE_ESTADOS["contactado"]),
        "interesado":    _count_estado(_FASE_ESTADOS["interesado"]),
        "negociando":    _count_estado(_FASE_ESTADOS["negociando"]),
        "cliente":       _count_estado(_FASE_ESTADOS["clientes"]),
        "no_contactar":  _count_estado(["no_contactar"]),
    }

    # Actividades recientes (últimas 60)
    actividades = (ActividadProspecto.query
                   .order_by(ActividadProspecto.creado_en.desc())
                   .limit(60).all())

    # Prospectos con actividad reciente (últimos 7 días)
    contactados_7d = (db.session.query(func.count(func.distinct(ActividadProspecto.prospecto_id)))
                      .filter(ActividadProspecto.creado_en >= hace7)
                      .scalar() or 0)

    # Actividades por canal (últimos 30 días)
    canal_stats = (db.session.query(ActividadProspecto.canal, func.count(ActividadProspecto.id))
                   .filter(ActividadProspecto.creado_en >= hace30,
                           ActividadProspecto.canal != None)
                   .group_by(ActividadProspecto.canal)
                   .all())
    canales = {c: n for c, n in canal_stats}

    # Seguimientos pendientes y vencidos
    seguimientos_pendientes = (SeguimientoProspecto.query
                                .filter_by(completado=False)
                                .filter(SeguimientoProspecto.fecha_programada <= en48h)
                                .order_by(SeguimientoProspecto.fecha_programada)
                                .limit(20).all())

    seguimientos_vencidos = (SeguimientoProspecto.query
                              .filter_by(completado=False)
                              .filter(SeguimientoProspecto.fecha_programada < hoy)
                              .count())

    # Emails enviados (últimos 7 días)
    emails_7d = (ActividadProspecto.query
                 .filter(ActividadProspecto.canal == "email",
                         ActividadProspecto.creado_en >= hace7)
                 .count())

    # WA enviados (últimos 7 días)
    wa_7d = (ActividadProspecto.query
             .filter(ActividadProspecto.canal == "whatsapp",
                     ActividadProspecto.creado_en >= hace7)
             .count())

    return render_template(
        "prospeccion/dashboard.html",
        kpis=kpis,
        actividades=actividades,
        contactados_7d=contactados_7d,
        canales=canales,
        seguimientos_pendientes=seguimientos_pendientes,
        seguimientos_vencidos=seguimientos_vencidos,
        emails_7d=emails_7d,
        wa_7d=wa_7d,
        hoy=hoy,
    )


# ── Reporte por trader (Master) ───────────────────────────────────────────────

@prospeccion_bp.route("/reporte-trader/<int:trader_id>")
@login_required
@require_role("Master")
def reporte_trader(trader_id):
    trader = db.get_or_404(User, trader_id)
    hoy    = now_peru().date()
    hace7  = now_peru() - timedelta(days=7)
    hace30 = now_peru() - timedelta(days=30)

    # Prospectos asignados activos
    prospectos = (db.session.query(Prospecto, AsignacionProspecto)
                  .join(AsignacionProspecto, AsignacionProspecto.prospecto_id == Prospecto.id)
                  .filter(AsignacionProspecto.trader_id == trader_id,
                          AsignacionProspecto.activo == True)
                  .order_by(Prospecto.actualizado_en.desc())
                  .all())

    # Pipeline counts
    estados = {"seguim": 0, "negoc": 0, "cliente": 0, "sin": 0}
    for p, _ in prospectos:
        ec = p.estado_comercial or ""
        if ec in ("P4", "cliente"):
            estados["cliente"] += 1
        elif ec in ("P3", "negociacion"):
            estados["negoc"] += 1
        elif ec in ("P1", "P2", "seguimiento"):
            estados["seguim"] += 1
        else:
            estados["sin"] += 1

    total_asig = len(prospectos)
    tasa_conv  = round(estados["cliente"] / total_asig * 100, 1) if total_asig else 0

    # Actividades últimos 30 días agrupadas por día
    act_rows = (db.session.query(
                    db.func.date(ActividadProspecto.creado_en).label("dia"),
                    db.func.count(ActividadProspecto.id).label("cnt"))
                .filter(ActividadProspecto.user_id == trader_id,
                        ActividadProspecto.creado_en >= hace30)
                .group_by(db.func.date(ActividadProspecto.creado_en))
                .order_by(db.func.date(ActividadProspecto.creado_en))
                .all())
    act_chart = {str(r.dia): r.cnt for r in act_rows}

    # Todas las actividades recientes (últimas 50)
    actividades_recientes = (ActividadProspecto.query
                             .filter_by(user_id=trader_id)
                             .order_by(ActividadProspecto.creado_en.desc())
                             .limit(50).all())

    act7d  = sum(1 for r in act_rows if r.dia >= hace7.date())
    act30d = sum(r.cnt for r in act_rows)

    # Vigencia por prospecto
    vigencia_map = {}
    for p, asig in prospectos:
        vt = DIAS_VIGENCIA + (asig.dias_extra or 0)
        dd = (hoy - asig.asignado_en.date()).days
        vigencia_map[p.id] = vt - dd

    # Prospectos sin actividad en últimos 7 días
    pids_con_act = set(
        r.prospecto_id for r in
        db.session.query(ActividadProspecto.prospecto_id)
        .filter(ActividadProspecto.user_id == trader_id,
                ActividadProspecto.creado_en >= hace7)
        .distinct().all()
    )
    inactivos = [(p, asig) for p, asig in prospectos
                 if p.id not in pids_con_act
                 and (p.estado_comercial or "") not in ("cliente", "P4")]

    return render_template(
        "prospeccion/reporte_trader.html",
        trader=trader,
        prospectos=prospectos,
        estados=estados,
        total_asig=total_asig,
        tasa_conv=tasa_conv,
        act7d=act7d,
        act30d=act30d,
        act_chart=act_chart,
        actividades_recientes=actividades_recientes,
        vigencia_map=vigencia_map,
        inactivos=inactivos,
        dias_vigencia=DIAS_VIGENCIA,
        hoy=hoy,
    )


# ── Lista de prospectos ───────────────────────────────────────────────────────

@prospeccion_bp.route("/lista")
@login_required
@require_role("Master", "Trader")
def lista():
    """Vista grid CRM — los datos se cargan vía /api/grid."""
    return render_template("prospeccion/lista.html",
                           is_master=(current_user.role in ('Master', 'Presidente de Negocios')))


# ── Detalle de prospecto ──────────────────────────────────────────────────────

@prospeccion_bp.route("/<int:pid>")
@login_required
@require_role("Master", "Trader")
def detalle(pid):
    p = db.get_or_404(Prospecto, pid)
    actividades = ActividadProspecto.query.filter_by(prospecto_id=p.id).order_by(ActividadProspecto.creado_en.desc()).limit(50).all()

    return render_template(
        "prospeccion/detalle.html",
        p=p, actividades=actividades,
    )


# ── Registrar actividad ───────────────────────────────────────────────────────

@prospeccion_bp.route("/<int:pid>/actividad", methods=["POST"])
@login_required
@require_role("Master", "Trader")
def agregar_actividad(pid):
    p = db.get_or_404(Prospecto, pid)
    _verificar_acceso(p)

    is_json     = request.is_json or request.headers.get("X-Requested-With") == "XMLHttpRequest"
    data        = request.get_json(silent=True) or request.form
    tipo        = data.get("tipo", "nota")
    descripcion = (data.get("descripcion") or "").strip()
    resultado   = (data.get("resultado") or "").strip()
    nuevo_est   = (data.get("nuevo_estado") or "").strip()
    fecha_prox  = (data.get("fecha_proximo_contacto") or "").strip()

    if not descripcion:
        if is_json:
            return jsonify({"ok": False, "error": "La descripción es requerida."}), 400
        flash("La descripcion es requerida.", "warning")
        return redirect(url_for("prospeccion.detalle", pid=pid))

    canal_map = {"email": "email", "whatsapp": "whatsapp", "llamada": "llamada",
                 "reunion": "reunion", "nota": "manual", "estado": "sistema"}
    canal = canal_map.get(tipo, "manual")

    _log_actividad(p.id, current_user.id, tipo, descripcion,
                   canal=canal, resultado=resultado or None,
                   nuevo_estado=nuevo_est or None)

    if nuevo_est:
        p.estado_comercial = _ESTADO_NORM.get(nuevo_est, nuevo_est)
    if fecha_prox:
        p.fecha_proximo_contacto = fecha_prox

    db.session.commit()

    if is_json:
        return jsonify({"ok": True})
    flash("Actividad registrada correctamente.", "success")
    return redirect(url_for("prospeccion.detalle", pid=pid))


# ── Editar prospecto ──────────────────────────────────────────────────────────

@prospeccion_bp.route("/<int:pid>/editar", methods=["GET", "POST"])
@login_required
@require_role("Master")
def editar(pid):
    p = db.get_or_404(Prospecto, pid)
    _verificar_acceso(p)

    if request.method == "POST":
        campos = [
            "razon_social", "ruc", "tipo", "rubro", "departamento", "provincia",
            "nombre_contacto", "cargo", "email", "email_alt", "telefono",
            "cliente_lfc", "canal", "fuente", "remitente",
            "fecha_proximo_contacto", "estado_comercial", "nivel_interes",
            "grupo", "notas",
        ]
        for campo in campos:
            val = request.form.get(campo)
            if val is not None:
                setattr(p, campo, val.strip() or None)

        db.session.commit()
        flash("Registro actualizado.", "success")
        return redirect(url_for("prospeccion.detalle", pid=pid))

    return render_template("prospeccion/editar.html", p=p, grupos=GRUPOS_ORDEN)


# ── Asignar trader (solo Master) ──────────────────────────────────────────────

@prospeccion_bp.route("/<int:pid>/asignar", methods=["POST"])
@login_required
@require_role("Master")
def asignar(pid):
    p          = db.get_or_404(Prospecto, pid)
    trader_id  = request.form.get("trader_id", type=int)

    if not trader_id:
        # Quitar asignacion
        AsignacionProspecto.query.filter_by(prospecto_id=p.id, activo=True).update({"activo": False})
        db.session.commit()
        flash("Asignacion eliminada.", "info")
        return redirect(url_for("prospeccion.detalle", pid=pid))

    trader = db.get_or_404(User, trader_id)

    # Verificar límite de 500 prospectos activos por trader
    active_count = AsignacionProspecto.query.filter_by(trader_id=trader_id, activo=True).count()
    if active_count >= 500:
        if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.is_json:
            return jsonify({"ok": False, "error": f"{trader.username} ya tiene 500 prospectos asignados (máximo permitido). Libera prospectos antes de asignar nuevos."}), 400
        flash(f"{trader.username} ya tiene 500 prospectos asignados (máximo permitido). Libera prospectos antes de asignar nuevos.", "warning")
        return redirect(request.referrer or url_for("prospeccion.detalle", pid=pid))

    # Desactivar asignacion anterior
    AsignacionProspecto.query.filter_by(prospecto_id=p.id, activo=True).update({"activo": False})

    # Buscar si ya existe (inactiva)
    existente = AsignacionProspecto.query.filter_by(
        prospecto_id=pid, trader_id=trader_id
    ).first()
    if existente:
        existente.activo       = True
        existente.asignado_por = current_user.id
        existente.asignado_en  = now_peru()
    else:
        nueva = AsignacionProspecto(
            prospecto_id=pid,
            trader_id=trader_id,
            asignado_por=current_user.id,
        )
        db.session.add(nueva)

    db.session.commit()
    # Responder JSON si viene del drawer (fetch), redirect si viene de form HTML
    if request.headers.get("X-Requested-With") == "XMLHttpRequest" or "application/json" in request.headers.get("Accept",""):
        return jsonify({"ok": True, "trader": trader.username})
    flash(f"Prospecto asignado a {trader.username}.", "success")
    return redirect(url_for("prospeccion.detalle", pid=pid))


# ── Asignacion masiva JSON (solo Master) ──────────────────────────────────────

@prospeccion_bp.route("/api/asignar-masivo", methods=["POST"])
@csrf.exempt
@login_required
@require_role("Master")
def api_asignar_masivo():
    """Asigna N prospectos sin trader al trader seleccionado."""
    data      = request.get_json(force=True) or {}
    trader_id = data.get("trader_id")
    cantidad  = min(int(data.get("cantidad") or 50), 500)
    estado    = (data.get("estado") or "").strip()

    if not trader_id:
        return jsonify({"ok": False, "error": "Selecciona un trader."}), 400

    trader = db.session.get(User, trader_id)
    if not trader:
        return jsonify({"ok": False, "error": "Trader no encontrado."}), 400

    active_count = AsignacionProspecto.query.filter_by(trader_id=trader_id, activo=True).count()
    if active_count >= 500:
        return jsonify({"ok": False, "error": f"{trader.username} ya tiene 500 prospectos asignados."}), 400
    cantidad = min(cantidad, 500 - active_count)

    ya_asignados = db.session.query(AsignacionProspecto.prospecto_id).filter_by(activo=True)
    q = Prospecto.query.filter(~Prospecto.id.in_(ya_asignados))
    if estado:
        _FASE_MAP = {
            "sin_contactar":  [None, "", "sin_contactar"],
            "presentado":     ["presentado", "seguimiento", "P1"],
            "precio_enviado": ["precio_enviado", "P2"],
            "negociando":     ["negociando", "negociacion", "P3"],
        }
        vals = _FASE_MAP.get(estado, [estado])
        from sqlalchemy import or_ as _or
        q = q.filter(_or(*[Prospecto.estado_comercial == v for v in vals]))
    prospectos = q.limit(cantidad).all()

    asignados = 0
    for p in prospectos:
        existente = AsignacionProspecto.query.filter_by(prospecto_id=p.id, trader_id=trader_id).first()
        if existente:
            existente.activo = True
            existente.asignado_por = current_user.id
            existente.asignado_en  = now_peru()
        else:
            db.session.add(AsignacionProspecto(
                prospecto_id=p.id, trader_id=trader_id, asignado_por=current_user.id
            ))
        asignados += 1

    db.session.commit()
    return jsonify({"ok": True, "asignados": asignados, "trader": trader.username})


# ── Asignacion masiva por grupo (solo Master) ──────────────────────────────────────────

@prospeccion_bp.route("/asignar-masivo", methods=["GET", "POST"])
@login_required
@require_role("Master")
def asignar_masivo():
    traders = User.query.filter_by(role="Trader", status="Activo").all()

    if request.method == "POST":
        trader_id  = request.form.get("trader_id", type=int)
        grupo      = request.form.get("grupo", "")
        cantidad   = request.form.get("cantidad", 50, type=int)

        if not trader_id or not grupo:
            flash("Selecciona trader y segmento.", "warning")
            return redirect(url_for("prospeccion.asignar_masivo"))

        # Verificar límite de 500 prospectos activos por trader
        active_count = AsignacionProspecto.query.filter_by(trader_id=trader_id, activo=True).count()
        if active_count >= 500:
            trader_obj = db.session.get(User, trader_id)
            nombre_t   = trader_obj.username if trader_obj else str(trader_id)
            flash(f"{nombre_t} ya tiene 500 prospectos asignados (máximo permitido). Libera prospectos antes de asignar nuevos.", "warning")
            return redirect(url_for("prospeccion.asignar_masivo"))
        cantidad = min(cantidad, 500 - active_count)

        # Prospectos sin asignacion activa del grupo seleccionado
        ya_asignados = db.session.query(AsignacionProspecto.prospecto_id).filter_by(activo=True)
        sin_asignar  = (Prospecto.query
                        .filter(Prospecto.grupo == grupo)
                        .filter(~Prospecto.id.in_(ya_asignados))
                        .filter(Prospecto.grupo != "NO CONTACTAR")
                        .order_by(Prospecto.score.desc())
                        .limit(cantidad).all())

        asignados = 0
        for prosp in sin_asignar:
            nueva = AsignacionProspecto(
                prospecto_id=prosp.id,
                trader_id=trader_id,
                asignado_por=current_user.id,
            )
            db.session.add(nueva)
            asignados += 1

        db.session.commit()
        trader = db.session.get(User, trader_id)
        flash(f"{asignados} prospectos asignados a {trader.username if trader else trader_id}.", "success")
        return redirect(url_for("prospeccion.asignar_masivo"))

    # Conteo por grupo sin asignar
    ya_asignados = db.session.query(AsignacionProspecto.prospecto_id).filter_by(activo=True)
    grupos_count = (db.session.query(Prospecto.grupo, func.count(Prospecto.id))
                    .filter(~Prospecto.id.in_(ya_asignados))
                    .filter(Prospecto.grupo != "NO CONTACTAR")
                    .group_by(Prospecto.grupo)
                    .order_by(func.count(Prospecto.id).desc()).all())

    return render_template("prospeccion/asignar_masivo.html",
                           traders=traders, grupos_count=grupos_count)


# ── Pipeline kanban ───────────────────────────────────────────────────────────

@prospeccion_bp.route("/pipeline")
@login_required
def pipeline():
    return redirect(url_for("prospeccion.lista"))


@prospeccion_bp.route("/<int:pid>/mover", methods=["POST"])
@login_required
@require_role("Master")
def mover_pipeline(pid):
    p = db.get_or_404(Prospecto, pid)
    _verificar_acceso(p)
    nuevo = request.form.get("estado")
    validos = {"presentado","precio_enviado","negociando","cliente","P1","P2","P3","P4"}
    if nuevo in validos:
        p.estado_comercial = nuevo
        act = ActividadProspecto(
            prospecto_id=p.id, user_id=current_user.id, tipo="estado",
            descripcion=f"Pipeline movido a {nuevo}.", nuevo_estado=nuevo,
        )
        db.session.add(act)
        db.session.commit()
    return redirect(request.referrer or url_for("prospeccion.lista"))


# ── No contactar ─────────────────────────────────────────────────────────────

@prospeccion_bp.route("/no-contactar")
@login_required
def no_contactar():
    return redirect(url_for("prospeccion.lista"))


# ── API JSON (para charts del dashboard) ─────────────────────────────────────

@prospeccion_bp.route("/api/charts/grupos")
@login_required
@require_role("Master")
def api_grupos():
    q = _base_query()
    rows = (db.session.query(Prospecto.grupo, func.count(Prospecto.id))
            .filter(Prospecto.id.in_(q.with_entities(Prospecto.id)))
            .group_by(Prospecto.grupo).all())
    colores = {
        "PIPELINE ACTIVO": "#8B5CF6", "CLIENTES LFC": "#22C55E",
        "PRIORITARIOS": "#B91C1C",    "CALIFICADOS": "#3B82F6",
        "POR CONTACTAR": "#0891B1",   "UNIVERSO CONTACTADOS": "#F59E0B",
        "SOFT BOUNCE": "#D97706",     "NO CONTACTAR": "#EF4444",
    }
    return jsonify([{"label": r[0] or "Sin grupo", "value": r[1],
                     "color": colores.get(r[0], "#94A3B8")} for r in rows])


@prospeccion_bp.route("/api/charts/rubros")
@login_required
@require_role("Master")
def api_rubros():
    q = _base_query()
    rows = (db.session.query(Prospecto.rubro, func.count(Prospecto.id))
            .filter(Prospecto.id.in_(q.with_entities(Prospecto.id)))
            .filter(Prospecto.rubro.isnot(None))
            .group_by(Prospecto.rubro)
            .order_by(func.count(Prospecto.id).desc())
            .limit(10).all())
    return jsonify([{"rubro": r[0], "total": r[1]} for r in rows])


# ── Import masivo via HTTP (endpoint temporal para carga inicial) ─────────────

@prospeccion_bp.route("/api/sincronizar-enviados", methods=["POST"])
@csrf.exempt
def sincronizar_enviados():
    """Marca como 'seguimiento' todos los prospectos cuyo email aparece en la lista."""
    import os
    key = request.headers.get("X-Import-Key", "")
    from app.utils.security import safe_compare
    expected_key = os.environ.get("PROSPECCION_IMPORT_KEY", "")
    if not expected_key or not safe_compare(key, expected_key):
        return jsonify({"error": "No autorizado"}), 401

    emails = request.get_json(force=True).get("emails", [])
    if not emails:
        return jsonify({"actualizados": 0, "no_encontrados": 0})

    emails_lower = [e.strip().lower() for e in emails if e and "@" in e]
    actualizados = 0
    no_encontrados = 0

    for email in emails_lower:
        p = Prospecto.query.filter(
            db.func.lower(Prospecto.email) == email,
            Prospecto.estado_comercial.notin_(["seguimiento","negociacion","cliente","P1","P2","P3","P4"])
        ).first()
        if p:
            p.estado_comercial = "seguimiento"
            actualizados += 1
        else:
            no_encontrados += 1

    db.session.commit()
    return jsonify({"actualizados": actualizados, "no_encontrados": no_encontrados})


@prospeccion_bp.route("/api/registrar-envios", methods=["POST"])
@csrf.exempt
def registrar_envios():
    """
    Recibe envios de los scripts de campaña y actualiza la BD en tiempo real.
    Body: {"envios": [{"email": "...", "tipo": "presentacion"|"precio", "remitente": "ggarcia"}]}
    """
    import os
    key = request.headers.get("X-Import-Key", "")
    from app.utils.security import safe_compare
    expected_key = os.environ.get("PROSPECCION_IMPORT_KEY", "")
    if not expected_key or not safe_compare(key, expected_key):
        return jsonify({"error": "No autorizado"}), 401

    data   = request.get_json(force=True)
    envios = data.get("envios", [])
    if not envios:
        return jsonify({"actualizados": 0})

    _NO_DEGRADAR = {"negociando", "negociacion", "P3", "cliente", "P4"}
    ahora        = now_peru().strftime("%Y-%m-%d %H:%M")
    actualizados = 0

    for item in envios:
        email     = (item.get("email") or "").strip().lower()
        tipo      = (item.get("tipo")  or "presentacion").strip()
        remitente = (item.get("remitente") or "ggarcia").strip()

        if not email or "@" not in email:
            continue

        nuevo_estado = "precio_enviado" if tipo == "precio" else "presentado"

        p = Prospecto.query.filter(
            db.func.lower(Prospecto.email) == email
        ).first()
        if not p:
            continue

        # No degradar si ya está en negociacion o cliente
        if (p.estado_comercial or "") not in _NO_DEGRADAR:
            p.estado_comercial = nuevo_estado

        p.tipo_ultimo_envio     = tipo
        p.remitente             = remitente
        p.fecha_ultimo_contacto = ahora
        p.num_contactos         = (p.num_contactos or 0) + 1
        if not p.fecha_primer_contacto:
            p.fecha_primer_contacto = ahora

        actualizados += 1

    db.session.commit()
    return jsonify({"actualizados": actualizados, "total": len(envios)})


@prospeccion_bp.route("/api/registrar-bounces", methods=["POST"])
@csrf.exempt
def registrar_bounces():
    """
    Recibe lista de emails rebotados desde monitor_rebotes.py y actualiza la BD.
    Body: {"emails": ["..."], "tipo": "Hard Bounce"}
    """
    import os
    key = request.headers.get("X-Import-Key", "")
    from app.utils.security import safe_compare
    expected_key = os.environ.get("PROSPECCION_IMPORT_KEY", "")
    if not expected_key or not safe_compare(key, expected_key):
        return jsonify({"error": "No autorizado"}), 401

    data         = request.get_json(force=True)
    emails       = data.get("emails", [])
    tipo_bounce  = data.get("tipo", "Hard Bounce")
    actualizados = 0

    for email in emails:
        email = (email or "").strip().lower().rstrip(".")
        if not email or "@" not in email:
            continue
        p = Prospecto.query.filter(
            db.func.lower(Prospecto.email) == email
        ).first()
        if p:
            p.estado_email = tipo_bounce
            if tipo_bounce == "Hard Bounce":
                p.email = None   # nullificar para limpiar la base
            actualizados += 1

    db.session.commit()
    return jsonify({"actualizados": actualizados, "total": len(emails)})


@prospeccion_bp.route("/api/import-batch", methods=["POST"])
@csrf.exempt
def import_batch():
    """
    Endpoint temporal para cargar los 11K prospectos desde el Mac.
    Protegido por API key en header X-Import-Key.
    """
    import os
    key = request.headers.get("X-Import-Key", "")
    from app.utils.security import safe_compare
    expected_key = os.environ.get("PROSPECCION_IMPORT_KEY", "")
    if not expected_key or not safe_compare(key, expected_key):
        return jsonify({"error": "No autorizado"}), 401

    data = request.get_json(force=True)
    action = data.get("action", "insert")

    if action == "truncate":
        Prospecto.query.delete()
        db.session.commit()
        return jsonify({"ok": True, "msg": "Tabla limpiada"})

    if action == "count":
        return jsonify({"total": Prospecto.query.count()})

    registros = data.get("registros", [])
    if not registros:
        return jsonify({"ok": True, "insertados": 0})

    objs = [Prospecto(**r) for r in registros]
    db.session.bulk_save_objects(objs)
    db.session.commit()
    return jsonify({"ok": True, "insertados": len(objs)})


@prospeccion_bp.route("/api/asignar-por-remitente", methods=["POST"])
@csrf.exempt
def asignar_por_remitente():
    """
    Asigna prospectos a usuarios segun el campo 'remitente'.
    Payload: {"remitente": "ggarcia", "email_usuario": "ggarcia@qoricash.pe"}
    O accion "preview" para ver cuantos se asignarian.
    """
    import os
    key = request.headers.get("X-Import-Key", "")
    from app.utils.security import safe_compare
    expected_key = os.environ.get("PROSPECCION_IMPORT_KEY", "")
    if not expected_key or not safe_compare(key, expected_key):
        return jsonify({"error": "No autorizado"}), 401

    data          = request.get_json(force=True)
    remitente     = data.get("remitente", "").strip().lower()
    email_usuario = data.get("email_usuario", "").strip().lower()
    accion        = data.get("accion", "asignar")

    if not remitente or not email_usuario:
        return jsonify({"error": "remitente y email_usuario son requeridos"}), 400

    # Buscar usuario por email
    usuario = User.query.filter(
        db.func.lower(User.email) == email_usuario
    ).first()
    if not usuario:
        # Intentar por username
        usuario = User.query.filter(
            db.func.lower(User.username) == remitente
        ).first()
    if not usuario:
        usuarios_disponibles = [
            {"id": u.id, "username": u.username, "email": u.email, "role": u.role}
            for u in User.query.filter(User.role.in_(["Master", "Trader"])).all()
        ]
        return jsonify({
            "error": f"No se encontro usuario con email '{email_usuario}'",
            "usuarios_disponibles": usuarios_disponibles
        }), 404

    # Buscar prospectos con ese remitente
    prospectos = Prospecto.query.filter(
        db.func.lower(Prospecto.remitente) == remitente
    ).all()

    if accion == "preview":
        ya_asignados = sum(
            1 for p in prospectos
            if next((a for a in p.asignaciones if a.trader_id == usuario.id and a.activo), None)
        )
        return jsonify({
            "remitente":      remitente,
            "usuario":        {"id": usuario.id, "username": usuario.username, "email": usuario.email, "role": usuario.role},
            "total_encontrados": len(prospectos),
            "ya_asignados":   ya_asignados,
            "por_asignar":    len(prospectos) - ya_asignados,
        })

    # Asignar
    asignados = 0
    ya_tenia  = 0
    ahora     = now_peru()

    for p in prospectos:
        existente = AsignacionProspecto.query.filter_by(
            prospecto_id=p.id, trader_id=usuario.id
        ).first()
        if existente:
            if not existente.activo:
                existente.activo      = True
                existente.asignado_en = ahora
                asignados += 1
            else:
                ya_tenia += 1
        else:
            nueva = AsignacionProspecto(
                prospecto_id=p.id,
                trader_id=usuario.id,
                activo=True,
                asignado_en=ahora,
            )
            db.session.add(nueva)
            asignados += 1

        if asignados % 500 == 0:
            db.session.flush()

    db.session.commit()
    return jsonify({
        "ok":        True,
        "remitente": remitente,
        "usuario":   {"id": usuario.id, "username": usuario.username, "role": usuario.role},
        "asignados": asignados,
        "ya_tenian": ya_tenia,
        "total":     len(prospectos),
    })


# ── Enviar correo a prospecto ─────────────────────────────────────────────────

LOGO_URL = "https://www.qoricash.pe/logofirma.png"

# ── Componentes del email de precios (modelo gerencia@qoricash.pe) ────────────

HEADER_HTML = f"""\
<tr>
  <td style="background:#0D1B2A;padding:18px 28px;">
    <table width="100%" cellpadding="0" cellspacing="0" border="0"><tr>
      <td>
        <img src="{LOGO_URL}" alt="QoriCash" height="32" style="display:block;height:32px;">
      </td>
      <td style="padding-left:6px;vertical-align:middle;">
        <span style="font-size:14px;font-weight:800;color:#FFFFFF;letter-spacing:2px;">QORICASH</span>
      </td>
      <td align="right">
        <span style="font-size:10px;font-weight:700;color:#5CB85C;text-transform:uppercase;
                     letter-spacing:1.5px;background:rgba(92,184,92,.12);
                     padding:4px 10px;border-radius:20px;border:1px solid rgba(92,184,92,.3);">
          {{fecha}}
        </span>
      </td>
    </tr></table>
  </td>
</tr>
<tr>
  <td style="background:#F8FAFC;padding:20px 28px;border-bottom:1px solid #E9EEF4;">
    <p style="margin:0;font-size:13px;color:#475569;line-height:1.7;">
      Estimado(a) <strong>{{nombre}}</strong>, a continuaci&oacute;n las tasas del tipo de cambio en estos momentos.
      Rentabilice sus operaciones con las mejores tasas del tipo de cambio del Per&uacute;.
    </p>
  </td>
</tr>"""

BANCOS_HTML = """\
<tr>
  <td style="padding:24px 28px 8px;">
    <p style="margin:0 0 14px;font-size:10px;font-weight:900;color:#94A3B8;text-transform:uppercase;letter-spacing:1.4px;">
      <strong>Operamos con los principales bancos del Per&uacute;</strong>
    </p>
    <table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-bottom:16px;">
      <tr><td style="border-left:3px solid #5CB85C;padding-left:10px;">
        <p style="margin:0;font-size:10px;font-weight:700;color:#0D1B2A;">QORICASH S.A.C.</p>
        <p style="margin:2px 0 0;font-size:9px;color:#94A3B8;">RUC 20615113698 &nbsp;&middot;&nbsp; Regulada por la SBS</p>
      </td></tr>
    </table>
    <table width="100%" cellpadding="0" cellspacing="0" border="0" style="border:1px solid #E9EEF4;border-radius:10px;overflow:hidden;margin-bottom:16px;">
      <tr style="border-bottom:1px solid #F1F5F9;">
        <td style="width:90px;padding:12px 8px 12px 18px;vertical-align:middle;">
          <span style="font-size:15px;font-weight:800;color:#F97316;">BCP</span>
        </td>
        <td style="padding:14px 12px;vertical-align:middle;border-left:1px solid #F1F5F9;">
          <p style="margin:0;font-size:9px;color:#94A3B8;text-transform:uppercase;letter-spacing:0.8px;">Soles</p>
          <p style="margin:2px 0 0;font-size:12px;font-weight:700;color:#1E293B;letter-spacing:0.5px;">1937353150041</p>
        </td>
        <td style="padding:14px 18px 14px 12px;vertical-align:middle;border-left:1px solid #F1F5F9;">
          <p style="margin:0;font-size:9px;color:#94A3B8;text-transform:uppercase;letter-spacing:0.8px;">D&oacute;lares</p>
          <p style="margin:2px 0 0;font-size:12px;font-weight:700;color:#1E293B;letter-spacing:0.5px;">1917357790119</p>
        </td>
      </tr>
      <tr style="border-bottom:1px solid #F1F5F9;background:#FAFBFC;">
        <td style="width:90px;padding:12px 8px 12px 18px;vertical-align:middle;">
          <span style="font-size:15px;font-weight:800;color:#00A859;">Interbank</span>
        </td>
        <td style="padding:14px 12px;vertical-align:middle;border-left:1px solid #F1F5F9;">
          <p style="margin:0;font-size:9px;color:#94A3B8;text-transform:uppercase;letter-spacing:0.8px;">Soles</p>
          <p style="margin:2px 0 0;font-size:12px;font-weight:700;color:#1E293B;letter-spacing:0.5px;">200-3007757571</p>
        </td>
        <td style="padding:14px 18px 14px 12px;vertical-align:middle;border-left:1px solid #F1F5F9;">
          <p style="margin:0;font-size:9px;color:#94A3B8;text-transform:uppercase;letter-spacing:0.8px;">D&oacute;lares</p>
          <p style="margin:2px 0 0;font-size:12px;font-weight:700;color:#1E293B;letter-spacing:0.5px;">200-3007757589</p>
        </td>
      </tr>
      <tr>
        <td style="width:90px;padding:12px 8px 12px 18px;vertical-align:middle;">
          <span style="font-size:15px;font-weight:800;color:#004B9D;">BanBif</span>
        </td>
        <td style="padding:14px 12px;vertical-align:middle;border-left:1px solid #F1F5F9;">
          <p style="margin:0;font-size:9px;color:#94A3B8;text-transform:uppercase;letter-spacing:0.8px;">Soles</p>
          <p style="margin:2px 0 0;font-size:12px;font-weight:700;color:#1E293B;letter-spacing:0.5px;">007000845805</p>
        </td>
        <td style="padding:14px 18px 14px 12px;vertical-align:middle;border-left:1px solid #F1F5F9;">
          <p style="margin:0;font-size:9px;color:#94A3B8;text-transform:uppercase;letter-spacing:0.8px;">D&oacute;lares</p>
          <p style="margin:2px 0 0;font-size:12px;font-weight:700;color:#1E293B;letter-spacing:0.5px;">007000845813</p>
        </td>
      </tr>
    </table>
    <p style="margin:0 0 24px;font-size:10px;color:#64748B;line-height:1.6;">
      <strong>Transferencia interbancaria (CCI) disponible desde BBVA, Scotiabank, Pichincha y cualquier banco del Per&uacute;.</strong>
    </p>
    <table cellpadding="0" cellspacing="0" border="0" style="margin-bottom:28px;">
      <tr><td style="border-radius:7px;background:#0D1B2A;box-shadow:0 4px 14px rgba(13,27,42,.25);">
        <a href="https://wa.me/51910624404" style="display:inline-block;padding:13px 30px;color:#FFFFFF;text-decoration:none;font-size:12px;font-weight:700;letter-spacing:0.8px;">
          Cotizar en l&iacute;nea &nbsp;&rarr;
        </a>
      </td></tr>
    </table>
  </td>
</tr>"""

FIRMA_HTML = f"""\
<tr>
  <td style="padding:16px 28px;border-top:1px solid #F1F5F9;background:#FAFAFA;">
    <table cellpadding="0" cellspacing="0" border="0"><tr>
      <td style="padding-right:12px;vertical-align:middle;">
        <img src="{LOGO_URL}" width="32" height="32" alt="QoriCash" style="display:block;border-radius:4px;">
      </td>
      <td style="vertical-align:middle;padding-right:24px;">
        <p style="margin:0;font-size:12px;font-weight:700;color:#0D1B2A;">{{trader_nombre}}</p>
        <p style="margin:1px 0 0;font-size:10px;color:#5CB85C;font-weight:600;">{{trader_cargo}}</p>
      </td>
      <td style="width:1px;background:#E2E8F0;padding:0;"></td>
      <td style="width:24px;"></td>
      <td style="vertical-align:middle;">
        <p style="margin:0;font-size:10px;color:#64748B;">
          <a href="https://wa.me/51910624404" style="color:#64748B;text-decoration:none;">+51 910 624 404</a>
          &nbsp;&middot;&nbsp;
          <a href="https://www.qoricash.pe" style="color:#5CB85C;text-decoration:none;font-weight:600;">www.qoricash.pe</a>
        </p>
        <p style="margin:2px 0 0;font-size:9px;color:#94A3B8;">Av. Brasil 2790, int. 504 &mdash; Pueblo Libre</p>
      </td>
    </tr></table>
  </td>
</tr>"""

PIE = """\
<tr>
  <td style="padding:12px 28px;background:#F8FAFC;border-top:1px solid #F1F5F9;">
    <p style="margin:0;font-size:9px;color:#CBD5E1;text-align:center;">
      Regulada por la SBS &nbsp;&middot;&nbsp; Res. N.&ordm; 00313-2026 &nbsp;&middot;&nbsp;
      Precios sujetos a variaci&oacute;n &nbsp;&middot;&nbsp;
      Para no recibir m&aacute;s comunicaciones, responda con asunto <em>NO CONTACTAR</em>.
    </p>
  </td>
</tr>"""

CUERPO_PRESENTACION = """\
<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8"></head>
<body style="font-family:Arial,sans-serif;font-size:14px;color:#1E293B;line-height:1.7;max-width:620px;margin:0 auto;padding:24px;">
  {header}
  <p>Estimado(a) <strong>{nombre}</strong>,</p>
  <p style="text-align:justify;">{presentacion_remitente}</p>
  <p style="text-align:justify;">Trabajamos con empresas que realizan operaciones frecuentes de compra y venta de
    dolares, y que en muchos casos estan dejando dinero sobre la mesa al operar con el tipo de cambio que les ofrece
    su entidad financiera actual.</p>
  <p style="text-align:justify;">Porque cada centavo cuenta, le ofrecemos <strong>tasas que superan
    consistentemente al sistema bancario tradicional</strong>, con ejecucion inmediata y cero costos ocultos.
    Nuestro sistema tecnologico monitorea el mercado en tiempo real, lo que nos permite notificarle cuando es el
    momento ideal para operar.</p>
  <div style="background:#F7F9FC;border-left:4px solid #4CAF50;border-radius:4px;padding:16px 20px;margin:24px 0;">
    <p style="margin:0 0 6px;font-weight:bold;color:#0D1B2A;">Le propongo algo concreto:</p>
    <p style="margin:0;color:#4A5568;text-align:justify;">Una comparativa sin compromiso entre las tasas que recibe
      hoy de su proveedor actual y las que podemos ofrecerle en QoriCash, en tiempo real.</p>
  </div>
  <p style="text-align:justify;">Si desea conocer mas, puede ver nuestra presentacion institucional aqui:</p>
  <div style="margin:16px 0 24px;">
    <a href="https://qoricash.pe/presentacion.pdf"
       style="display:inline-block;padding:10px 24px;background:linear-gradient(135deg,#5CB85C 0%,#4a9b4a 100%);
              color:#ffffff;text-decoration:none;border-radius:6px;font-size:13px;font-weight:700;letter-spacing:0.3px;"
       target="_blank">Ver presentacion QoriCash</a>
  </div>
  <p style="text-align:justify;">Contamos con cuentas corrientes en soles y d&oacute;lares en los bancos
    m&aacute;s importantes del Per&uacute;:</p>
  {bancos}
  <p style="text-align:justify;">Quedamos atentos a su respuesta.</p>
  <br>{firma}{pie}
</body></html>"""

CUERPO_PRECIO = """\
<!DOCTYPE html>
<html lang="es">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#F4F6F8;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#F4F6F8;padding:32px 0;">
<tr><td align="center">
<table width="560" cellpadding="0" cellspacing="0" border="0"
  style="max-width:560px;width:100%;background:#FFFFFF;border-radius:12px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,.07);">
  {header}
  {mensaje}
  {ticker}
  {bancos}
  {firma}
  {pie}
</table>
</td></tr>
</table>
</body></html>"""


def _build_ticker(compra, venta):
    hoy = now_peru().strftime("%d/%m/%Y")
    return f"""\
<tr>
  <td style="padding:20px 28px;">
    <table width="100%" cellpadding="0" cellspacing="0" border="0"
      style="border:1.5px solid #E2E8F0;border-radius:10px;overflow:hidden;">
      <tr>
        <td width="50%" style="padding:24px 28px;border-right:1.5px solid #E2E8F0;text-align:center;">
          <p style="margin:0 0 6px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;letter-spacing:1.8px;">Compramos</p>
          <p style="margin:0;font-size:34px;font-weight:800;color:#0D1B2A;letter-spacing:-1px;line-height:1;">S/. {compra}</p>
          <p style="margin:8px 0 0;font-size:10px;color:#94A3B8;">por d&oacute;lar &middot; USD</p>
        </td>
        <td width="50%" style="padding:24px 28px;text-align:center;">
          <p style="margin:0 0 6px;font-size:9px;font-weight:700;color:#94A3B8;text-transform:uppercase;letter-spacing:1.8px;">Vendemos</p>
          <p style="margin:0;font-size:34px;font-weight:800;color:#16a34a;letter-spacing:-1px;line-height:1;">S/. {venta}</p>
          <p style="margin:8px 0 0;font-size:10px;color:#94A3B8;">por d&oacute;lar &middot; USD</p>
        </td>
      </tr>
      <tr>
        <td colspan="2" style="padding:10px 28px;border-top:1.5px solid #E2E8F0;text-align:center;">
          <span style="font-size:10px;font-weight:600;color:#64748B;letter-spacing:0.3px;">
            &bull;&nbsp; Operaci&oacute;n en minutos &nbsp;&middot;&nbsp; Sin costo de transferencia
          </span><br>
          <span style="font-size:9.5px;color:#94A3B8;font-style:italic;">
            Precios del momento &mdash; sujetos a variaci&oacute;n por volatilidad del mercado.
          </span>
        </td>
      </tr>
    </table>
  </td>
</tr>"""


MENSAJE_SEG_HTML = """\
<tr>
  <td style="padding:28px 28px 8px;">
    <p style="margin:0 0 16px;font-size:14px;color:#1E293B;line-height:1.7;">
      Estimado(a) <strong>{nombre}</strong>,
    </p>
    <p style="margin:0 0 16px;font-size:13px;color:#475569;line-height:1.7;text-align:justify;">
      Me comunico nuevamente desde QoriCash para darle seguimiento a la presentaci&oacute;n que le enviamos
      recientemente. Quiero asegurarme de que haya tenido la oportunidad de revisarla y, sobre todo,
      saber si tiene alguna consulta sobre c&oacute;mo podemos optimizar sus operaciones de cambio de d&oacute;lares.
    </p>
    <div style="background:#F0FDF4;border-left:4px solid #5CB85C;border-radius:4px;padding:16px 20px;margin:20px 0;">
      <p style="margin:0 0 8px;font-weight:700;font-size:13px;color:#0D1B2A;">En QoriCash ofrecemos:</p>
      <p style="margin:0;font-size:13px;color:#4A5568;line-height:1.85;">
        &#10003;&nbsp; Tasas que superan consistentemente al sistema bancario<br>
        &#10003;&nbsp; Operaci&oacute;n 100% digital, en minutos y sin costos ocultos<br>
        &#10003;&nbsp; Atenci&oacute;n personalizada con su ejecutivo asignado
      </p>
    </div>
    <p style="margin:0 0 22px;font-size:13px;color:#475569;line-height:1.7;text-align:justify;">
      &iquest;Le parece si coordinamos una llamada breve esta semana para mostrarle en tiempo real la diferencia
      entre nuestras tasas y las que recibe actualmente? No toma m&aacute;s de 10 minutos y puede representar
      un ahorro significativo para su empresa.
    </p>
    <table cellpadding="0" cellspacing="0" border="0" style="margin-bottom:28px;">
      <tr><td style="border-radius:7px;background:#5CB85C;box-shadow:0 4px 14px rgba(92,184,92,.3);">
        <a href="https://wa.me/51910624404"
           style="display:inline-block;padding:12px 28px;color:#FFFFFF;text-decoration:none;font-size:13px;font-weight:700;letter-spacing:0.5px;">
          Coordinar llamada &nbsp;&rarr;
        </a>
      </td></tr>
    </table>
  </td>
</tr>"""

CUERPO_SEGUIMIENTO = """\
<!DOCTYPE html>
<html lang="es">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#F4F6F8;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#F4F6F8;padding:32px 0;">
<tr><td align="center">
<table width="560" cellpadding="0" cellspacing="0" border="0"
  style="max-width:560px;width:100%;background:#FFFFFF;border-radius:12px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,.07);">
  {header_seg}
  {cuerpo_seg}
  {firma}
  {pie}
</table>
</td></tr>
</table>
</body></html>"""


def _construir_email(p, tipo, compra, venta, sender_email, nombre_completo, cargo, firma_pre=None):
    """Construye el HTML y asunto. Sin dependencia en current_user (usable en hilos)."""
    nombre_saludo = (p.nombre_contacto or p.razon_social or "estimado cliente").split()[0].capitalize()

    presentacion_remitente = (
        f"Mi nombre es <strong>{nombre_completo}</strong>, {cargo} de "
        "<strong>QoriCash SAC</strong>, fintech de cambio de divisas 100% digital, "
        "regulada por la Superintendencia de Banca, Seguros y AFP del Peru."
    )

    fecha = now_peru().strftime("%d/%m/%Y")

    # firma_pre: firma ya obtenida (evita llamada extra a Gmail API en campañas masivas)
    firma_gmail = firma_pre if firma_pre is not None else _get_firma_gmail(sender_email)
    if firma_gmail:
        firma = f'<tr><td style="padding:16px 28px;border-top:1px solid #F1F5F9;background:#FAFAFA;">{firma_gmail}</td></tr>'
    else:
        firma = FIRMA_HTML.replace("{trader_nombre}", nombre_completo).replace("{trader_cargo}", cargo)

    header = HEADER_HTML.replace("{fecha}", fecha).replace("{nombre}", nombre_saludo)

    if tipo == "precio":
        ticker = _build_ticker(compra, venta)
        html   = CUERPO_PRECIO.format(
            header=header,
            mensaje="",   # placeholder requerido por la plantilla
            ticker=ticker, bancos=BANCOS_HTML, firma=firma, pie=PIE,
        )
        return html, "Tipo de cambio QORICASH", "precio_enviado"
    elif tipo == "seguimiento":
        header_seg = (
            f'<tr><td style="background:#0D1B2A;padding:18px 28px;">'
            f'<table width="100%" cellpadding="0" cellspacing="0" border="0"><tr>'
            f'<td><img src="{LOGO_URL}" alt="QoriCash" height="32" style="display:block;height:32px;"></td>'
            f'<td style="padding-left:6px;vertical-align:middle;">'
            f'<span style="font-size:14px;font-weight:800;color:#FFFFFF;letter-spacing:2px;">QORICASH</span></td>'
            f'<td align="right"><span style="font-size:10px;font-weight:700;color:#5CB85C;text-transform:uppercase;'
            f'letter-spacing:1.5px;background:rgba(92,184,92,.12);padding:4px 10px;border-radius:20px;'
            f'border:1px solid rgba(92,184,92,.3);">{fecha}</span></td>'
            f'</tr></table></td></tr>'
        )
        cuerpo_seg = MENSAJE_SEG_HTML.replace("{nombre}", nombre_saludo)
        html = CUERPO_SEGUIMIENTO.format(
            header_seg=header_seg,
            cuerpo_seg=cuerpo_seg,
            firma=firma, pie=PIE,
        )
        return html, "Seguimiento — QoriCash", "presentado"
    else:
        html = CUERPO_PRESENTACION.format(
            header=HEADER_HTML.replace("{fecha}", fecha).replace("{nombre}", nombre_saludo),
            nombre=nombre_saludo,
            presentacion_remitente=presentacion_remitente,
            bancos=BANCOS_HTML, firma=firma, pie=PIE,
        )
        return html, "QoriCash - El mejor tipo de cambio para empresas", "presentado"


@prospeccion_bp.route("/<int:pid>/preview-email", methods=["POST"])
@login_required
@require_role("Master")
def preview_email(pid):
    """Genera el borrador HTML sin enviarlo."""
    p = db.get_or_404(Prospecto, pid)
    _verificar_acceso(p)

    if not p.email:
        return jsonify({"ok": False, "error": "El prospecto no tiene email registrado."}), 400

    data   = request.get_json(force=True)
    tipo   = data.get("tipo", "presentacion")
    compra = data.get("compra", "")
    venta  = data.get("venta", "")

    if tipo == "precio" and (not compra or not venta):
        return jsonify({"ok": False, "error": "Debes indicar COMPRA y VENTA."}), 400

    sender_email            = (getattr(current_user, "email", "") or "").lower()
    nombre_completo, cargo  = _get_trader_info(sender_email, current_user.role)
    html, asunto, _         = _construir_email(p, tipo, compra, venta, sender_email, nombre_completo, cargo)
    return jsonify({"ok": True, "html": html, "asunto": asunto, "para": p.email})


@prospeccion_bp.route("/<int:pid>/crear-borrador", methods=["POST"])
@login_required
@require_role("Master")
@csrf.exempt
def crear_borrador(pid):
    """Genera el HTML del correo y lo guarda como borrador en Gmail del trader."""
    p = db.get_or_404(Prospecto, pid)
    _verificar_acceso(p)

    if not p.email:
        return jsonify({"ok": False, "error": "El prospecto no tiene email registrado."}), 400

    data   = request.get_json(force=True)
    tipo   = data.get("tipo", "presentacion")
    compra = data.get("compra", "")
    venta  = data.get("venta", "")

    if tipo == "precio" and (not compra or not venta):
        return jsonify({"ok": False, "error": "Debes indicar compra y venta para enviar precio."}), 400

    sender_email           = (getattr(current_user, "email", "") or "").lower()
    nombre_completo, cargo = _get_trader_info(sender_email, current_user.role)

    try:
        html, asunto, _ = _construir_email(p, tipo, compra, venta, sender_email, nombre_completo, cargo)
        draft_id, message_id = _crear_borrador_gmail(sender_email, p.email, asunto, html)

        # URL directa al borrador en la cuenta correcta.
        # Google Workspace → mail/a/{dominio}/#drafts/{message_id}
        # Gmail personal   → mail/u/0/#drafts/{message_id}
        domain = sender_email.split("@")[-1] if "@" in sender_email else ""
        if domain and domain != "gmail.com":
            base = f"https://mail.google.com/mail/a/{domain}/#drafts"
        else:
            base = "https://mail.google.com/mail/u/0/#drafts"
        url = f"{base}/{message_id}" if message_id else base

        return jsonify({
            "ok":       True,
            "draft_id": draft_id,
            "url":      url,
            "asunto":   asunto,
            "para":     p.email,
        })
    except ValueError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@prospeccion_bp.route("/<int:pid>/enviar-email", methods=["POST"])
@login_required
@require_role("Master")
def enviar_email(pid):
    p = db.get_or_404(Prospecto, pid)
    _verificar_acceso(p)

    if not p.email:
        return jsonify({"ok": False, "error": "El prospecto no tiene email registrado."}), 400

    data         = request.get_json(force=True)
    tipo         = data.get("tipo", "presentacion")
    compra       = data.get("compra", "")
    venta        = data.get("venta", "")
    html_editado = data.get("html_editado", "").strip()

    if tipo == "precio" and (not compra or not venta):
        return jsonify({"ok": False, "error": "Debes indicar COMPRA y VENTA."}), 400

    sender_email            = (getattr(current_user, "email", "") or "").lower()
    nombre_completo, cargo  = _get_trader_info(sender_email, current_user.role)

    if not sender_email:
        return jsonify({"ok": False, "error": "Tu usuario no tiene email configurado."}), 400

    if html_editado:
        html   = html_editado
        asunto = data.get("asunto", "QoriCash")
    else:
        html, asunto, _ = _construir_email(p, tipo, compra, venta, sender_email, nombre_completo, cargo)

    try:
        _send_via_gmail_api(sender_email, p.email, asunto, html)
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500

    # Enviar también al email alternativo si existe
    enviado_alt = False
    if p.email_alt and p.email_alt.strip() and p.email_alt.strip() != p.email:
        try:
            if html_editado:
                html_alt   = html_editado
                asunto_alt = asunto
            else:
                html_alt, asunto_alt, _ = _construir_email(
                    p, tipo, compra, venta, sender_email, nombre_completo, cargo)
            _send_via_gmail_api(sender_email, p.email_alt.strip(), asunto_alt, html_alt)
            enviado_alt = True
        except Exception:
            pass  # El principal ya fue enviado; el alt falla silencioso

    nuevo_estado = "interesado" if tipo == "precio" else "contactado"
    desc = f"Correo de {tipo} enviado a {p.email}."
    if enviado_alt:
        desc += f" También enviado a {p.email_alt.strip()}."
    _log_actividad(p.id, current_user.id, "email", desc,
                   canal="email", bandeja=sender_email,
                   resultado="Enviado", nuevo_estado=nuevo_estado)
    p.estado_comercial  = nuevo_estado
    p.tipo_ultimo_envio = tipo
    db.session.commit()

    tipo_label = "presentacion" if tipo == "presentacion" else "precio del dia"
    msg = f"Correo de {tipo_label} enviado a {p.email}."
    if enviado_alt:
        msg += f" Copia también enviada a {p.email_alt.strip()}."
    return jsonify({"ok": True, "msg": msg})


# ── Cambiar estado (seguimiento → negociacion) ────────────────────────────────

@prospeccion_bp.route("/<int:pid>/cambiar-estado", methods=["POST"])
@login_required
@require_role("Master", "Trader")
def cambiar_estado(pid):
    p = db.get_or_404(Prospecto, pid)
    _verificar_acceso(p)

    data   = request.get_json(force=True)
    nuevo  = data.get("estado", "")
    motivo = data.get("motivo", "").strip()

    # Normalizar legacy
    nuevo = _ESTADO_NORM.get(nuevo, nuevo)
    if nuevo not in ESTADOS_VALIDOS:
        return jsonify({"ok": False, "error": "Estado inválido."}), 400

    if nuevo == "no_contactar" and motivo:
        descripcion = f"Marcado como No Contactar. Motivo: {motivo}"
    else:
        labels = {"contactado": "Contactado", "interesado": "Interesado",
                  "negociando": "Negociando", "cliente": "Cliente",
                  "no_interesado": "No Interesado", "no_contactar": "No Contactar",
                  "inactivo": "Inactivo", "sin_contactar": "Sin Contactar"}
        descripcion = f"Estado cambiado a: {labels.get(nuevo, nuevo)}."

    _log_actividad(p.id, current_user.id, "estado", descripcion,
                   canal="sistema", nuevo_estado=nuevo)
    p.estado_comercial = nuevo
    db.session.commit()
    ahora = p.fecha_ultimo_contacto or now_peru().strftime("%Y-%m-%d %H:%M")
    return jsonify({"ok": True, "fecha_ultimo_contacto": ahora})


# ── Registrar como cliente ────────────────────────────────────────────────────

@prospeccion_bp.route("/<int:pid>/registrar-cliente", methods=["POST"])
@login_required
@require_role("Master")
def registrar_cliente(pid):
    from app.models.client import Client
    from werkzeug.security import generate_password_hash
    import secrets

    p = db.get_or_404(Prospecto, pid)
    _verificar_acceso(p)

    # 1. Buscar cliente existente por RUC o email
    cliente = None
    if p.ruc:
        cliente = Client.query.filter_by(dni=p.ruc).first()
    if not cliente and p.email:
        cliente = Client.query.filter_by(email=p.email).first()

    if cliente:
        # Ya existe — solo marcar prospecto como convertido
        p.estado_comercial = "cliente"
        act = ActividadProspecto(
            prospecto_id=p.id, user_id=current_user.id,
            tipo="estado",
            descripcion=f"Cruzado con cliente existente ID {cliente.id} ({cliente.email}).",
            nuevo_estado="cliente",
        )
        db.session.add(act)
        db.session.commit()
        return jsonify({"ok": True, "accion": "vinculado",
                        "msg": f"Vinculado con cliente existente: {cliente.email}"})

    # 2. No existe — crear cliente basico con los datos del prospecto
    if not p.email:
        return jsonify({"ok": False,
                        "error": "El prospecto no tiene email. Agrega uno antes de registrarlo como cliente."}), 400

    temp_pass = secrets.token_urlsafe(10)
    nuevo_cliente = Client(
        document_type = "RUC" if p.ruc else "DNI",
        dni           = p.ruc or p.email,
        razon_social  = p.razon_social,
        persona_contacto = p.nombre_contacto,
        email         = p.email,
        phone         = p.telefono,
        departamento  = p.departamento,
        provincia     = p.provincia,
        password_hash = generate_password_hash(temp_pass),
    )
    db.session.add(nuevo_cliente)

    p.estado_comercial = "cliente"
    act = ActividadProspecto(
        prospecto_id=p.id, user_id=current_user.id,
        tipo="estado",
        descripcion=f"Registrado como nuevo cliente ({p.email}).",
        nuevo_estado="cliente",
    )
    db.session.add(act)

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"ok": False, "error": f"Error al crear cliente: {exc}"}), 500

    return jsonify({"ok": True, "accion": "creado",
                    "msg": f"Cliente creado exitosamente: {p.email}"})


# ── Campaña masiva ────────────────────────────────────────────────────────────

def _campana_worker(app, trader_id, prospecto_ids, tipo, sender_email,
                    nombre_completo, cargo, firma_pre):
    """Hilo que envía emails con pausa configurable entre cada uno."""
    pausa = 7  # segundos entre envíos

    with app.app_context():
        stop_event = _campanas[trader_id]["stop"]

        for pid in prospecto_ids:
            if stop_event.is_set():
                break
            try:
                with _campanas_lock:
                    if trader_id not in _campanas or stop_event.is_set():
                        break
                    compra = _campanas[trader_id]["compra"]
                    venta  = _campanas[trader_id]["venta"]

                p = db.session.get(Prospecto, pid)
                if not p or not p.email:
                    with _campanas_lock:
                        if trader_id in _campanas:
                            _campanas[trader_id]["errores"] += 1
                    continue

                html, asunto, nuevo_estado = _construir_email(
                    p, tipo, compra, venta,
                    sender_email, nombre_completo, cargo, firma_pre=firma_pre,
                )
                _send_via_gmail_api(sender_email, p.email, asunto, html)

                # Enviar al email alternativo si existe
                enviado_alt = False
                if p.email_alt and p.email_alt.strip() and p.email_alt.strip() != p.email:
                    try:
                        stop_event.wait(timeout=3)  # pausa corta entre principal y alt
                        if not stop_event.is_set():
                            _send_via_gmail_api(sender_email, p.email_alt.strip(), asunto, html)
                            enviado_alt = True
                    except Exception:
                        pass

                p.estado_comercial      = nuevo_estado
                p.tipo_ultimo_envio     = tipo
                p.fecha_ultimo_contacto = now_peru().strftime("%Y-%m-%d %H:%M")

                nuevo_estado = "precio_enviado" if tipo == "precio" else "presentado"
                desc = f"Campaña masiva: correo de {tipo} enviado."
                if enviado_alt:
                    desc += f" También a email alternativo ({p.email_alt.strip()})."
                act = ActividadProspecto(
                    prospecto_id=p.id,
                    user_id=trader_id,
                    tipo="email",
                    descripcion=desc,
                    resultado="Enviado",
                    nuevo_estado=nuevo_estado,
                )
                db.session.add(act)
                db.session.commit()

                with _campanas_lock:
                    if trader_id in _campanas:
                        _campanas[trader_id]["enviados"] += 1

            except Exception:
                db.session.rollback()
                with _campanas_lock:
                    if trader_id in _campanas:
                        _campanas[trader_id]["errores"] += 1

            # Pausa interruptible
            stop_event.wait(timeout=pausa)

        with _campanas_lock:
            if trader_id in _campanas:
                if stop_event.is_set():
                    _campanas[trader_id]["estado"] = "detenida"
                else:
                    _campanas[trader_id]["estado"] = "completada"


@prospeccion_bp.route("/campana/iniciar", methods=["POST"])
@login_required
@require_role("Master")
def campana_iniciar():
    data   = request.get_json(force=True)
    ids    = [int(i) for i in data.get("ids", []) if str(i).isdigit()]
    tipo   = data.get("tipo", "presentacion")
    compra = data.get("compra", "").strip()
    venta  = data.get("venta", "").strip()

    if not ids:
        return jsonify({"ok": False, "error": "No se seleccionaron prospectos."}), 400
    if tipo == "precio" and (not compra or not venta):
        return jsonify({"ok": False, "error": "Debes indicar COMPRA y VENTA."}), 400

    trader_id = current_user.id
    with _campanas_lock:
        camp = _campanas.get(trader_id)
        if camp and camp["estado"] == "activa":
            return jsonify({"ok": False,
                            "error": "Ya hay una campaña activa. Detén la actual primero."}), 400

    sender_email            = (getattr(current_user, "email", "") or "").lower()
    nombre_completo, cargo  = _get_trader_info(sender_email, current_user.role)
    firma_pre               = _get_firma_gmail(sender_email)

    stop_event = threading.Event()
    with _campanas_lock:
        _campanas[trader_id] = {
            "estado":   "activa",
            "tipo":     tipo,
            "compra":   compra,
            "venta":    venta,
            "total":    len(ids),
            "enviados": 0,
            "errores":  0,
            "stop":     stop_event,
        }

    app = current_app._get_current_object()
    t = threading.Thread(
        target=_campana_worker,
        args=(app, trader_id, ids, tipo, sender_email, nombre_completo, cargo, firma_pre),
        daemon=True,
    )
    t.start()
    return jsonify({"ok": True, "total": len(ids)})


@prospeccion_bp.route("/campana/estado")
@login_required
@require_role("Master")
def campana_estado():
    trader_id = current_user.id
    with _campanas_lock:
        camp = _campanas.get(trader_id)
        if not camp:
            return jsonify({"activa": False})
        resp = {
            "activa":   camp["estado"] == "activa",
            "estado":   camp["estado"],
            "tipo":     camp["tipo"],
            "total":    camp["total"],
            "enviados": camp["enviados"],
            "errores":  camp["errores"],
            "compra":   camp.get("compra", ""),
            "venta":    camp.get("venta", ""),
        }
        # Limpiar estado final para que no persista entre recargas
        if camp["estado"] in ("completada", "detenida"):
            del _campanas[trader_id]
        return jsonify(resp)


@prospeccion_bp.route("/campana/detener", methods=["POST"])
@login_required
@require_role("Master")
def campana_detener():
    trader_id = current_user.id
    with _campanas_lock:
        camp = _campanas.get(trader_id)
        if not camp or camp["estado"] != "activa":
            return jsonify({"ok": False, "error": "No hay campaña activa."}), 400
        camp["stop"].set()
    return jsonify({"ok": True})


@prospeccion_bp.route("/<int:pid>/solicitar-extension", methods=["POST"])
@login_required
@require_role("Trader")
def solicitar_extension(pid):
    """Trader solicita 45 días extra para un prospecto en negociación."""
    p = db.get_or_404(Prospecto, pid)
    _verificar_acceso(p)

    if p.estado_comercial not in ("negociacion", "P2", "P3"):
        return jsonify({"ok": False,
                        "error": "Solo puedes solicitar extensión para prospectos En Negociacion."}), 400

    asig = next((a for a in p.asignaciones if a.trader_id == current_user.id and a.activo), None)
    if not asig:
        return jsonify({"ok": False, "error": "No tienes este prospecto asignado."}), 400
    if asig.extension_solicitada:
        return jsonify({"ok": False, "error": "Ya hay una solicitud pendiente de aprobacion."}), 400

    asig.extension_solicitada = True
    db.session.add(ActividadProspecto(
        prospecto_id=pid,
        user_id=current_user.id,
        tipo="nota",
        descripcion="Solicitud de extension de vigencia enviada al Master (prospecto en negociacion).",
    ))
    db.session.commit()
    return jsonify({"ok": True})


@prospeccion_bp.route("/<int:pid>/resolver-extension", methods=["POST"])
@login_required
@require_role("Master")
def resolver_extension(pid):
    """Master aprueba o rechaza la extensión de vigencia."""
    data    = request.get_json(force=True)
    accion  = data.get("accion")          # "aprobar" | "rechazar"
    asig_id = int(data.get("asig_id", 0))

    asig = db.session.get(AsignacionProspecto, asig_id)
    if not asig or asig.prospecto_id != pid:
        return jsonify({"ok": False, "error": "Asignacion no encontrada."}), 404

    if accion == "aprobar":
        asig.dias_extra           = (asig.dias_extra or 0) + DIAS_VIGENCIA
        asig.extension_solicitada = False
        desc = f"Extension de {DIAS_VIGENCIA} dias aprobada por {current_user.username}."
    else:
        asig.extension_solicitada = False
        desc = f"Solicitud de extension rechazada por {current_user.username}."

    db.session.add(ActividadProspecto(
        prospecto_id=pid,
        user_id=current_user.id,
        tipo="nota",
        descripcion=desc,
    ))
    db.session.commit()
    return jsonify({"ok": True, "accion": accion})


@prospeccion_bp.route("/api/migrar-vigencia", methods=["POST"])
@login_required
@require_role("Master")
def migrar_vigencia():
    """Agrega columnas dias_extra y extension_solicitada a asignaciones_prospecto (idempotente)."""
    from sqlalchemy import text
    try:
        with db.engine.connect() as conn:
            conn.execute(text(
                "ALTER TABLE asignaciones_prospecto "
                "ADD COLUMN IF NOT EXISTS dias_extra INTEGER DEFAULT 0"
            ))
            conn.execute(text(
                "ALTER TABLE asignaciones_prospecto "
                "ADD COLUMN IF NOT EXISTS extension_solicitada BOOLEAN DEFAULT FALSE"
            ))
            conn.commit()
        return jsonify({"ok": True, "msg": "Migracion completada."})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@prospeccion_bp.route("/reasignar-vencidos", methods=["POST"])
@login_required
@require_role("Master")
def reasignar_vencidos():
    data            = request.get_json(force=True)
    ids             = [int(i) for i in data.get("ids", []) if str(i).isdigit()]
    nuevo_trader_id = int(data.get("trader_id", 0))

    if not ids or not nuevo_trader_id:
        return jsonify({"ok": False, "error": "Faltan parámetros."}), 400

    nuevo_trader = db.session.get(User, nuevo_trader_id)
    if not nuevo_trader:
        return jsonify({"ok": False, "error": "Trader no encontrado."}), 404

    ahora       = now_peru()
    reasignados = 0

    for pid in ids:
        p = db.session.get(Prospecto, pid)
        if not p:
            continue

        # Desactivar asignación actual
        AsignacionProspecto.query.filter_by(prospecto_id=p.id, activo=True).update({"activo": False})

        # Crear o reactivar asignación al nuevo trader
        existente = AsignacionProspecto.query.filter_by(
            prospecto_id=pid, trader_id=nuevo_trader_id
        ).first()
        if existente:
            existente.activo       = True
            existente.asignado_en  = ahora
            existente.asignado_por = current_user.id
        else:
            db.session.add(AsignacionProspecto(
                prospecto_id=pid,
                trader_id=nuevo_trader_id,
                asignado_por=current_user.id,
                asignado_en=ahora,
            ))

        # Registrar actividad
        db.session.add(ActividadProspecto(
            prospecto_id=pid,
            user_id=current_user.id,
            tipo="nota",
            descripcion=(f"Reasignado a {nuevo_trader.username} "
                         f"por vencimiento de vigencia ({DIAS_VIGENCIA} días sin actividad)."),
        ))
        reasignados += 1

    db.session.commit()
    return jsonify({"ok": True, "reasignados": reasignados})


@prospeccion_bp.route("/campana/actualizar-precios", methods=["POST"])
@login_required
@require_role("Master")
def campana_actualizar_precios():
    data   = request.get_json(force=True)
    compra = data.get("compra", "").strip()
    venta  = data.get("venta", "").strip()
    if not compra or not venta:
        return jsonify({"ok": False, "error": "Debes indicar COMPRA y VENTA."}), 400

    trader_id = current_user.id
    with _campanas_lock:
        camp = _campanas.get(trader_id)
        if not camp or camp["estado"] != "activa":
            return jsonify({"ok": False, "error": "No hay campaña activa."}), 400
        camp["compra"] = compra
        camp["venta"]  = venta
    return jsonify({"ok": True})


# ── Limpieza de emails invalidos ─────────────────────────────────────────────

_limpiar_estado: dict = {}   # {"estado": "corriendo"|"listo"|"error", "total":n, "procesados":n, "limpiados":n}
_limpiar_lock   = threading.Lock()


def _limpiar_emails_worker(app):
    """Corre en background: valida MX de cada email y nullifica los invalidos."""
    import re, socket
    try:
        import dns.resolver as _res
    except ImportError:
        _res = None

    FREE_DOMAINS = {
        "gmail.com","hotmail.com","yahoo.com","outlook.com","live.com",
        "icloud.com","me.com","aol.com","msn.com","protonmail.com",
        "yahoo.es","hotmail.es","yahoo.com.pe","gmail.pe",
    }
    SYNTAX_RE = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')

    _mx_cache: dict = {}

    def tiene_mx(domain: str) -> bool:
        if domain in _mx_cache:
            return _mx_cache[domain]
        if _res is None:
            _mx_cache[domain] = True
            return True
        try:
            records = _res.resolve(domain, "MX", lifetime=5)
            ok = len(records) > 0
        except Exception:
            ok = False
        _mx_cache[domain] = ok
        return ok

    def email_valido(email: str) -> bool:
        if not email or not SYNTAX_RE.match(email.strip()):
            return False
        domain = email.strip().split("@")[1].lower()
        # Dominios gratis: se mantienen (son contactos reales)
        if domain in FREE_DOMAINS:
            return True
        return tiene_mx(domain)

    with app.app_context():
        try:
            # Traer todos los prospectos con email
            rows = db.session.execute(
                db.text("SELECT id, email FROM prospectos WHERE email IS NOT NULL AND email != ''")
            ).fetchall()

            total = len(rows)
            with _limpiar_lock:
                _limpiar_estado.update({"estado": "corriendo", "total": total,
                                        "procesados": 0, "limpiados": 0})

            limpiados = 0
            for i, (pid, email) in enumerate(rows):
                if not email_valido(email):
                    db.session.execute(
                        db.text("UPDATE prospectos SET email = NULL WHERE id = :id"),
                        {"id": pid}
                    )
                    db.session.execute(
                        db.text("UPDATE prospectos SET estado_email = 'Dominio inválido' WHERE id = :id"),
                        {"id": pid}
                    )
                    limpiados += 1

                if (i + 1) % 100 == 0:
                    db.session.commit()
                    with _limpiar_lock:
                        _limpiar_estado.update({"procesados": i + 1, "limpiados": limpiados})

            db.session.commit()
            with _limpiar_lock:
                _limpiar_estado.update({
                    "estado": "listo",
                    "procesados": total,
                    "limpiados": limpiados,
                })
        except Exception as e:
            db.session.rollback()
            with _limpiar_lock:
                _limpiar_estado["estado"] = f"error: {e}"


@prospeccion_bp.route("/api/limpiar-emails", methods=["POST"])
@login_required
@require_role("Master")
def limpiar_emails_iniciar():
    with _limpiar_lock:
        if _limpiar_estado.get("estado") == "corriendo":
            return jsonify({"ok": False, "error": "Ya está corriendo."})
        _limpiar_estado.clear()
        _limpiar_estado["estado"] = "corriendo"

    t = threading.Thread(
        target=_limpiar_emails_worker,
        args=(current_app._get_current_object(),),
        daemon=True,
    )
    t.start()
    return jsonify({"ok": True})


@prospeccion_bp.route("/api/limpiar-emails/estado")
@login_required
@require_role("Master")
def limpiar_emails_estado():
    with _limpiar_lock:
        return jsonify(dict(_limpiar_estado))


# ── CRM Grid API ──────────────────────────────────────────────────────────────

_CAMPOS_EDITABLES = {
    "razon_social", "ruc", "tipo", "rubro", "departamento", "provincia",
    "nombre_contacto", "cargo", "email", "email_alt", "telefono", "telefono_alt",
    "tamano_empresa", "volumen_estimado_usd", "prioridad",
    "cliente_lfc", "canal", "fuente", "remitente",
    "fecha_proximo_contacto", "fecha_ultimo_contacto", "fecha_primer_contacto",
    "estado_comercial", "nivel_interes", "grupo", "notas",
    "tipo_ultimo_envio", "estado_email", "clasificacion", "score",
}


@prospeccion_bp.route("/api/grid")
@login_required
@require_role("Master", "Trader")
def api_grid():
    """Retorna hasta 5000 prospectos como JSON para AG Grid."""
    try:
        return _api_grid_impl()
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"ok": False, "error": str(e), "rows": [], "total": 0, "counts": {}, "traders": [], "rubros": [], "deptos": []}), 500


def _api_grid_impl():
    q_str     = request.args.get("q", "").strip()
    tab       = request.args.get("tab", "todos")
    depto     = request.args.get("depto", "")
    rubro_f   = request.args.get("rubro", "")
    estado_f  = request.args.get("estado", "")
    tamano_f  = request.args.get("tamano", "")
    nivel_f   = request.args.get("nivel_interes", "")
    prioridad_f = request.args.get("prioridad", "")
    trader_id_f = request.args.get("trader_id", type=int)

    query = _base_query()

    # Trader solo ve sus prospectos asignados activos
    if current_user.role == "Trader":
        query = query.join(
            AsignacionProspecto,
            (AsignacionProspecto.prospecto_id == Prospecto.id) &
            (AsignacionProspecto.activo == True) &
            (AsignacionProspecto.trader_id == current_user.id)
        )

    _clientes_vals   = _FASE_ESTADOS["clientes"]
    _contactado_vals = _FASE_ESTADOS["contactado"]
    _interesado_vals = _FASE_ESTADOS["interesado"]
    _negociando_vals = _FASE_ESTADOS["negociando"]
    _excluir_todos   = _clientes_vals + ["no_contactar", "no_interesado", "inactivo"]

    if tab == "clientes":
        query = query.filter(Prospecto.estado_comercial.in_(_clientes_vals))
    elif tab in ("presentado", "contactado"):
        query = query.filter(Prospecto.estado_comercial.in_(_contactado_vals))
    elif tab in ("precio", "interesado"):
        query = query.filter(Prospecto.estado_comercial.in_(_interesado_vals))
    elif tab == "negociando":
        query = query.filter(Prospecto.estado_comercial.in_(_negociando_vals))
    elif tab == "sin_contactar":
        query = query.filter(or_(
            Prospecto.estado_comercial == None,
            Prospecto.estado_comercial == "",
            Prospecto.estado_comercial == "sin_contactar"))
    elif tab == "no_contactar":
        query = query.filter(Prospecto.estado_comercial == "no_contactar")
    else:
        # "todos": excluye clientes, no_contactar, no_interesado, inactivo
        query = query.filter(or_(
            Prospecto.estado_comercial == None,
            Prospecto.estado_comercial == "",
            Prospecto.estado_comercial.notin_(_excluir_todos),
        ))

    if q_str:
        like = f"%{q_str}%"
        query = query.filter(or_(
            Prospecto.razon_social.ilike(like),
            Prospecto.email.ilike(like),
            Prospecto.nombre_contacto.ilike(like),
            Prospecto.ruc.ilike(like),
            Prospecto.telefono.ilike(like),
        ))
    if depto:
        query = query.filter(Prospecto.departamento.ilike(f"%{depto}%"))
    if rubro_f:
        query = query.filter(Prospecto.rubro.ilike(f"%{rubro_f}%"))
    if estado_f:
        query = query.filter(Prospecto.estado_comercial == estado_f)
    if tamano_f:
        query = query.filter(Prospecto.tamano_empresa == tamano_f)
    if nivel_f:
        query = query.filter(Prospecto.nivel_interes == nivel_f)
    if prioridad_f:
        query = query.filter(Prospecto.prioridad == prioridad_f)
    if trader_id_f:
        query = query.join(
            AsignacionProspecto,
            (AsignacionProspecto.prospecto_id == Prospecto.id) &
            (AsignacionProspecto.activo == True)
        ).filter(AsignacionProspecto.trader_id == trader_id_f)

    prospectos_list = query.order_by(Prospecto.actualizado_en.desc()).limit(5000).all()
    ids = [p.id for p in prospectos_list]

    # Asignaciones (bulk, sin N+1)
    asigs = AsignacionProspecto.query.filter(
        AsignacionProspecto.prospecto_id.in_(ids),
        AsignacionProspecto.activo == True,
    ).all() if ids else []
    trader_map = {}
    for a in asigs:
        if a.prospecto_id not in trader_map and a.trader:
            trader_map[a.prospecto_id] = a.trader.username

    # Emails extra (bulk)
    from app.models.prospecto import ProspectoEmail
    try:
        extras = ProspectoEmail.query.filter(
            ProspectoEmail.prospecto_id.in_(ids),
            ProspectoEmail.activo == True,
        ).all() if ids else []
        emails_map = {}
        for e in extras:
            emails_map.setdefault(e.prospecto_id, []).append(e.email)
    except Exception:
        emails_map = {}

    # Última actividad WhatsApp por prospecto (bulk)
    wa_acts = (
        db.session.query(
            ActividadProspecto.prospecto_id,
            func.max(ActividadProspecto.creado_en).label('ultima_wa')
        )
        .filter(
            ActividadProspecto.canal == 'whatsapp',
            ActividadProspecto.prospecto_id.in_(ids),
        )
        .group_by(ActividadProspecto.prospecto_id)
        .all()
    ) if ids else []
    wa_map = {r.prospecto_id: r.ultima_wa for r in wa_acts}

    # Última actividad Email por prospecto (bulk)
    email_acts = (
        db.session.query(
            ActividadProspecto.prospecto_id,
            func.max(ActividadProspecto.creado_en).label('ultima_email')
        )
        .filter(
            ActividadProspecto.canal == 'email',
            ActividadProspecto.prospecto_id.in_(ids),
        )
        .group_by(ActividadProspecto.prospecto_id)
        .all()
    ) if ids else []
    email_map = {r.prospecto_id: r.ultima_email for r in email_acts}

    rows = []
    for p in prospectos_list:
        all_emails = []
        if p.email:
            all_emails.append(p.email)
        if p.email_alt and p.email_alt not in all_emails:
            all_emails.append(p.email_alt)
        for em in emails_map.get(p.id, []):
            if em not in all_emails:
                all_emails.append(em)

        rows.append({
            "id":                    p.id,
            "razon_social":          p.razon_social or "",
            "ruc":                   p.ruc or "",
            "tipo":                  p.tipo or "",
            "rubro":                 p.rubro or "",
            "departamento":          p.departamento or "",
            "provincia":             p.provincia or "",
            "distrito":              getattr(p, "distrito", "") or "",
            "web":                   getattr(p, "web", "") or "",
            "nombre_contacto":       p.nombre_contacto or "",
            "cargo":                 p.cargo or "",
            "email":                 p.email or "",
            "emails":                all_emails,
            "telefono":              p.telefono or "",
            "telefono_alt":          getattr(p, "telefono_alt", "") or "",
            "telefono_3":            getattr(p, "telefono_3", "") or "",
            "telefono_4":            getattr(p, "telefono_4", "") or "",
            "tamano_empresa":        getattr(p, "tamano_empresa", "") or "",
            "volumen_estimado_usd":  float(getattr(p, "volumen_estimado_usd", None) or 0) or None,
            "prioridad":             getattr(p, "prioridad", "") or "",
            "cliente_lfc":           p.cliente_lfc or "",
            "score":                 p.score or 0,
            "clasificacion":         p.clasificacion or "",
            "canal":                 p.canal or "",
            "fuente":                p.fuente or "",
            "remitente":             p.remitente or "",
            "tipo_ultimo_envio":     p.tipo_ultimo_envio or "",
            "fecha_primer_contacto": p.fecha_primer_contacto or "",
            "fecha_ultimo_contacto": p.fecha_ultimo_contacto or "",
            "fecha_proximo_contacto":p.fecha_proximo_contacto or "",
            "num_contactos":         p.num_contactos or 0,
            "estado_email":          p.estado_email or "",
            "estado_comercial":      p.estado_comercial or "sin_contactar",
            "nivel_interes":         p.nivel_interes or "",
            "grupo":                 p.grupo or "",
            "notas":                 p.notas or "",
            "trader":                trader_map.get(p.id, ""),
            "creado_en":             p.creado_en.strftime("%Y-%m-%d") if p.creado_en else "",
            "actualizado_en":        p.actualizado_en.strftime("%Y-%m-%d %H:%M") if p.actualizado_en else "",
            "ultima_wa":             wa_map[p.id].strftime("%Y-%m-%d %H:%M") if p.id in wa_map and wa_map[p.id] else "",
            "ultima_email":          email_map[p.id].strftime("%Y-%m-%d %H:%M") if p.id in email_map and email_map[p.id] else "",
        })

    base = _base_query()
    _excl = _FASE_ESTADOS["clientes"] + ["no_contactar", "no_interesado", "inactivo"]
    counts = {
        "todos":         base.filter(or_(Prospecto.estado_comercial == None, Prospecto.estado_comercial == "", Prospecto.estado_comercial.notin_(_excl))).count(),
        "sin_contactar": base.filter(or_(Prospecto.estado_comercial == None, Prospecto.estado_comercial == "", Prospecto.estado_comercial == "sin_contactar")).count(),
        "contactado":    base.filter(Prospecto.estado_comercial.in_(_FASE_ESTADOS["contactado"])).count(),
        "interesado":    base.filter(Prospecto.estado_comercial.in_(_FASE_ESTADOS["interesado"])).count(),
        "negociando":    base.filter(Prospecto.estado_comercial.in_(_FASE_ESTADOS["negociando"])).count(),
        "clientes":      base.filter(Prospecto.estado_comercial.in_(_FASE_ESTADOS["clientes"])).count(),
        "no_contactar":  base.filter(Prospecto.estado_comercial == "no_contactar").count(),
    }

    traders = [{"id": u.id, "username": u.username} for u in
               User.query.filter_by(role="Trader", status="Activo").order_by(User.username).all()]
    rubros  = [r[0] for r in db.session.query(Prospecto.rubro).distinct().order_by(Prospecto.rubro).all() if r[0]]
    deptos  = [d[0] for d in db.session.query(Prospecto.departamento).distinct().order_by(Prospecto.departamento).all() if d[0]]
    tamanos = [t[0] for t in db.session.query(Prospecto.tamano_empresa).distinct().order_by(Prospecto.tamano_empresa).all() if t[0]]

    return jsonify({
        "ok": True,
        "rows": rows, "total": len(rows),
        "counts": counts,
        "traders": traders,
        "rubros": rubros,
        "deptos": deptos,
        "tamanos": tamanos,
    })


@prospeccion_bp.route("/api/<int:pid>/campo", methods=["PATCH"])
@csrf.exempt
@login_required
@require_role("Master", "Trader")
def api_campo(pid):
    """Actualiza un campo inline (autosave del grid)."""
    p = db.get_or_404(Prospecto, pid)
    data  = request.get_json() or {}
    campo = data.get("campo", "").strip()
    valor = data.get("valor")

    if campo not in _CAMPOS_EDITABLES:
        return jsonify({"ok": False, "error": f"Campo no permitido: {campo}"}), 400

    try:
        if campo == "volumen_estimado_usd":
            from decimal import Decimal
            valor = Decimal(str(valor)) if valor else None
        elif campo == "score":
            valor = int(valor) if valor else 0
        elif valor == "":
            valor = None

        setattr(p, campo, valor)

        if campo == "estado_comercial" and valor:
            valor = _ESTADO_NORM.get(valor, valor)
            setattr(p, campo, valor)
            _log_actividad(p.id, current_user.id, "estado",
                           f"Estado cambiado a: {valor} (inline)",
                           canal="sistema", nuevo_estado=valor)

        db.session.commit()
        return jsonify({"ok": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500


@prospeccion_bp.route("/api/<int:pid>/notas")
@login_required
@require_role("Master", "Trader")
def api_notas_get(pid):
    """Lista todas las notas (actividades tipo nota) del prospecto."""
    db.get_or_404(Prospecto, pid)
    acts = (ActividadProspecto.query
            .filter_by(prospecto_id=pid, tipo="nota")
            .order_by(ActividadProspecto.creado_en.desc())
            .limit(100).all())
    return jsonify({"ok": True, "notas": [
        {
            "id": a.id,
            "texto": a.descripcion,
            "fecha": a.creado_en.strftime("%d/%m/%Y %H:%M") if a.creado_en else "",
            "usuario": a.usuario.username if a.usuario else "",
        } for a in acts
    ]})


@prospeccion_bp.route("/api/<int:pid>/nota", methods=["POST"])
@csrf.exempt
@login_required
@require_role("Master", "Trader")
def api_nota_add(pid):
    """Agrega una nota al prospecto como actividad."""
    db.get_or_404(Prospecto, pid)
    data  = request.get_json() or {}
    texto = data.get("texto", "").strip()
    if not texto:
        return jsonify({"ok": False, "error": "Texto requerido"}), 400
    act = ActividadProspecto(
        prospecto_id=pid,
        user_id=current_user.id,
        tipo="nota",
        descripcion=texto,
    )
    db.session.add(act)
    db.session.commit()
    return jsonify({"ok": True, "nota": {
        "id": act.id,
        "texto": act.descripcion,
        "fecha": act.creado_en.strftime("%d/%m/%Y %H:%M") if act.creado_en else "",
        "usuario": current_user.username,
    }})


@prospeccion_bp.route("/api/<int:pid>/emails")
@login_required
@require_role("Master")
def api_emails_list(pid):
    """Lista todos los emails del prospecto."""
    from app.models.prospecto import ProspectoEmail
    p = db.get_or_404(Prospecto, pid)
    result = []
    if p.email:
        result.append({"id": None, "email": p.email, "activo": True, "es_principal": True})
    if p.email_alt:
        result.append({"id": None, "email": p.email_alt, "activo": True, "es_principal": False, "es_alt": True})
    try:
        for e in ProspectoEmail.query.filter_by(prospecto_id=pid).order_by(ProspectoEmail.creado_en).all():
            result.append({"id": e.id, "email": e.email, "activo": e.activo, "es_principal": False})
    except Exception:
        pass
    return jsonify({"emails": result, "razon_social": p.razon_social})


@prospeccion_bp.route("/api/<int:pid>/emails", methods=["POST"])
@csrf.exempt
@login_required
@require_role("Master")
def api_emails_add(pid):
    """Agrega un nuevo email al prospecto."""
    from app.models.prospecto import ProspectoEmail
    p = db.get_or_404(Prospecto, pid)
    data  = request.get_json() or {}
    email = (data.get("email") or "").strip().lower()

    if not email or "@" not in email or "." not in email:
        return jsonify({"ok": False, "error": "Email inválido"}), 400

    existing = ([p.email, p.email_alt] +
                [e.email for e in ProspectoEmail.query.filter_by(prospecto_id=pid).all()])
    if email in [e for e in existing if e]:
        return jsonify({"ok": False, "error": "Este email ya existe para este prospecto"}), 400

    nuevo = ProspectoEmail(prospecto_id=pid, email=email, activo=True)
    db.session.add(nuevo)
    db.session.commit()
    return jsonify({"ok": True, "id": nuevo.id, "email": nuevo.email})


@prospeccion_bp.route("/api/email/<int:eid>/toggle", methods=["PATCH"])
@csrf.exempt
@login_required
@require_role("Master")
def api_email_toggle(eid):
    """Activa/desactiva un email extra."""
    from app.models.prospecto import ProspectoEmail
    e = db.get_or_404(ProspectoEmail, eid)
    e.activo = not e.activo
    db.session.commit()
    return jsonify({"ok": True, "activo": e.activo})


@prospeccion_bp.route("/api/export-excel")
@login_required
@require_role("Master")
def api_export_excel():
    """Exporta todos los prospectos a Excel."""
    import openpyxl
    from io import BytesIO
    from openpyxl.styles import Font, PatternFill, Alignment

    prospectos_list = _base_query().order_by(Prospecto.actualizado_en.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prospectos"

    headers = [
        "ID", "Razón Social", "RUC", "Tipo", "Rubro",
        "Departamento", "Provincia", "Distrito", "Web",
        "Contacto", "Cargo", "Email", "Email Alt", "Teléfono", "Teléfono Alt",
        "Tamaño Empresa", "Volumen Est. USD", "Prioridad",
        "Estado Comercial", "Nivel Interés", "Grupo", "Canal", "Fuente",
        "Tipo Último Envío", "Fecha Primer Contacto", "Fecha Último Contacto",
        "Fecha Próximo Contacto", "N° Contactos", "Notas",
        "Trader Asignado",
        "Última Actividad Tipo", "Última Actividad Fecha", "Última Actividad Descripción",
        "Creado En", "Actualizado En",
    ]

    header_fill = PatternFill("solid", fgColor="1a1a2e")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Trader map
    asigs = AsignacionProspecto.query.filter_by(activo=True).all()
    tmap  = {}
    for a in asigs:
        if a.prospecto_id not in tmap and a.trader:
            tmap[a.prospecto_id] = a.trader.username

    # Última actividad por prospecto
    from sqlalchemy import select as sa_select
    latest_act_subq = (
        db.session.query(
            ActividadProspecto.prospecto_id,
            func.max(ActividadProspecto.creado_en).label("max_ts"),
        )
        .group_by(ActividadProspecto.prospecto_id)
        .subquery()
    )
    act_rows = (
        db.session.query(ActividadProspecto)
        .join(latest_act_subq, (ActividadProspecto.prospecto_id == latest_act_subq.c.prospecto_id) &
              (ActividadProspecto.creado_en == latest_act_subq.c.max_ts))
        .all()
    )
    amap = {a.prospecto_id: a for a in act_rows}

    for p in prospectos_list:
        ult = amap.get(p.id)
        ws.append([
            p.id, p.razon_social, p.ruc, p.tipo, p.rubro,
            p.departamento, p.provincia,
            getattr(p, "distrito", "") or "",
            getattr(p, "web", "") or "",
            p.nombre_contacto, p.cargo, p.email, p.email_alt, p.telefono,
            getattr(p, "telefono_alt", "") or "",
            getattr(p, "tamano_empresa", "") or "",
            float(getattr(p, "volumen_estimado_usd", None) or 0) or None,
            getattr(p, "prioridad", "") or "",
            p.estado_comercial, p.nivel_interes, p.grupo, p.canal, p.fuente,
            p.tipo_ultimo_envio, p.fecha_primer_contacto, p.fecha_ultimo_contacto,
            p.fecha_proximo_contacto, p.num_contactos, p.notas,
            tmap.get(p.id, ""),
            ult.tipo if ult else "",
            ult.creado_en.strftime("%Y-%m-%d %H:%M") if ult and ult.creado_en else "",
            (ult.descripcion or "")[:200] if ult else "",
            p.creado_en.strftime("%Y-%m-%d %H:%M") if p.creado_en else "",
            p.actualizado_en.strftime("%Y-%m-%d %H:%M") if p.actualizado_en else "",
        ])

    # Ajustar anchos
    col_widths = [6,30,14,10,20,15,12,12,25,25,20,30,30,14,14,12,14,8,16,12,20,12,12,14,16,16,16,6,40,15,14,16,40,16,16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    from datetime import date as _date
    return current_app.response_class(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename=prospectos_{_date.today()}.xlsx"},
    )


@prospeccion_bp.route("/api/import-excel", methods=["POST"])
@csrf.exempt
@login_required
@require_role("Master")
def api_import_excel():
    """Importa prospectos desde Excel (archivo o link Google Sheets/Drive).
    Validación estricta: 15 columnas en orden exacto."""
    import unicodedata
    import openpyxl
    from io import BytesIO

    # ── Resolver fuente: archivo subido o URL ──────────────────────────────────
    file_bytes = None

    raw_url = (request.form.get("url") or "").strip()
    if raw_url:
        download_url = _resolve_import_url(raw_url)
        if not download_url:
            return jsonify({"ok": False, "error": "URL no reconocida. Debe ser un link de Google Sheets o Google Drive."}), 400
        try:
            import requests as _req
            _headers = {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                )
            }
            resp = _req.get(download_url, timeout=30, allow_redirects=True, headers=_headers)
            if resp.status_code != 200:
                return jsonify({"ok": False, "error": (
                    f"Google no permitió la descarga (HTTP {resp.status_code}). "
                    "Verifica que la hoja esté compartida como 'Cualquier persona con el enlace puede ver'."
                )}), 400
            # Verificar que la respuesta es xlsx y no HTML (página de error/login de Google)
            ct = resp.headers.get("Content-Type", "")
            if "html" in ct or resp.content[:2] != b"PK":
                return jsonify({"ok": False, "error": (
                    "Google devolvió una página web en lugar del archivo. "
                    "Asegúrate de que la hoja esté compartida públicamente: "
                    "Archivo → Compartir → 'Cualquier persona con el enlace' → Ver."
                )}), 400
            file_bytes = BytesIO(resp.content)
        except Exception as e:
            return jsonify({"ok": False, "error": f"Error al descargar desde Google: {e}"}), 400
    else:
        f = request.files.get("file")
        if not f:
            return jsonify({"ok": False, "error": "No se envió archivo ni URL"}), 400
        file_bytes = f

    # Columnas esperadas en ORDEN EXACTO (posición 0-14)
    EXPECTED_COLS = [
        "RUC",
        "RAZÓN SOCIAL",
        "SECTOR",
        "TAMAÑO EMPRESA",
        "DEPARTAMENTO",
        "PROVINCIA",
        "DISTRITO",
        "CELULAR 1",
        "CELULAR 2",
        "EMAIL",
        "WEB",
        "CARGO",
        "NOMBRES",
        "APELLIDO PATERNO",
        "APELLIDO MATERNO",
    ]

    def _norm(s):
        """Normaliza a mayúsculas sin tildes ni espacios extra."""
        if not s:
            return ""
        s = str(s).strip().upper()
        return "".join(
            c for c in unicodedata.normalize("NFD", s)
            if unicodedata.category(c) != "Mn"
        )

    EXPECTED_NORM = [_norm(c) for c in EXPECTED_COLS]

    f = request.files.get("file")
    if not f:
        return jsonify({"ok": False, "error": "No se envió archivo"}), 400

    try:
        wb   = openpyxl.load_workbook(file_bytes, data_only=True)
        ws   = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        return jsonify({"ok": False, "error": f"No se pudo leer el archivo: {e}"}), 400

    if not rows:
        return jsonify({"ok": False, "error": "Archivo vacío"}), 400

    # ── Validación estricta de encabezados ─────────────────────────────────────
    raw_header = [str(c).strip() if c is not None else "" for c in rows[0]]
    actual_norm = [_norm(c) for c in raw_header]

    # Debe tener al menos 15 columnas
    if len(actual_norm) < 15:
        return jsonify({
            "ok":    False,
            "error": (
                f"El archivo tiene {len(actual_norm)} columna(s) pero se requieren exactamente 15. "
                "Descarga la plantilla oficial e inténtalo de nuevo."
            ),
            "columnas_esperadas": EXPECTED_COLS,
            "columnas_recibidas": raw_header,
        }), 400

    # Verificar cada columna en su posición exacta
    errores_col = []
    for i, (exp, act) in enumerate(zip(EXPECTED_NORM, actual_norm[:15])):
        if exp != act:
            errores_col.append(
                f"Columna {i+1}: esperada '{EXPECTED_COLS[i]}', recibida '{raw_header[i] or '(vacía)'}'"
            )

    if errores_col:
        return jsonify({
            "ok":    False,
            "error": "El archivo no tiene el formato correcto. " + "; ".join(errores_col),
            "columnas_esperadas": EXPECTED_COLS,
            "columnas_recibidas": raw_header[:15],
        }), 400

    # ── Deduplicación previa ───────────────────────────────────────────────────
    existing_emails = set(
        r[0].lower() for r in
        db.session.query(Prospecto.email).filter(Prospecto.email != None).all()
    )
    existing_rucs = set(
        r[0] for r in
        db.session.query(Prospecto.ruc).filter(Prospecto.ruc != None).all()
    )

    def cell(row, idx):
        if idx >= len(row):
            return None
        v = row[idx]
        return str(v).strip() if v is not None else None

    from app.models.prospecto import ProspectoEmail

    inserted = skipped = 0
    # Acumular (prospecto, emails_extra) para insertar emails extra post-flush
    pending_extras: list = []

    for i, row in enumerate(rows[1:], start=2):
        if not any(c for c in row if c):
            continue

        ruc_v = (cell(row, 0) or "").strip()

        # Separar emails mergeados (pueden venir con ";" del archivo limpio)
        raw_email_field = (cell(row, 9) or "")
        all_emails = [e.strip().lower() for e in raw_email_field.split(";") if e.strip()]
        email_v = all_emails[0] if all_emails else ""
        extra_emails = all_emails[1:]

        # Deduplicación: saltar si el primer email ya existe
        if email_v and email_v in existing_emails:
            skipped += 1
            continue
        if ruc_v and ruc_v in existing_rucs and not email_v:
            skipped += 1
            continue

        # Combinar NOMBRES + APELLIDO PATERNO + APELLIDO MATERNO
        nombres = cell(row, 12) or ""
        ap_pat  = cell(row, 13) or ""
        ap_mat  = cell(row, 14) or ""
        nombre_completo = " ".join(filter(None, [nombres, ap_pat, ap_mat])) or None

        p = Prospecto(
            ruc             = ruc_v or None,
            razon_social    = cell(row, 1),
            rubro           = cell(row, 2),
            tamano_empresa  = cell(row, 3),
            departamento    = cell(row, 4),
            provincia       = cell(row, 5),
            distrito        = cell(row, 6),
            telefono        = cell(row, 7),
            telefono_alt    = cell(row, 8),
            email           = email_v or None,
            web             = cell(row, 10),
            cargo           = cell(row, 11),
            nombre_contacto = nombre_completo,
        )

        db.session.add(p)
        if extra_emails:
            pending_extras.append((p, extra_emails))
        if email_v:
            existing_emails.add(email_v)
        if ruc_v:
            existing_rucs.add(ruc_v)
        inserted += 1

    try:
        # Flush para obtener IDs antes de insertar emails extra
        db.session.flush()
        for p, extras in pending_extras:
            seen = {p.email.lower()} if p.email else set()
            for em in extras:
                if em and em not in seen and em not in existing_emails:
                    db.session.add(ProspectoEmail(prospecto_id=p.id, email=em))
                    seen.add(em)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500

    return jsonify({
        "ok":       True,
        "inserted": inserted,
        "skipped":  skipped,
    })


@prospeccion_bp.route("/api/normalize-deptos", methods=["POST"])
@csrf.exempt
@login_required
@require_role("Master")
def api_normalize_deptos():
    """Normaliza departamento y tamano_empresa en todos los prospectos (solo Master)."""
    import unicodedata

    DEPTOS = [
        "Amazonas", "Áncash", "Apurímac", "Arequipa", "Ayacucho",
        "Cajamarca", "Callao", "Cusco", "Huancavelica", "Huánuco",
        "Ica", "Junín", "La Libertad", "Lambayeque", "Lima",
        "Loreto", "Madre de Dios", "Moquegua", "Pasco", "Piura",
        "Puno", "San Martín", "Tacna", "Tumbes", "Ucayali",
    ]
    TAMANOS = ["MYPE", "Pequeña", "Mediana", "Grande"]

    def _key(s):
        s = s.strip()
        s = unicodedata.normalize("NFD", s)
        s = "".join(c for c in s if unicodedata.category(c) != "Mn")
        return s.lower()

    depto_map  = {_key(d): d for d in DEPTOS}
    tamano_map = {_key(t): t for t in TAMANOS}
    depto_map.update({
        "ancash": "Áncash", "apurimac": "Apurímac", "huanuco": "Huánuco",
        "junin": "Junín", "san martin": "San Martín",
        "madre de dios": "Madre de Dios", "la libertad": "La Libertad",
    })
    tamano_map.update({
        "micro": "MYPE", "microempresa": "MYPE",
        "pequena empresa": "Pequeña", "pequena": "Pequeña",
        "mediana empresa": "Mediana", "grande empresa": "Grande",
    })

    try:
        prospectos = Prospecto.query.all()
        d_up = t_up = 0
        for p in prospectos:
            if p.departamento:
                canon = depto_map.get(_key(p.departamento))
                if canon and canon != p.departamento:
                    p.departamento = canon
                    d_up += 1
            if p.tamano_empresa:
                canon = tamano_map.get(_key(p.tamano_empresa))
                if canon and canon != p.tamano_empresa:
                    p.tamano_empresa = canon
                    t_up += 1
        db.session.commit()
        return jsonify({"ok": True, "departamentos_normalizados": d_up, "tamanos_normalizados": t_up})
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500


@prospeccion_bp.route("/api/limpiar-todo", methods=["POST"])
@login_required
@require_role("Master")
def api_limpiar_todo():
    """Elimina TODOS los prospectos y datos relacionados. Solo Master."""
    try:
        deleted_act  = db.session.query(ActividadProspecto).delete()
        from app.models.prospecto import ProspectoEmail
        deleted_em   = db.session.query(ProspectoEmail).delete()
        deleted_asig = db.session.query(AsignacionProspecto).delete()
        deleted_pros = db.session.query(Prospecto).delete()
        db.session.commit()
        total = deleted_pros + deleted_asig + deleted_act + deleted_em
        return jsonify({"ok": True, "deleted": deleted_pros,
                        "detail": {"prospectos": deleted_pros, "asignaciones": deleted_asig,
                                   "actividades": deleted_act, "emails_extra": deleted_em}})
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500



@prospeccion_bp.route("/api/eliminar-por-nombre", methods=["POST"])
@csrf.exempt
@login_required
@require_role("Master")
def api_eliminar_por_nombre():
    """Elimina todos los prospectos cuya razon social coincida con el termino dado. Solo Master."""
    data   = request.get_json(silent=True) or {}
    nombre = (data.get("nombre") or "").strip()
    if not nombre or len(nombre) < 3:
        return jsonify({"ok": False, "error": "Debe indicar al menos 3 caracteres"}), 400
    try:
        ids = [r.id for r in Prospecto.query.filter(
            Prospecto.razon_social.ilike(f"%{nombre}%")
        ).all()]
        if not ids:
            return jsonify({"ok": True, "deleted": 0, "ids": [], "mensaje": "Ningun prospecto encontrado"})
        from app.models.prospecto import ProspectoEmail, SeguimientoProspecto
        db.session.query(ActividadProspecto).filter(ActividadProspecto.prospecto_id.in_(ids)).delete(synchronize_session=False)
        db.session.query(ProspectoEmail).filter(ProspectoEmail.prospecto_id.in_(ids)).delete(synchronize_session=False)
        db.session.query(AsignacionProspecto).filter(AsignacionProspecto.prospecto_id.in_(ids)).delete(synchronize_session=False)
        db.session.query(SeguimientoProspecto).filter(SeguimientoProspecto.prospecto_id.in_(ids)).delete(synchronize_session=False)
        deleted = db.session.query(Prospecto).filter(Prospecto.id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()
        return jsonify({"ok": True, "deleted": deleted, "ids": ids})
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500


@prospeccion_bp.route("/api/bulk-campo", methods=["POST"])
@csrf.exempt
@login_required
@require_role("Master")
def api_bulk_campo():
    """Actualiza un campo en múltiples prospectos a la vez."""
    data  = request.get_json() or {}
    ids   = data.get("ids", [])
    campo = data.get("campo", "").strip()
    valor = data.get("valor")

    if not ids or campo not in _CAMPOS_EDITABLES:
        return jsonify({"ok": False, "error": "Parámetros inválidos"}), 400

    try:
        Prospecto.query.filter(Prospecto.id.in_(ids)).update(
            {campo: valor or None}, synchronize_session=False)
        db.session.commit()
        return jsonify({"ok": True, "updated": len(ids)})
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500


# ── Helpers ───────────────────────────────────────────────────────────────────

def _resolve_import_url(raw_url: str):
    """Convierte un link de Google Sheets o Drive a una URL de descarga directa en xlsx.
    Retorna None si el formato no es reconocido.
    """
    import re
    # Google Sheets: https://docs.google.com/spreadsheets/d/{ID}/...
    m = re.search(r"docs\.google\.com/spreadsheets/d/([^/\?&]+)", raw_url)
    if m:
        sheet_id = m.group(1)
        # Detectar gid (hoja específica)
        gid_m = re.search(r"[#&?]gid=(\d+)", raw_url)
        gid = f"&gid={gid_m.group(1)}" if gid_m else ""
        return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx{gid}"
    # Google Drive file: https://drive.google.com/file/d/{ID}/view
    m = re.search(r"drive\.google\.com/file/d/([^/\?&]+)", raw_url)
    if m:
        file_id = m.group(1)
        return f"https://drive.google.com/uc?export=download&id={file_id}"
    # Google Drive open: https://drive.google.com/open?id={ID}
    m = re.search(r"drive\.google\.com/open\?id=([^&]+)", raw_url)
    if m:
        return f"https://drive.google.com/uc?export=download&id={m.group(1)}"
    return None


# ── Timeline de actividades de un prospecto ───────────────────────────────────

@prospeccion_bp.route("/api/<int:pid>/timeline")
@login_required
@require_role("Master", "Trader")
@csrf.exempt
def api_timeline(pid):
    """Retorna el timeline completo de actividades + seguimientos de un prospecto."""
    p = db.get_or_404(Prospecto, pid)
    actividades = (ActividadProspecto.query
                   .filter_by(prospecto_id=pid)
                   .order_by(ActividadProspecto.creado_en.desc())
                   .limit(100).all())
    seguimientos = (SeguimientoProspecto.query
                    .filter_by(prospecto_id=pid)
                    .order_by(SeguimientoProspecto.fecha_programada.desc())
                    .all())
    return jsonify({
        "ok": True,
        "actividades":  [a.to_dict() for a in actividades],
        "seguimientos": [s.to_dict() for s in seguimientos],
    })


# ── Seguimientos (recordatorios programados) ──────────────────────────────────

@prospeccion_bp.route("/api/<int:pid>/seguimientos", methods=["GET"])
@login_required
@require_role("Master", "Trader")
def api_seguimientos_list(pid):
    """Lista seguimientos de un prospecto."""
    db.get_or_404(Prospecto, pid)
    segs = (SeguimientoProspecto.query
            .filter_by(prospecto_id=pid)
            .order_by(SeguimientoProspecto.fecha_programada)
            .all())
    return jsonify({"ok": True, "seguimientos": [s.to_dict() for s in segs]})


@prospeccion_bp.route("/api/<int:pid>/seguimientos", methods=["POST"])
@csrf.exempt
@login_required
@require_role("Master", "Trader")
def api_seguimientos_crear(pid):
    """Crea un nuevo seguimiento programado."""
    db.get_or_404(Prospecto, pid)
    data        = request.get_json(silent=True) or {}
    tipo        = data.get("tipo", "otro")
    descripcion = (data.get("descripcion") or "").strip()
    fecha_str   = (data.get("fecha_programada") or "").strip()

    if not fecha_str:
        return jsonify({"ok": False, "error": "Fecha requerida."}), 400

    try:
        for fmt in ("%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                fecha = datetime.strptime(fecha_str[:16], fmt[:len(fecha_str[:16])])
                break
            except ValueError:
                continue
        else:
            return jsonify({"ok": False, "error": "Formato de fecha inválido."}), 400
    except Exception:
        return jsonify({"ok": False, "error": "Fecha inválida."}), 400

    seg = SeguimientoProspecto(
        prospecto_id=pid,
        user_id=current_user.id,
        tipo=tipo,
        descripcion=descripcion or None,
        fecha_programada=fecha,
    )
    db.session.add(seg)
    db.session.commit()
    return jsonify({"ok": True, "seguimiento": seg.to_dict()})


@prospeccion_bp.route("/api/seguimientos/<int:sid>/completar", methods=["POST"])
@csrf.exempt
@login_required
@require_role("Master", "Trader")
def api_seguimiento_completar(sid):
    """Marca un seguimiento como completado y lo registra en el timeline."""
    seg = db.get_or_404(SeguimientoProspecto, sid)
    if not seg.completado:
        seg.completado    = True
        seg.completado_en = now_peru()
        _log_actividad(seg.prospecto_id, current_user.id,
                       seg.tipo or "seguimiento",
                       f"Seguimiento completado: {seg.descripcion or seg.tipo}",
                       canal=seg.tipo if seg.tipo in ("llamada","email","whatsapp","reunion") else "manual")
        db.session.commit()
    return jsonify({"ok": True, "seguimiento": seg.to_dict()})


@prospeccion_bp.route("/api/seguimientos/<int:sid>", methods=["DELETE"])
@csrf.exempt
@login_required
@require_role("Master", "Trader")
def api_seguimiento_eliminar(sid):
    """Elimina un seguimiento pendiente."""
    seg = db.get_or_404(SeguimientoProspecto, sid)
    db.session.delete(seg)
    db.session.commit()
    return jsonify({"ok": True})


# ── API — Registrar contacto manual por WhatsApp ──────────────────────────────

@prospeccion_bp.route("/api/<int:pid>/registrar-wa", methods=["POST"])
@csrf.exempt
@login_required
@require_role("Master", "Trader")
def api_registrar_wa(pid):
    """Registra un contacto manual por WhatsApp desde el drawer del prospecto."""
    p = db.get_or_404(Prospecto, pid)
    ahora = now_peru()
    ts = ahora.strftime("%Y-%m-%d %H:%M")

    _log_actividad(
        p.id, current_user.id,
        tipo='whatsapp',
        descripcion='WhatsApp enviado (manual desde CRM)',
        canal='whatsapp',
        resultado='Enviado',
    )

    # Avanzar estado solo si aún no ha sido contactado
    estados_sin_contacto = (None, '', 'sin_contactar')
    if p.estado_comercial in estados_sin_contacto:
        p.estado_comercial = 'contactado'
        _log_actividad(
            p.id, current_user.id,
            tipo='estado',
            descripcion='Estado cambiado a: contactado (WhatsApp manual)',
            canal='sistema',
            nuevo_estado='contactado',
        )

    db.session.commit()
    return jsonify({"ok": True, "ultima_wa": ts, "fecha_ultimo_contacto": ts, "estado_comercial": p.estado_comercial})


@prospeccion_bp.route("/api/<int:pid>/registrar-email", methods=["POST"])
@csrf.exempt
@login_required
@require_role("Master", "Trader")
def api_registrar_email(pid):
    """Registra un contacto manual por email desde el drawer del prospecto."""
    p = db.get_or_404(Prospecto, pid)
    ts = now_peru().strftime("%Y-%m-%d %H:%M")

    _log_actividad(
        p.id, current_user.id,
        tipo='email',
        descripcion='Email registrado manualmente desde CRM',
        canal='email',
        resultado='Enviado',
    )

    estados_sin_contacto = (None, '', 'sin_contactar')
    if p.estado_comercial in estados_sin_contacto:
        p.estado_comercial = 'contactado'
        _log_actividad(
            p.id, current_user.id,
            tipo='estado',
            descripcion='Estado cambiado a: contactado (email manual)',
            canal='sistema',
            nuevo_estado='contactado',
        )

    db.session.commit()
    return jsonify({"ok": True, "ultima_email": ts, "fecha_ultimo_contacto": ts, "estado_comercial": p.estado_comercial})


# ── Dashboard API — seguimientos pendientes ────────────────────────────────────

@prospeccion_bp.route("/api/seguimientos-pendientes")
@login_required
@require_role("Master", "Trader")
def api_seguimientos_pendientes():
    """Seguimientos no completados con fecha <= ahora + 24h. Para alertas de navbar."""
    limite = now_peru() + timedelta(hours=24)
    segs = (SeguimientoProspecto.query
            .filter_by(completado=False)
            .filter(SeguimientoProspecto.fecha_programada <= limite)
            .order_by(SeguimientoProspecto.fecha_programada)
            .limit(50).all())
    total_vencidos = (SeguimientoProspecto.query
                      .filter_by(completado=False)
                      .filter(SeguimientoProspecto.fecha_programada < now_peru())
                      .count())
    result = []
    for s in segs:
        p = db.session.get(Prospecto, s.prospecto_id)
        result.append({
            **s.to_dict(),
            "empresa": p.razon_social if p else "",
            "pid":     s.prospecto_id,
        })
    return jsonify({"ok": True, "pendientes": result, "vencidos": total_vencidos})


# ── API — elegibles para campaña WA (prospectos con email previo + teléfono válido) ──
@prospeccion_bp.route("/api/wa-campana-elegibles")
@csrf.exempt
def api_wa_campana_elegibles():
    """
    Retorna prospectos con teléfono celular peruano válido que ya fueron
    contactados por email (canal='email' en actividades O estado implica contacto previo).
    Auth: X-API-Key = CRM_API_KEY.
    """
    import re as _re
    api_key  = request.headers.get('X-API-Key', '')
    crm_key  = os.environ.get('CRM_API_KEY', 'qoricash_crm_2026')
    if api_key != crm_key:
        return jsonify({'ok': False, 'error': 'Unauthorized'}), 401

    def _valid_mobile(val):
        if not val: return None
        d = _re.sub(r'\D', '', str(val))
        if d.startswith('51') and len(d) == 11: d = d[2:]
        if len(d) == 9 and d.startswith('9'): return d
        return None

    # IDs que tienen al menos 1 actividad de canal email
    email_act_ids = (
        db.session.query(ActividadProspecto.prospecto_id)
        .filter(or_(
            ActividadProspecto.canal == 'email',
            ActividadProspecto.tipo  == 'email',
        ))
        .distinct()
        .subquery()
    )

    # IDs con actividad WA en las últimas 24h (respetar ventana Meta)
    cutoff_wa = now_peru() - timedelta(hours=24)
    wa_recientes_ids = (
        db.session.query(ActividadProspecto.prospecto_id)
        .filter(
            ActividadProspecto.canal == 'whatsapp',
            ActividadProspecto.creado_en >= cutoff_wa,
        )
        .distinct()
        .subquery()
    )

    # Estados que implican contacto previo por email
    _ESTADOS_CONTACTADOS = (
        'contactado', 'interesado', 'negociando', 'cliente',
        'P1', 'P2', 'P3', 'P4', 'presentado', 'precio_enviado', 'seguimiento',
    )
    _EXCLUIR = ('no_contactar',)

    prospectos = (
        Prospecto.query
        .filter(
            or_(
                Prospecto.id.in_(email_act_ids),
                Prospecto.estado_comercial.in_(_ESTADOS_CONTACTADOS),
                Prospecto.tipo_ultimo_envio.isnot(None),
            ),
            ~Prospecto.estado_comercial.in_(_EXCLUIR),
            ~Prospecto.id.in_(wa_recientes_ids),   # excluir contactados por WA < 24h
        )
        .all()
    )

    resultado = []
    seen_phones = set()
    for p in prospectos:
        phone = None
        for campo in (p.contacto_wa, p.telefono, p.telefono_alt, p.telefono_3, p.telefono_4):
            phone = _valid_mobile(campo)
            if phone:
                break
        if not phone or phone in seen_phones:
            continue
        seen_phones.add(phone)

        empresa = (p.razon_social or '').strip()
        nombre  = (p.nombre_contacto or empresa or '').strip() or 'Estimado/a'

        resultado.append({
            'id':      p.id,
            'phone':   phone,          # 9 dígitos sin prefijo
            'nombre':  nombre,
            'empresa': empresa,
            'estado':  p.estado_comercial or '',
            'email':   p.email or '',
        })

    return jsonify({'ok': True, 'total': len(resultado), 'prospectos': resultado})


# ── API — Campaña Nuevos Prospectos ───────────────────────────────────────────

_ESTADOS_EXCLUIDOS_CAMPANA = {
    'no_contactar', 'correo_invalido', 'correo_rebotado', 'rebote',
    'cliente', 'no_interesado', 'inactivo',
}


@prospeccion_bp.route("/api/campana-nuevos-elegibles")
@csrf.exempt
def api_campana_nuevos_elegibles():
    """
    Retorna prospectos sin contactar elegibles para Campaña Nuevos Prospectos.
    Auth: X-API-Key = CRM_API_KEY env var.
    """
    api_key = request.headers.get('X-API-Key', '')
    crm_key = os.environ.get('CRM_API_KEY', 'qoricash_crm_2026')
    if api_key != crm_key:
        return jsonify({'ok': False, 'error': 'Unauthorized'}), 401

    prospectos = (
        Prospecto.query
        .filter(Prospecto.email.isnot(None))
        .filter(Prospecto.email != '')
        .filter(
            or_(
                Prospecto.fecha_ultimo_contacto.is_(None),
                Prospecto.fecha_ultimo_contacto == '',
            )
        )
        .filter(
            or_(
                Prospecto.fecha_primer_contacto.is_(None),
                Prospecto.fecha_primer_contacto == '',
            )
        )
        .filter(
            or_(
                Prospecto.estado_comercial.is_(None),
                Prospecto.estado_comercial == '',
                ~Prospecto.estado_comercial.in_(_ESTADOS_EXCLUIDOS_CAMPANA),
            )
        )
        .order_by(Prospecto.id.asc())
        .all()
    )

    result = []
    for p in prospectos:
        result.append({
            'id':              p.id,
            'razon_social':    p.razon_social or '',
            'nombre_contacto': p.nombre_contacto or '',
            'email':           p.email,
            'email_alt':       p.email_alt or '',
            'num_contactos':   p.num_contactos or 0,
        })

    return jsonify({'ok': True, 'total': len(result), 'prospectos': result})


@prospeccion_bp.route("/api/campana-precios-elegibles")
@csrf.exempt
def api_campana_precios_elegibles():
    """
    Retorna prospectos ya contactados elegibles para Campaña de Precios.
    Elegibles: fecha_primer_contacto != NULL, no excluidos, con email válido.
    El cutoff de 5 días hábiles se evalúa en el script cliente.
    Auth: X-API-Key = CRM_API_KEY env var.
    """
    api_key = request.headers.get('X-API-Key', '')
    crm_key = os.environ.get('CRM_API_KEY', 'qoricash_crm_2026')
    if api_key != crm_key:
        return jsonify({'ok': False, 'error': 'Unauthorized'}), 401

    prospectos = (
        Prospecto.query
        .filter(Prospecto.email.isnot(None))
        .filter(Prospecto.email != '')
        .filter(Prospecto.fecha_primer_contacto.isnot(None))
        .filter(Prospecto.fecha_primer_contacto != '')
        .filter(
            or_(
                Prospecto.estado_comercial.is_(None),
                Prospecto.estado_comercial == '',
                ~Prospecto.estado_comercial.in_(_ESTADOS_EXCLUIDOS_CAMPANA),
            )
        )
        .order_by(Prospecto.fecha_ultimo_contacto.asc())
        .all()
    )

    result = []
    for p in prospectos:
        result.append({
            'id':                   p.id,
            'razon_social':         p.razon_social or '',
            'email':                p.email,
            'fecha_ultimo_contacto': p.fecha_ultimo_contacto or '',
        })

    return jsonify({'ok': True, 'total': len(result), 'prospectos': result})


@prospeccion_bp.route("/api/campana-registrar-envio", methods=["POST"])
@csrf.exempt
def api_campana_registrar_envio():
    """
    Registra un envío de Campaña Nuevos Prospectos.
    Actualiza el prospecto y loggea la actividad.
    Auth: X-API-Key = CRM_API_KEY env var.
    Body: {
        prospecto_id, remitente, bandeja, tipo,
        ok (bool), fecha_envio (YYYY-MM-DD HH:MM),
        fecha_proximo_contacto (YYYY-MM-DD),
        gmail_message_id (str, opcional),
        error (str, opcional)
    }
    """
    api_key = request.headers.get('X-API-Key', '')
    crm_key = os.environ.get('CRM_API_KEY', 'qoricash_crm_2026')
    if api_key != crm_key:
        return jsonify({'ok': False, 'error': 'Unauthorized'}), 401

    data = request.get_json(force=True)
    pid  = data.get('prospecto_id')
    if not pid:
        return jsonify({'ok': False, 'error': 'prospecto_id requerido'}), 400

    p = db.session.get(Prospecto, int(pid))
    if not p:
        return jsonify({'ok': False, 'error': 'Prospecto no encontrado'}), 404

    enviado_ok  = data.get('ok', True)
    remitente   = (data.get('remitente') or '').strip()
    bandeja     = (data.get('bandeja') or remitente).strip()
    tipo        = (data.get('tipo') or 'presentacion').strip()
    fecha_envio = (data.get('fecha_envio') or now_peru().strftime('%Y-%m-%d %H:%M')).strip()
    fecha_prox  = (data.get('fecha_proximo_contacto') or '').strip()
    msg_id      = (data.get('gmail_message_id') or '').strip()
    error_txt   = (data.get('error') or '').strip()

    # Obtener usuario sistema para actividad
    bot_user = (
        User.query.filter_by(role='Master').order_by(User.id.asc()).first()
        or User.query.order_by(User.id.asc()).first()
    )

    if not enviado_ok:
        if bot_user:
            act = ActividadProspecto(
                prospecto_id=p.id,
                user_id=bot_user.id,
                tipo='sistema',
                canal='email',
                bandeja=bandeja,
                descripcion=f'Campaña Nuevos Prospectos — error de envío: {error_txt}',
                resultado='Error',
            )
            db.session.add(act)
            db.session.commit()
        return jsonify({'ok': True, 'updated': False})

    _NO_DEGRADAR = {'negociando', 'negociacion', 'P3', 'cliente', 'P4'}
    if (p.estado_comercial or '') not in _NO_DEGRADAR:
        p.estado_comercial = 'presentado'

    p.tipo_ultimo_envio     = tipo
    p.remitente             = remitente
    p.bandeja               = bandeja
    p.fecha_ultimo_contacto = fecha_envio
    p.num_contactos         = (p.num_contactos or 0) + 1
    if not p.fecha_primer_contacto:
        p.fecha_primer_contacto = fecha_envio
    if fecha_prox:
        p.fecha_proximo_contacto = fecha_prox

    if bot_user:
        desc = 'Campaña Nuevos Prospectos — correo de presentación enviado'
        if msg_id:
            desc += f' [ID: {msg_id}]'
        act = ActividadProspecto(
            prospecto_id=p.id,
            user_id=bot_user.id,
            tipo='email',
            canal='email',
            bandeja=bandeja,
            descripcion=desc,
            resultado='Enviado',
            nuevo_estado='presentado',
        )
        db.session.add(act)

    db.session.commit()
    return jsonify({'ok': True, 'updated': True})


# ── API — Limpiar Bandejas (integración comercial) ────────────────────────────

_BOUNCE_ESTADOS_NO_DEGRADAR = {'negociando', 'negociacion', 'P3', 'cliente', 'P4'}
_INTERES_ESTADOS_NO_SOBREESCRIBIR = {
    'interesado', 'negociando', 'negociacion', 'cliente',
    'precio_enviado', 'P2', 'P3', 'P4',
}


def _find_prospecto_by_email(email_lower: str):
    """Busca un Prospecto por email en todos los campos de email disponibles."""
    p = Prospecto.query.filter(
        or_(
            func.lower(Prospecto.email)     == email_lower,
            func.lower(Prospecto.email_alt) == email_lower,
            func.lower(Prospecto.email_3)   == email_lower,
            func.lower(Prospecto.email_4)   == email_lower,
        )
    ).first()
    if not p:
        pe = ProspectoEmail.query.filter(
            func.lower(ProspectoEmail.email) == email_lower,
            ProspectoEmail.activo == True,
        ).first()
        if pe:
            p = db.session.get(Prospecto, pe.prospecto_id)
    return p


@prospeccion_bp.route("/api/todos-emails")
@csrf.exempt
def api_todos_emails():
    """
    Retorna todos los prospectos con email (todos los estados).
    Usado por el script de validación MX masiva.
    Auth: X-API-Key = CRM_API_KEY env var.
    """
    api_key = request.headers.get('X-API-Key', '')
    crm_key = os.environ.get('CRM_API_KEY', 'qoricash_crm_2026')
    if api_key != crm_key:
        return jsonify({'ok': False, 'error': 'Unauthorized'}), 401

    # Excluir los que ya están marcados como inválidos o rebotados
    EXCLUIR = {'correo_invalido', 'correo_rebotado', 'Dominio inválido'}
    prospectos = (
        Prospecto.query
        .filter(Prospecto.email.isnot(None))
        .filter(Prospecto.email != '')
        .filter(
            or_(
                Prospecto.estado_email.is_(None),
                Prospecto.estado_email == '',
                ~Prospecto.estado_email.in_(list(EXCLUIR)),
            )
        )
        .order_by(Prospecto.id.asc())
        .all()
    )

    result = [{'id': p.id, 'email': p.email, 'estado_email': p.estado_email or ''} for p in prospectos]
    return jsonify({'ok': True, 'total': len(result), 'prospectos': result})


@prospeccion_bp.route("/api/limpiar-bandejas/registrar-batch", methods=["POST"])
@csrf.exempt
def api_limpiar_bandejas_registrar_batch():
    """
    Registra en batch todos los eventos detectados durante la limpieza de bandejas.
    Auth: X-API-Key = CRM_API_KEY env var.

    Optimizado: bulk-lookup de emails (2 queries totales por batch) en lugar de
    1 query por email, para evitar timeouts en Render con batches grandes.
    """
    api_key = request.headers.get('X-API-Key', '')
    crm_key = os.environ.get('CRM_API_KEY', 'qoricash_crm_2026')
    if api_key != crm_key:
        return jsonify({'ok': False, 'error': 'Unauthorized'}), 401

    data    = request.get_json(force=True)
    eventos = data.get('eventos', [])
    if not eventos:
        return jsonify({'ok': True, 'procesados': 0, 'actualizados': 0})

    # ── Pre-procesar eventos ────────────────────────────────────────────────────
    evs_parsed = []
    emails_lower = set()
    for ev in eventos:
        email = (ev.get('email') or '').strip().lower()
        if not email or '@' not in email:
            continue
        emails_lower.add(email)
        evs_parsed.append({
            'email':       email,
            'tipo':        (ev.get('tipo_evento') or '').strip(),
            'tipo_rebote': (ev.get('tipo_rebote') or 'hard').strip(),
            'motivo':      (ev.get('motivo') or '').strip()[:200],
            'cuenta':      (ev.get('cuenta_origen') or '').strip(),
            'fecha':       (ev.get('fecha') or now_peru().strftime('%Y-%m-%d %H:%M')).strip(),
            'asunto':      (ev.get('asunto') or '').strip()[:150],
            'obs':         (ev.get('observaciones') or '').strip()[:300],
        })

    if not evs_parsed:
        return jsonify({'ok': True, 'procesados': 0, 'actualizados': 0})

    # ── Bulk lookup: 2 queries para todos los emails del batch ─────────────────
    email_list = list(emails_lower)

    # Query 1: campos directos de Prospecto
    prospectos_directos = Prospecto.query.filter(
        or_(
            func.lower(Prospecto.email).in_(email_list),
            func.lower(Prospecto.email_alt).in_(email_list),
            func.lower(Prospecto.email_3).in_(email_list),
            func.lower(Prospecto.email_4).in_(email_list),
        )
    ).all()

    # Construir mapa email → Prospecto
    email_map: dict = {}
    for p in prospectos_directos:
        for campo in [p.email, p.email_alt, p.email_3, p.email_4]:
            if campo:
                k = campo.strip().lower()
                if k in emails_lower and k not in email_map:
                    email_map[k] = p

    # Query 2: ProspectoEmail para los que no se encontraron
    emails_faltantes = [e for e in email_list if e not in email_map]
    if emails_faltantes:
        pes = ProspectoEmail.query.filter(
            func.lower(ProspectoEmail.email).in_(emails_faltantes),
            ProspectoEmail.activo == True,
        ).all()
        if pes:
            ids_extra = list({pe.prospecto_id for pe in pes})
            prospectos_extra = Prospecto.query.filter(Prospecto.id.in_(ids_extra)).all()
            id_map = {p.id: p for p in prospectos_extra}
            for pe in pes:
                k = pe.email.strip().lower()
                if k not in email_map and pe.prospecto_id in id_map:
                    email_map[k] = id_map[pe.prospecto_id]

    # ── Usuario sistema para actividades ───────────────────────────────────────
    bot_user = (
        User.query.filter_by(role='Master').order_by(User.id.asc()).first()
        or User.query.order_by(User.id.asc()).first()
    )
    bot_id = bot_user.id if bot_user else None

    procesados              = len(evs_parsed)
    prospectos_encontrados  = 0
    prospectos_actualizados = 0
    fuera_de_crm            = []

    # ── Procesar eventos usando el mapa en memoria ─────────────────────────────
    for ev in evs_parsed:
        email       = ev['email']
        tipo        = ev['tipo']
        tipo_rebote = ev['tipo_rebote']
        motivo      = ev['motivo']
        cuenta      = ev['cuenta']
        fecha       = ev['fecha']
        asunto      = ev['asunto']
        obs         = ev['obs']

        p = email_map.get(email)
        if not p:
            fuera_de_crm.append(email)
            continue

        prospectos_encontrados += 1
        cambios = False

        if tipo == 'bounce':
            p.estado_email = 'correo_rebotado'
            if tipo_rebote == 'hard':
                if (p.estado_comercial or '') not in _BOUNCE_ESTADOS_NO_DEGRADAR:
                    p.estado_comercial = 'correo_rebotado'
            nota = f'[REBOTE {tipo_rebote.upper()} {fecha}] {motivo or asunto}'
            p.notas = ((p.notas or '') + '\n' + nota).strip()[-2000:]
            p.fecha_ultimo_contacto = fecha
            if bot_id:
                db.session.add(ActividadProspecto(
                    prospecto_id=p.id, user_id=bot_id,
                    tipo='bounce', canal='email', bandeja=cuenta,
                    descripcion=nota,
                    resultado=f'Rebote {tipo_rebote}',
                    nuevo_estado='correo_rebotado' if tipo_rebote == 'hard' else p.estado_comercial,
                ))
            cambios = True

        elif tipo == 'no_contactar':
            p.estado_comercial      = 'no_contactar'
            p.fecha_ultimo_contacto = fecha
            nota = f'[NO CONTACTAR {fecha}] {asunto or obs}'
            p.notas = ((p.notas or '') + '\n' + nota).strip()[-2000:]
            if bot_id:
                db.session.add(ActividadProspecto(
                    prospecto_id=p.id, user_id=bot_id,
                    tipo='email', canal='email', bandeja=cuenta,
                    descripcion=nota,
                    resultado='No desea ser contactado',
                    nuevo_estado='no_contactar',
                ))
            cambios = True

        elif tipo == 'respuesta_positiva':
            p.fecha_ultimo_contacto = fecha
            if (p.estado_comercial or '') not in _INTERES_ESTADOS_NO_SOBREESCRIBIR:
                p.estado_comercial = 'interesado'
            desc = f'[RESPUESTA POSITIVA {fecha}] {asunto} — {obs}'
            if bot_id:
                db.session.add(ActividadProspecto(
                    prospecto_id=p.id, user_id=bot_id,
                    tipo='email', canal='email', bandeja=cuenta,
                    descripcion=desc,
                    resultado='Respuesta positiva',
                    nuevo_estado=p.estado_comercial,
                ))
            cambios = True

        elif tipo == 'consulta_tc':
            p.fecha_ultimo_contacto = fecha
            if (p.estado_comercial or '') not in _INTERES_ESTADOS_NO_SOBREESCRIBIR:
                p.estado_comercial = 'interesado'
            desc = f'[CONSULTA TC {fecha}] {asunto} — {obs}'
            if bot_id:
                db.session.add(ActividadProspecto(
                    prospecto_id=p.id, user_id=bot_id,
                    tipo='email', canal='email', bandeja=cuenta,
                    descripcion=desc,
                    resultado='Consulta tipo de cambio',
                    nuevo_estado=p.estado_comercial,
                ))
            cambios = True

        elif tipo == 'oportunidad':
            p.fecha_ultimo_contacto = fecha
            desc = f'[OPORTUNIDAD {fecha}] {asunto} — {obs}'
            if bot_id:
                db.session.add(ActividadProspecto(
                    prospecto_id=p.id, user_id=bot_id,
                    tipo='email', canal='email', bandeja=cuenta,
                    descripcion=desc,
                    resultado='Oportunidad comercial',
                ))
            cambios = True

        elif tipo == 'mx_invalido':
            p.estado_email = 'correo_invalido'
            nota = f'[MX INVALIDO {fecha}] Dominio sin registro MX — {motivo or obs}'
            p.notas = ((p.notas or '') + '\n' + nota).strip()[-2000:]
            if bot_id:
                db.session.add(ActividadProspecto(
                    prospecto_id=p.id, user_id=bot_id,
                    tipo='sistema', canal='email', bandeja='validacion_mx',
                    descripcion=nota,
                    resultado='Dominio sin MX — correo inválido',
                    nuevo_estado=p.estado_comercial,
                ))
            cambios = True

        elif tipo == 'auto_reply':
            desc = f'[AUTO-REPLY {fecha}] {obs or asunto}'
            p.notas = ((p.notas or '') + '\n' + desc).strip()[-2000:]
            if bot_id:
                db.session.add(ActividadProspecto(
                    prospecto_id=p.id, user_id=bot_id,
                    tipo='sistema', canal='email', bandeja=cuenta,
                    descripcion=desc,
                    resultado='Auto-respuesta detectada',
                ))
            cambios = True

        if cambios:
            prospectos_actualizados += 1

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500

    return jsonify({
        'ok':                      True,
        'procesados':              procesados,
        'prospectos_encontrados':  prospectos_encontrados,
        'prospectos_actualizados': prospectos_actualizados,
        'fuera_de_crm':            fuera_de_crm,
        'fuera_de_crm_count':      len(fuera_de_crm),
    })


def _verificar_acceso(p):
    """Trader solo puede ver sus prospectos asignados. Llama abort(403) si no tiene acceso."""
    if current_user.role in ('Master', 'Presidente de Negocios'):
        return
    asig = next((a for a in p.asignaciones if a.trader_id == current_user.id and a.activo), None)
    if not asig:
        from flask import abort
        abort(403)
