"""
Rutas de Operaciones para QoriCash Trading V2
"""
from flask import Blueprint, render_template, request, jsonify, send_file
from flask_login import login_required, current_user
from app.services.operation_service import OperationService
from app.socketio_events import emit_operation_event
from app.services.file_service import FileService
from app.services.notification_service import NotificationService
from app.utils.decorators import require_role
from app.utils.formatters import now_peru
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from datetime import datetime

operations_bp = Blueprint('operations', __name__)


def get_operator_name(operator_id):
    """
    Helper para obtener nombre de operador en templates

    Args:
        operator_id: ID del operador

    Returns:
        str: Nombre del operador o 'N/A'
    """
    if not operator_id:
        return 'N/A'

    from app.models.user import User
    operator = User.query.get(operator_id)
    return operator.username if operator else 'N/A'


# Registrar función helper en el contexto de templates
@operations_bp.app_template_filter('get_operator_name')
def get_operator_name_filter(operator_id):
    return get_operator_name(operator_id)


@operations_bp.route('/')
@operations_bp.route('/list')
@login_required
def operations_list():
    """
    Página de listado de operaciones

    Muestra solo las operaciones del día actual
    Todos los roles (Master, Trader, Operador) ven las operaciones de hoy
    """
    operations = OperationService.get_today_operations()
    return render_template('operations/list.html',
                         user=current_user,
                         operations=operations,
                         get_operator_name=get_operator_name)


@operations_bp.route('/create')
@login_required
@require_role('Master', 'Trader')
def create_page():
    """
    Página de creación de operación

    Los clientes se buscan dinámicamente desde el modal de búsqueda
    """
    return render_template('operations/create.html',
                         user=current_user)


@operations_bp.route('/history')
@login_required
def history():
    """
    Página de historial de operaciones

    Muestra todas las operaciones de todos los estados y fechas
    Disponible para todos los roles (Master, Trader, Operador)
    """
    operations = OperationService.get_all_operations(include_relations=False)
    return render_template('operations/history.html',
                         user=current_user,
                         operations=operations)


def get_bank_account_info(operation, account_number):
    """
    Obtener información completa de una cuenta bancaria (Banco - Número)

    Args:
        operation: Objeto Operation
        account_number: Número de cuenta a buscar

    Returns:
        str: Formato "BANCO - NUMERO" o "-" si no existe
    """
    if not account_number or not operation.client:
        return '-'

    bank_accounts = operation.client.bank_accounts or []
    for account in bank_accounts:
        if account.get('account_number') == account_number:
            bank_name = account.get('bank_name', 'N/A')
            return f"{bank_name} - {account_number}"

    # Si no se encuentra en las cuentas del cliente, solo retornar el número
    return account_number


@operations_bp.route('/api/export_today')
@login_required
def export_today():
    """
    API: Exportar operaciones del día actual a Excel

    Las columnas varían según el rol:
    - Master y Operador: Incluyen columna de Usuario (email)
    - Trader: No incluye columna de Usuario
    """
    operations = OperationService.get_today_operations()

    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Operaciones del Día"

    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Definir encabezados según el rol
    if current_user.role in ['Master', 'Operador']:
        headers = ['ID OP.', 'DOCUMENTO', 'CLIENTE', 'USD', 'T.C.', 'PEN', 'CUENTA CARGO', 'CUENTA DESTINO', 'ESTADO', 'FECHA', 'USUARIO']
    else:
        headers = ['ID OP.', 'DOCUMENTO', 'CLIENTE', 'USD', 'T.C.', 'PEN', 'CUENTA CARGO', 'CUENTA DESTINO', 'ESTADO', 'FECHA']

    # Escribir encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Escribir datos
    for row_num, op in enumerate(operations, 2):
        ws.cell(row=row_num, column=1, value=op.operation_id)
        ws.cell(row=row_num, column=2, value=op.client.dni if op.client else '-')
        ws.cell(row=row_num, column=3, value=op.client.full_name if op.client else '-')
        ws.cell(row=row_num, column=4, value=float(op.amount_usd))
        ws.cell(row=row_num, column=5, value=float(op.exchange_rate))
        ws.cell(row=row_num, column=6, value=float(op.amount_pen))
        ws.cell(row=row_num, column=7, value=get_bank_account_info(op, op.source_account))
        ws.cell(row=row_num, column=8, value=get_bank_account_info(op, op.destination_account))
        ws.cell(row=row_num, column=9, value=op.status)
        ws.cell(row=row_num, column=10, value=op.created_at.strftime('%d/%m/%Y %H:%M') if op.created_at else '-')

        # Solo agregar columna de usuario para Master y Operador
        if current_user.role in ['Master', 'Operador']:
            ws.cell(row=row_num, column=11, value=op.user.email if op.user else '-')

    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 35  # CUENTA CARGO (más ancho para banco + número)
    ws.column_dimensions['H'].width = 35  # CUENTA DESTINO (más ancho para banco + número)
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 18
    if current_user.role in ['Master', 'Operador']:
        ws.column_dimensions['K'].width = 30

    # Guardar en memoria
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Nombre del archivo con fecha actual
    filename = f"operaciones_del_dia_{now_peru().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@operations_bp.route('/api/export_history')
@login_required
def export_history():
    """
    API: Exportar historial de operaciones a Excel con filtro de fechas

    Query params opcionales:
        start_date: Fecha inicio (formato: YYYY-MM-DD)
        end_date: Fecha fin (formato: YYYY-MM-DD)

    Las columnas varían según el rol:
    - Master y Operador: Incluyen columna de Usuario (email)
    - Trader: No incluye columna de Usuario
    """
    from app.models.operation import Operation
    from datetime import datetime
    from app.utils.formatters import now_peru
    from sqlalchemy.orm import joinedload

    # Obtener parámetros de filtro de fechas
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    # Construir query base con EAGER LOADING para evitar N+1
    query = Operation.query.options(
        joinedload(Operation.client),
        joinedload(Operation.user)
    )

    # Aplicar filtros de fecha si existen
    if start_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            query = query.filter(Operation.created_at >= start)
        except ValueError:
            pass

    if end_date:
        try:
            # Agregar 23:59:59 al final del día
            end = datetime.strptime(end_date, '%Y-%m-%d')
            end = end.replace(hour=23, minute=59, second=59)
            query = query.filter(Operation.created_at <= end)
        except ValueError:
            pass

    # Ordenar por fecha descendente
    operations = query.order_by(Operation.created_at.desc()).all()

    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial de Operaciones"

    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Definir encabezados según el rol
    if current_user.role in ['Master', 'Operador']:
        headers = ['ID OP.', 'DOCUMENTO', 'CLIENTE', 'USD', 'T.C.', 'PEN', 'CUENTA CARGO', 'CUENTA DESTINO', 'ESTADO', 'FECHA', 'USUARIO']
    else:
        headers = ['ID OP.', 'DOCUMENTO', 'CLIENTE', 'USD', 'T.C.', 'PEN', 'CUENTA CARGO', 'CUENTA DESTINO', 'ESTADO', 'FECHA']

    # Escribir encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Escribir datos
    for row_num, op in enumerate(operations, 2):
        ws.cell(row=row_num, column=1, value=op.operation_id)
        ws.cell(row=row_num, column=2, value=op.client.dni if op.client else '-')
        ws.cell(row=row_num, column=3, value=op.client.full_name if op.client else '-')
        ws.cell(row=row_num, column=4, value=float(op.amount_usd))
        ws.cell(row=row_num, column=5, value=float(op.exchange_rate))
        ws.cell(row=row_num, column=6, value=float(op.amount_pen))
        ws.cell(row=row_num, column=7, value=get_bank_account_info(op, op.source_account))
        ws.cell(row=row_num, column=8, value=get_bank_account_info(op, op.destination_account))
        ws.cell(row=row_num, column=9, value=op.status)
        ws.cell(row=row_num, column=10, value=op.created_at.strftime('%d/%m/%Y %H:%M') if op.created_at else '-')

        # Solo agregar columna de usuario para Master y Operador
        if current_user.role in ['Master', 'Operador']:
            ws.cell(row=row_num, column=11, value=op.user.email if op.user else '-')

    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 35  # CUENTA CARGO (más ancho para banco + número)
    ws.column_dimensions['H'].width = 35  # CUENTA DESTINO (más ancho para banco + número)
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 18
    if current_user.role in ['Master', 'Operador']:
        ws.column_dimensions['K'].width = 30

    # Guardar en memoria
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Nombre del archivo con fecha actual
    filename = f"historial_operaciones_{now_peru().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@operations_bp.route('/api/list')
@login_required
def api_list():
    """
    API: Listar operaciones del día actual

    Query params:
        status: Filtrar por estado (opcional)
        client_id: Filtrar por cliente (opcional)
        all: Si es 'true', devuelve todas las operaciones (para historial)
    """
    status = request.args.get('status')
    client_id = request.args.get('client_id', type=int)
    show_all = request.args.get('all', 'false').lower() == 'true'

    if show_all:
        # Para el historial, mostrar todas las operaciones
        if status:
            operations = OperationService.get_operations_by_status(status)
        elif client_id:
            operations = OperationService.get_operations_by_client(client_id)
        else:
            operations = OperationService.get_all_operations(include_relations=True)
            # Ya viene como diccionarios
            return jsonify({'success': True, 'operations': operations})
    else:
        # Por defecto, solo operaciones del día actual
        operations = OperationService.get_today_operations()
        # get_today_operations devuelve objetos, necesitamos convertir a dict
        return jsonify({
            'success': True,
            'operations': [op.to_dict(include_relations=True) for op in operations]
        })

    return jsonify({
        'success': True,
        'operations': [op.to_dict(include_relations=True) for op in operations]
    })


@operations_bp.route('/api/create', methods=['POST'])
@login_required
@require_role('Master', 'Trader')
def create_operation():
    """
    API: Crear nueva operación
    
    POST JSON:
        client_id: int (required)
        operation_type: string (required) - 'Compra' o 'Venta'
        amount_usd: float (required)
        exchange_rate: float (required)
        source_account: string (optional)
        destination_account: string (optional)
        notes: string (optional)
    """
    data = request.get_json()
    
    # Crear operación
    success, message, operation = OperationService.create_operation(
        current_user=current_user,
        client_id=data.get('client_id'),
        operation_type=data.get('operation_type'),
        amount_usd=data.get('amount_usd'),
        exchange_rate=data.get('exchange_rate'),
        source_account=data.get('source_account'),
        destination_account=data.get('destination_account'),
        notes=data.get('notes')
    )
    
    if success:
        # Notificar creación
        NotificationService.notify_new_operation(operation)
        NotificationService.notify_dashboard_update()
        NotificationService.notify_position_update()
        
        # Emitir evento Socket.IO para tiempo real
        emit_operation_event('created', operation.to_dict(include_relations=True))
        
        return jsonify({
            'success': True,
            'message': message,
            'operation': operation.to_dict(include_relations=True)
        }), 201
    else:
        return jsonify({
            'success': False,
            'message': message
        }), 400


@operations_bp.route('/api/update_status/<int:operation_id>', methods=['PATCH'])
@login_required
def update_status(operation_id):
    """
    API: Actualizar estado de operación
    
    PATCH JSON:
        status: string (required) - 'Pendiente', 'En proceso', 'Completada', 'Cancelado'
        notes: string (optional)
    """
    data = request.get_json()
    new_status = data.get('status')
    notes = data.get('notes')
    
    if not new_status:
        return jsonify({
            'success': False,
            'message': 'El estado es requerido'
        }), 400
    
    # Obtener operación para guardar estado anterior
    operation = OperationService.get_operation_by_id(operation_id)
    old_status = operation.status if operation else None
    
    # Actualizar estado
    success, message, operation = OperationService.update_operation_status(
        current_user=current_user,
        operation_id=operation_id,
        new_status=new_status,
        notes=notes
    )
    
    if success:
        # Notificar según el nuevo estado
        NotificationService.notify_operation_updated(operation, old_status)

        if new_status == 'Completada':
            NotificationService.notify_operation_completed(operation)

        NotificationService.notify_dashboard_update()
        NotificationService.notify_position_update()
        
        return jsonify({
            'success': True,
            'message': message,
            'operation': operation.to_dict(include_relations=True)
        })
    else:
        return jsonify({
            'success': False,
            'message': message
        }), 400


@operations_bp.route('/api/upload_proof/<int:operation_id>', methods=['POST'])
@login_required
def upload_proof(operation_id):
    """
    API: Subir comprobantes de operación
    
    Form data:
        payment_proof: file (optional)
        operator_proof: file (optional)
    """
    operation = OperationService.get_operation_by_id(operation_id)
    if not operation:
        return jsonify({'success': False, 'message': 'Operación no encontrada'}), 404
    
    file_service = FileService()
    
    payment_proof_url = None
    operator_proof_url = None
    
    # Subir comprobante de pago (cliente)
    if 'payment_proof' in request.files:
        file = request.files['payment_proof']
        success, message, url = file_service.upload_payment_proof(file, operation.operation_id)
        if success:
            payment_proof_url = url
        else:
            return jsonify({'success': False, 'message': f'Error comprobante de pago: {message}'}), 400
    
    # Subir comprobante del operador
    if 'operator_proof' in request.files:
        file = request.files['operator_proof']
        success, message, url = file_service.upload_operator_proof(file, operation.operation_id)
        if success:
            operator_proof_url = url
        else:
            return jsonify({'success': False, 'message': f'Error comprobante del operador: {message}'}), 400
    
    # Actualizar operación con URLs
    if payment_proof_url or operator_proof_url:
        success, message, operation = OperationService.update_operation_proofs(
            current_user=current_user,
            operation_id=operation_id,
            payment_proof_url=payment_proof_url,
            operator_proof_url=operator_proof_url
        )
        
        if success:
            return jsonify({
                'success': True,
                'message': message,
                'operation': operation.to_dict(include_relations=True)
            })
        else:
            return jsonify({'success': False, 'message': message}), 400
    
    return jsonify({'success': False, 'message': 'No se seleccionó ningún archivo'}), 400


@operations_bp.route('/api/cancel/<int:operation_id>', methods=['POST'])
@login_required
@require_role('Master', 'Trader')
def cancel_operation(operation_id):
    """
    API: Cancelar operación
    
    POST JSON:
        reason: string (required)
    """
    data = request.get_json()
    reason = data.get('reason', '').strip()
    
    if not reason:
        return jsonify({
            'success': False,
            'message': 'La razón de cancelación es requerida'
        }), 400
    
    # Cancelar operación
    success, message, operation = OperationService.cancel_operation(
        current_user=current_user,
        operation_id=operation_id,
        reason=reason
    )
    
    if success:
        # Notificar cancelación
        NotificationService.notify_operation_canceled(operation, reason)
        NotificationService.notify_dashboard_update()
        NotificationService.notify_position_update()
        
        return jsonify({
            'success': True,
            'message': message,
            'operation': operation.to_dict(include_relations=True)
        })
    else:
        return jsonify({
            'success': False,
            'message': message
        }), 400


@operations_bp.route('/api/<string:operation_id>')
@login_required
def get_operation(operation_id):
    """
    API: Obtener detalles de una operación
    
    Args:
        operation_id: Puede ser ID numérico o operation_id (EXP-XXXX)
    """
    # Intentar como ID numérico
    try:
        op_id = int(operation_id)
        operation = OperationService.get_operation_by_id(op_id)
    except ValueError:
        # Buscar por operation_id string
        operation = OperationService.get_operation_by_operation_id(operation_id)
    
    if not operation:
        return jsonify({
            'success': False,
            'message': 'Operación no encontrada'
        }), 404
    
    return jsonify({
        'success': True,
        'operation': operation.to_dict(include_relations=True)
    })


@operations_bp.route('/api/today')
@login_required
def get_today_operations():
    """
    API: Obtener operaciones de hoy
    """
    operations = OperationService.get_today_operations()
    return jsonify({
        'success': True,
        'operations': [op.to_dict(include_relations=True) for op in operations]
    })


@operations_bp.route('/api/for_operator')
@login_required
@require_role('Operador')
def get_for_operator():
    """
    API: Obtener operaciones para operador (Pendientes y En proceso)
    """
    operations = OperationService.get_operations_for_operator()
    return jsonify({
        'success': True,
        'operations': [op.to_dict(include_relations=True) for op in operations]
    })


# === NUEVOS ENDPOINTS ===

@operations_bp.route('/api/update/<int:operation_id>', methods=['PUT', 'PATCH'])
@login_required
def update_operation(operation_id):
    """
    API: Actualizar operación (monto, abonos, pagos)

    Solo Trader puede editar en estado Pendiente
    Master puede editar siempre

    PUT/PATCH JSON:
        amount_usd: float (optional) - Solo editable
        client_deposits: array (optional) - Abonos del cliente
        client_payments: array (optional) - Pagos al cliente
    """
    from app.extensions import db
    from app.models.operation import Operation
    from app.models.audit_log import AuditLog

    data = request.get_json()
    operation = Operation.query.get(operation_id)

    if not operation:
        return jsonify({'success': False, 'message': 'Operación no encontrada'}), 404

    # Validar permisos
    if current_user.role == 'Trader':
        if not operation.can_trader_edit():
            return jsonify({'success': False, 'message': 'No puedes editar esta operación en estado: ' + operation.status}), 403
    elif current_user.role == 'Operador':
        # Operador solo puede ver, no editar (excepto al finalizar)
        return jsonify({'success': False, 'message': 'Los operadores no pueden editar operaciones'}), 403
    # Master puede editar siempre

    try:
        # Actualizar monto USD si se proporciona
        if 'amount_usd' in data:
            new_amount = float(data['amount_usd'])
            if new_amount <= 0:
                return jsonify({'success': False, 'message': 'El monto debe ser mayor a 0'}), 400

            # Registrar log de modificación
            old_amount = float(operation.amount_usd)
            if old_amount != new_amount:
                operation.add_modification_log(
                    user=current_user,
                    campo='amount_usd',
                    valor_anterior=old_amount,
                    valor_nuevo=new_amount
                )

                operation.amount_usd = new_amount
                # Recalcular monto en PEN
                operation.amount_pen = new_amount * float(operation.exchange_rate)

                # Registrar en auditoría
                AuditLog.log_action(
                    user_id=current_user.id,
                    action='UPDATE_OPERATION',
                    entity='Operation',
                    entity_id=operation.id,
                    details=f'Monto modificado de ${old_amount} a ${new_amount}'
                )

        # Actualizar abonos del cliente
        if 'client_deposits' in data:
            new_deposits = data['client_deposits']
            if not isinstance(new_deposits, list):
                return jsonify({'success': False, 'message': 'client_deposits debe ser un array'}), 400

            # Preservar comprobante_url existente si el frontend no envía uno
            existing_deposits = operation.client_deposits or []
            for i, dep in enumerate(new_deposits):
                if i < len(existing_deposits) and existing_deposits[i].get('comprobante_url'):
                    if not dep.get('comprobante_url'):
                        dep['comprobante_url'] = existing_deposits[i]['comprobante_url']

            operation.client_deposits = new_deposits

        # Actualizar pagos al cliente
        if 'client_payments' in data:
            payments = data['client_payments']
            if not isinstance(payments, list):
                return jsonify({'success': False, 'message': 'client_payments debe ser un array'}), 400
            if len(payments) > 4:
                return jsonify({'success': False, 'message': 'Máximo 4 pagos permitidos'}), 400
            operation.client_payments = payments

        db.session.commit()

        # Notificar actualización de posición
        NotificationService.notify_position_update()

        return jsonify({
            'success': True,
            'message': 'Operación actualizada correctamente',
            'operation': operation.to_dict(include_relations=True)
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al actualizar: {str(e)}'}), 500


@operations_bp.route('/api/send_to_process/<int:operation_id>', methods=['POST'])
@login_required
@require_role('Master', 'Trader')
def send_to_process(operation_id):
    """
    API: Enviar operación a proceso (Trader)

    Valida que:
    - Suma de abonos coincida con el total
    - Suma de pagos coincida con el total
    - La operación esté en estado Pendiente

    Cambia estado a 'En proceso'
    """
    from app.extensions import db
    from app.models.operation import Operation
    from app.models.audit_log import AuditLog

    operation = Operation.query.get(operation_id)

    if not operation:
        return jsonify({'success': False, 'message': 'Operación no encontrada'}), 404

    if operation.status != 'Pendiente':
        return jsonify({'success': False, 'message': 'Solo se pueden enviar a proceso operaciones Pendientes'}), 400

    # Validar suma de abonos
    if not operation.validate_deposits_sum():
        if operation.operation_type == 'Compra':
            expected = float(operation.amount_usd)
            currency = 'USD'
        else:
            expected = float(operation.amount_pen)
            currency = 'PEN'
        return jsonify({
            'success': False,
            'message': f'La suma de abonos ({operation.get_total_deposits():.2f}) no coincide con el total de la operación ({expected:.2f} {currency})'
        }), 400

    # Validar suma de pagos
    if not operation.validate_payments_sum():
        if operation.operation_type == 'Venta':
            expected = float(operation.amount_usd)
            currency = 'USD'
        else:
            expected = float(operation.amount_pen)
            currency = 'PEN'
        return jsonify({
            'success': False,
            'message': f'La suma de pagos ({operation.get_total_payments():.2f}) no coincide con el total de la operación ({expected:.2f} {currency})'
        }), 400

    try:
        operation.status = 'En proceso'
        operation.in_process_since = now_peru()  # Registrar cuándo entró en proceso

        # Asignar operador automáticamente de forma balanceada
        assigned_operator_id = OperationService.assign_operator_balanced()
        if assigned_operator_id:
            operation.assigned_operator_id = assigned_operator_id
            print(f"Operación {operation.operation_id} asignada al operador ID: {assigned_operator_id}")
        else:
            print(f"ADVERTENCIA: No se pudo asignar operador a {operation.operation_id}")

        AuditLog.log_action(
            user_id=current_user.id,
            action='UPDATE_OPERATION',
            entity='Operation',
            entity_id=operation.id,
            details=f'Operación {operation.operation_id} enviada a proceso' +
                   (f' y asignada al operador ID {assigned_operator_id}' if assigned_operator_id else '')
        )

        db.session.commit()

        # Notificar actualización general
        NotificationService.notify_operation_updated(operation, 'Pendiente')
        NotificationService.notify_dashboard_update()

        # Notificar al operador asignado
        if assigned_operator_id:
            from app.models.user import User
            assigned_operator = User.query.get(assigned_operator_id)
            if assigned_operator:
                NotificationService.notify_operation_assigned(operation, assigned_operator)

        return jsonify({
            'success': True,
            'message': 'Operación enviada a proceso correctamente',
            'operation': operation.to_dict(include_relations=True)
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500


@operations_bp.route('/api/return_to_pending/<int:operation_id>', methods=['POST'])
@login_required
@require_role('Operador', 'Master')
def return_to_pending(operation_id):
    """
    API: Devolver operación a pendiente (Operador o Master)

    Solo disponible cuando el estado es 'En proceso'
    Devuelve al trader la capacidad de editar
    """
    from app.extensions import db
    from app.models.operation import Operation
    from app.models.audit_log import AuditLog

    data = request.get_json() or {}
    reason = data.get('reason', '').strip()

    operation = Operation.query.get(operation_id)

    if not operation:
        return jsonify({'success': False, 'message': 'Operación no encontrada'}), 404

    if operation.status != 'En proceso':
        return jsonify({'success': False, 'message': 'Solo se pueden devolver operaciones en proceso'}), 400

    # Verificar que el operador esté asignado a esta operación (Master puede editar todas)
    if current_user.role == 'Operador':
        if not operation.is_assigned_to_operator(current_user.id):
            return jsonify({
                'success': False,
                'message': 'Esta operación está asignada a otro operador. Solo puedes editar operaciones asignadas a ti.'
            }), 403

    try:
        operation.status = 'Pendiente'
        operation.in_process_since = None  # Limpiar timestamp al devolver a pendiente
        operation.assigned_operator_id = None  # Limpiar asignación de operador

        # Agregar nota con razón de devolución
        if reason:
            operation.notes = (operation.notes or '') + f'\n[DEVUELTO] {reason}'

        AuditLog.log_action(
            user_id=current_user.id,
            action='UPDATE_OPERATION',
            entity='Operation',
            entity_id=operation.id,
            details=f'Operación {operation.operation_id} devuelta a pendiente. Razón: {reason or "No especificada"}'
        )

        db.session.commit()

        # Notificar
        NotificationService.notify_operation_updated(operation, 'En proceso')
        NotificationService.notify_dashboard_update()

        return jsonify({
            'success': True,
            'message': 'Operación devuelta a pendiente correctamente',
            'operation': operation.to_dict(include_relations=True)
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500


@operations_bp.route('/api/complete/<int:operation_id>', methods=['POST'])
@login_required
@require_role('Operador', 'Master')
def complete_operation(operation_id):
    """
    API: Finalizar operación (Operador o Master)

    POST JSON:
        operator_proofs: array (optional) - Comprobantes del operador
        operator_comments: string (optional) - Comentarios

    Cambia estado a 'Completada'
    """
    from app.extensions import db
    from app.models.operation import Operation
    from app.models.audit_log import AuditLog
    from datetime import datetime

    data = request.get_json() or {}
    operation = Operation.query.get(operation_id)

    if not operation:
        return jsonify({'success': False, 'message': 'Operación no encontrada'}), 404

    if operation.status != 'En proceso':
        return jsonify({'success': False, 'message': 'Solo se pueden completar operaciones en proceso'}), 400

    # Verificar que el operador esté asignado a esta operación (Master puede editar todas)
    if current_user.role == 'Operador':
        if not operation.is_assigned_to_operator(current_user.id):
            return jsonify({
                'success': False,
                'message': 'Esta operación está asignada a otro operador. Solo puedes editar operaciones asignadas a ti.'
            }), 403

    try:
        # Guardar comprobantes del operador
        if 'operator_proofs' in data:
            proofs = data['operator_proofs']
            if len(proofs) > 4:
                return jsonify({'success': False, 'message': 'Máximo 4 comprobantes permitidos'}), 400
            operation.operator_proofs = proofs

        # Guardar comentarios
        if 'operator_comments' in data:
            operation.operator_comments = data['operator_comments']

        operation.status = 'Completada'
        operation.completed_at = now_peru()
        operation.in_process_since = None  # Limpiar timestamp al completar
        operation.assigned_operator_id = None  # Limpiar asignación de operador

        AuditLog.log_action(
            user_id=current_user.id,
            action='COMPLETE_OPERATION',
            entity='Operation',
            entity_id=operation.id,
            details=f'Operación {operation.operation_id} completada'
        )

        db.session.commit()

        # Enviar email de confirmación
        try:
            import logging
            logger = logging.getLogger(__name__)
            logger.info(f'[COMPLETE] Operacion {operation.operation_id} completada, enviando email...')

            from app.services.email_service import EmailService
            success, message = EmailService.send_completed_operation_email(operation)

            if success:
                logger.info(f'[COMPLETE] Email enviado exitosamente para {operation.operation_id}')
            else:
                logger.warning(f'[COMPLETE] Error al enviar email para {operation.operation_id}: {message}')
        except Exception as e:
            # Log el error pero no falla la operación completada
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'[COMPLETE] Excepcion al enviar email para {operation.operation_id}: {str(e)}')
            logger.exception(e)

        # Notificar
        NotificationService.notify_operation_completed(operation)
        NotificationService.notify_dashboard_update()

        return jsonify({
            'success': True,
            'message': 'Operación completada exitosamente',
            'operation': operation.to_dict(include_relations=True)
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500


@operations_bp.route('/api/upload_deposit_proof/<int:operation_id>', methods=['POST'])
@login_required
@require_role('Master', 'Trader')
def upload_deposit_proof(operation_id):
    """
    API: Subir comprobante de abono del cliente

    Form data:
        file: archivo
        deposit_index: índice del abono a actualizar
    """
    from app.extensions import db
    from app.models.operation import Operation

    print(f"[DEBUG] upload_deposit_proof llamado para operación {operation_id}")

    operation = Operation.query.get(operation_id)
    if not operation:
        print(f"[DEBUG] Operación {operation_id} no encontrada")
        return jsonify({'success': False, 'message': 'Operación no encontrada'}), 404

    if 'file' not in request.files:
        print(f"[DEBUG] No hay archivo en request.files: {list(request.files.keys())}")
        return jsonify({'success': False, 'message': 'No se envió ningún archivo'}), 400

    deposit_index = request.form.get('deposit_index', type=int)
    if deposit_index is None:
        print(f"[DEBUG] deposit_index no proporcionado. Form data: {dict(request.form)}")
        return jsonify({'success': False, 'message': 'deposit_index es requerido'}), 400

    file = request.files['file']
    print(f"[DEBUG] Archivo recibido: {file.filename}, deposit_index: {deposit_index}")

    file_service = FileService()
    print(f"[DEBUG] FileService configurado: {file_service.configured}")

    success, message, url = file_service.upload_file(
        file,
        'deposits',
        f"{operation.operation_id}_deposit_{deposit_index}"
    )

    print(f"[DEBUG] Resultado upload: success={success}, message={message}, url={url}")

    if not success:
        return jsonify({'success': False, 'message': message}), 400

    # Actualizar el abono con la URL del comprobante
    deposits = operation.client_deposits
    # Asegurar que existe el índice (crear entradas vacías si es necesario)
    while len(deposits) <= deposit_index:
        deposits.append({})
    deposits[deposit_index]['comprobante_url'] = url
    operation.client_deposits = deposits
    db.session.commit()

    return jsonify({
        'success': True,
        'message': 'Comprobante subido correctamente',
        'url': url
    })


@operations_bp.route('/api/upload_operator_proof/<int:operation_id>', methods=['POST'])
@login_required
@require_role('Operador', 'Master')
def upload_operator_proof_new(operation_id):
    """
    API: Subir comprobante del operador

    Form data:
        file: archivo
        proof_index: índice del comprobante
        comentario: string (optional)
    """
    from app.extensions import db
    from app.models.operation import Operation

    operation = Operation.query.get(operation_id)
    if not operation:
        return jsonify({'success': False, 'message': 'Operación no encontrada'}), 404

    if operation.status != 'En proceso':
        return jsonify({'success': False, 'message': 'Solo se pueden subir comprobantes en operaciones en proceso'}), 400

    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No se envió ningún archivo'}), 400

    file = request.files['file']
    proof_index = request.form.get('proof_index', type=int, default=0)
    comentario = request.form.get('comentario', '')

    file_service = FileService()

    success, message, url = file_service.upload_file(
        file,
        'operator_proofs',
        f"{operation.operation_id}_proof_{proof_index}"
    )

    if not success:
        return jsonify({'success': False, 'message': message}), 400

    # Agregar comprobante a la lista
    proofs = operation.operator_proofs
    if len(proofs) >= 4:
        return jsonify({'success': False, 'message': 'Máximo 4 comprobantes permitidos'}), 400

    proofs.append({
        'comprobante_url': url,
        'comentario': comentario
    })
    operation.operator_proofs = proofs
    db.session.commit()

    return jsonify({
        'success': True,
        'message': 'Comprobante subido correctamente',
        'url': url,
        'proofs': operation.operator_proofs
    })


@operations_bp.route('/api/mark_notes_read/<int:operation_id>', methods=['POST'])
@login_required
@require_role('Operador', 'Master')
def mark_notes_read(operation_id):
    """
    API: Marcar notas como leídas por el usuario actual

    POST: Marca las notas de la operación como leídas por el usuario autenticado
    """
    from app.extensions import db
    from app.models.operation import Operation

    operation = Operation.query.get(operation_id)
    if not operation:
        return jsonify({'success': False, 'message': 'Operación no encontrada'}), 404

    try:
        # Marcar notas como leídas por el usuario actual
        operation.mark_notes_as_read(current_user.id)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Notas marcadas como leídas',
            'notes_read_by': operation.notes_read_by
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500


@operations_bp.route('/api/check_pending_operations', methods=['GET'])
@login_required
@require_role('Operador')
def check_pending_operations():
    """
    API: Verificar operaciones en estado 'En proceso' que requieren atención (Operador)

    Retorna operaciones que:
    - Están en estado 'En proceso'
    - Han estado en ese estado por 10 minutos o más
    - Son del día actual (según in_process_since)

    Returns:
        JSON con lista de operaciones que necesitan atención y tiempo transcurrido
    """
    from app.extensions import db
    from app.models.operation import Operation
    from sqlalchemy import and_, func
    from datetime import datetime

    # Obtener inicio y fin del día actual en Perú
    now = now_peru()
    start_of_day = datetime(now.year, now.month, now.day, 0, 0, 0)
    end_of_day = datetime(now.year, now.month, now.day, 23, 59, 59)

    # Obtener operaciones en proceso asignadas al operador actual y del día actual
    operations = Operation.query.filter(
        and_(
            Operation.status == 'En proceso',
            Operation.in_process_since.isnot(None),
            Operation.assigned_operator_id == current_user.id,  # Solo sus operaciones asignadas
            Operation.in_process_since >= start_of_day,  # Solo del día actual
            Operation.in_process_since <= end_of_day
        )
    ).all()

    pending_operations = []
    current_time = now_peru()

    for operation in operations:
        time_in_process = operation.get_time_in_process_minutes()

        # Solo incluir si ha pasado 10 minutos o más
        if time_in_process and time_in_process >= 10:
            # Determinar si debe mostrar notificación
            # Primera notificación: a los 10 minutos
            # Notificaciones subsiguientes: cada 5 minutos después de los 10
            minutes_since_first_alert = time_in_process - 10
            should_alert = (time_in_process == 10) or (minutes_since_first_alert > 0 and minutes_since_first_alert % 5 == 0)

            op_data = operation.to_dict(include_relations=True)
            op_data['time_in_process_minutes'] = time_in_process
            op_data['should_alert'] = should_alert
            pending_operations.append(op_data)

    return jsonify({
        'success': True,
        'pending_operations': pending_operations,
        'count': len(pending_operations)
    })


@operations_bp.route('/api/reassign_operator/<int:operation_id>', methods=['POST'])
@login_required
@require_role('Master')
def reassign_operator(operation_id):
    """
    API: Reasignar operador a una operación (Solo Master)

    POST JSON:
        new_operator_id: int - ID del nuevo operador

    Returns:
        JSON con resultado de la reasignación
    """
    from app.extensions import db
    from app.models.operation import Operation
    from app.models.user import User
    from app.models.audit_log import AuditLog

    data = request.get_json() or {}
    new_operator_id = data.get('new_operator_id')

    if not new_operator_id:
        return jsonify({'success': False, 'message': 'Debe especificar el nuevo operador'}), 400

    operation = Operation.query.get(operation_id)
    if not operation:
        return jsonify({'success': False, 'message': 'Operación no encontrada'}), 404

    # Validar que la operación esté en proceso
    if operation.status != 'En proceso':
        return jsonify({'success': False, 'message': 'Solo se pueden reasignar operaciones en proceso'}), 400

    # Validar que el nuevo operador exista y esté activo
    new_operator = User.query.get(new_operator_id)
    if not new_operator:
        return jsonify({'success': False, 'message': 'Operador no encontrado'}), 404

    if new_operator.role != 'Operador':
        return jsonify({'success': False, 'message': 'El usuario seleccionado no es un operador'}), 400

    if new_operator.status != 'Activo':
        return jsonify({'success': False, 'message': 'El operador seleccionado no está activo'}), 400

    # Validar que el operador esté conectado al sistema
    if not new_operator.is_online():
        return jsonify({'success': False, 'message': 'El operador seleccionado no está conectado al sistema'}), 400

    # Guardar operador anterior para el log
    old_operator_id = operation.assigned_operator_id
    old_operator_name = 'No asignado'
    if old_operator_id:
        old_operator = User.query.get(old_operator_id)
        old_operator_name = old_operator.username if old_operator else 'Desconocido'

    try:
        # Reasignar operador
        operation.assigned_operator_id = new_operator_id

        # Registrar en audit log
        AuditLog.log_action(
            user_id=current_user.id,
            action='REASSIGN_OPERATOR',
            entity='Operation',
            entity_id=operation.id,
            details=f'Operación {operation.operation_id} reasignada de {old_operator_name} a {new_operator.username}'
        )

        db.session.commit()

        # Notificar actualización general
        NotificationService.notify_operation_updated(operation, 'En proceso')

        # Notificar reasignación al operador anterior y al nuevo
        old_operator_obj = User.query.get(old_operator_id) if old_operator_id else None
        NotificationService.notify_operation_reassigned(
            operation=operation,
            old_operator=old_operator_obj,
            new_operator=new_operator,
            reassigned_by=current_user
        )

        return jsonify({
            'success': True,
            'message': f'Operación reasignada exitosamente a {new_operator.username}',
            'operation': operation.to_dict(include_relations=True)
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500


@operations_bp.route('/api/get_active_operators', methods=['GET'])
@login_required
@require_role('Master')
def get_active_operators():
    """
    API: Obtener lista de operadores activos y conectados (Solo Master)

    Returns:
        JSON con lista de operadores activos y conectados al sistema
    """
    from app.models.user import User

    operators = User.query.filter(
        User.role == 'Operador',
        User.status == 'Activo'
    ).all()

    # Filtrar solo operadores conectados
    online_operators = [op for op in operators if op.is_online()]

    operators_list = [{
        'id': op.id,
        'username': op.username,
        'email': op.email
    } for op in online_operators]

    return jsonify({
        'success': True,
        'operators': operators_list,
        'count': len(operators_list)
    })
