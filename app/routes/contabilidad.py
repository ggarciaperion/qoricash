"""
Rutas del módulo contable QoriCash — Libro Diario, Gastos, Períodos.
Solo accesible por rol Master.
"""
from flask import Blueprint, render_template, request, jsonify, send_file
from flask_login import login_required, current_user
from app.extensions import db
from app.utils.decorators import require_role
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

contabilidad_bp = Blueprint('contabilidad', __name__)


def _current_period_defaults():
    """Retorna year y month del período activo (hoy)."""
    today = date.today()
    return today.year, today.month


def _company_info() -> dict:
    """Lee razón social y RUC desde SystemConfig (M-03). Fallback a valores por defecto."""
    from app.models.system_config import SystemConfig
    try:
        return {
            'razon_social': SystemConfig.get('RAZON_SOCIAL', 'QORICASH SAC'),
            'ruc':          SystemConfig.get('RUC', '20615113698'),
        }
    except Exception:
        return {'razon_social': 'QORICASH SAC', 'ruc': '20615113698'}


# ── Dashboard contable ─────────────────────────────────────────────────────────

@contabilidad_bp.route('/')
@contabilidad_bp.route('/dashboard')
@login_required
@require_role('Master')
def dashboard():
    from app.models.accounting_period import AccountingPeriod
    from app.models.journal_entry import JournalEntry
    from app.models.expense_record import ExpenseRecord
    from sqlalchemy import func, extract

    year, month = _current_period_defaults()

    # Período actual
    current_period = AccountingPeriod.query.filter_by(year=year, month=month).first()

    # Asientos del mes actual
    entries_this_month = JournalEntry.query.filter(
        extract('year', JournalEntry.entry_date) == year,
        extract('month', JournalEntry.entry_date) == month,
        JournalEntry.status == 'activo',
    ).order_by(JournalEntry.entry_date.desc()).limit(10).all()

    # Totales del mes
    totals = db.session.query(
        func.sum(JournalEntry.total_debe).label('debe'),
        func.sum(JournalEntry.total_haber).label('haber'),
        func.count(JournalEntry.id).label('count'),
    ).filter(
        extract('year', JournalEntry.entry_date) == year,
        extract('month', JournalEntry.entry_date) == month,
        JournalEntry.status == 'activo',
    ).first()

    # Gastos del mes
    gastos_total = db.session.query(
        func.sum(ExpenseRecord.amount_pen)
    ).filter(
        extract('year', ExpenseRecord.expense_date) == year,
        extract('month', ExpenseRecord.expense_date) == month,
    ).scalar() or Decimal('0')

    # Todos los períodos para el selector
    periods = AccountingPeriod.query.order_by(
        AccountingPeriod.year.desc(), AccountingPeriod.month.desc()
    ).all()

    return render_template(
        'contabilidad/dashboard.html',
        current_period=current_period,
        entries_this_month=entries_this_month,
        total_debe=totals.debe or Decimal('0'),
        total_haber=totals.haber or Decimal('0'),
        entries_count=totals.count or 0,
        gastos_total=gastos_total,
        periods=periods,
        selected_year=year,
        selected_month=month,
        user=current_user,
    )


# ── Libro Diario ───────────────────────────────────────────────────────────────

@contabilidad_bp.route('/diario')
@login_required
@require_role('Master')
def diario():
    from app.models.journal_entry import JournalEntry
    from app.models.accounting_period import AccountingPeriod

    year  = request.args.get('year',  type=int, default=date.today().year)
    month = request.args.get('month', type=int, default=date.today().month)
    status_filter = request.args.get('status', 'activo')

    q = JournalEntry.query.filter_by(status=status_filter)

    if year and month:
        from sqlalchemy import extract
        q = q.filter(
            extract('year',  JournalEntry.entry_date) == year,
            extract('month', JournalEntry.entry_date) == month,
        )
    elif year:
        from sqlalchemy import extract
        q = q.filter(extract('year', JournalEntry.entry_date) == year)

    entries = q.order_by(JournalEntry.entry_date.asc(), JournalEntry.id.asc()).all()

    periods = AccountingPeriod.query.order_by(
        AccountingPeriod.year.desc(), AccountingPeriod.month.desc()
    ).all()

    return render_template(
        'contabilidad/diario.html',
        entries=entries,
        periods=periods,
        selected_year=year,
        selected_month=month,
        status_filter=status_filter,
        user=current_user,
    )


@contabilidad_bp.route('/diario/nuevo', methods=['POST'])
@login_required
@require_role('Master')
def nuevo_asiento():
    """Crear un asiento manual con N líneas de debe/haber."""
    from app.services.accounting.journal_service import JournalService

    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'Sin datos'}), 400

    try:
        entry_date = date.fromisoformat(data.get('entry_date', str(date.today())))
        description = data.get('description', '').strip()
        raw_lines   = data.get('lines', [])

        if not description:
            return jsonify({'success': False, 'error': 'La glosa es obligatoria'}), 400
        if len(raw_lines) < 2:
            return jsonify({'success': False, 'error': 'Se requieren al menos 2 líneas'}), 400

        lines = []
        total_debe = Decimal('0')
        total_haber = Decimal('0')
        for l in raw_lines:
            debe  = Decimal(str(l.get('debe',  0) or 0))
            haber = Decimal(str(l.get('haber', 0) or 0))
            if not l.get('account_code'):
                return jsonify({'success': False, 'error': 'Todas las líneas requieren código de cuenta'}), 400
            lines.append({
                'account_code': l['account_code'].strip(),
                'description':  l.get('description', '').strip() or None,
                'debe':         debe,
                'haber':        haber,
                'currency':     l.get('currency', 'PEN'),
            })
            total_debe  += debe
            total_haber += haber

        # Validar códigos de cuenta contra catálogo PCGE (A-02)
        from app.models.accounting_account import AccountingAccount
        catalog_codes = {
            a.code for a in AccountingAccount.query.filter_by(is_active=True).all()
        }
        if catalog_codes:  # Solo bloquear si el catálogo tiene datos
            for l in lines:
                code = l['account_code']
                # Válido si es una cuenta del catálogo o sub-cuenta de alguna
                if not any(code == cat or code.startswith(cat) for cat in catalog_codes):
                    return jsonify({
                        'success': False,
                        'error': f'Cuenta "{code}" no existe en el catálogo PCGE. '
                                 'Verifique el código o regístrelo en el Plan de Cuentas.'
                    }), 400

        diff = abs(total_debe - total_haber)
        if diff != Decimal('0'):
            return jsonify({
                'success': False,
                'error': f'El asiento no cuadra: DEBE {total_debe:.2f} ≠ HABER {total_haber:.2f} '
                         f'(diferencia: {diff:.2f})'
            }), 400

        entry = JournalService.create_entry(
            entry_type='manual',
            description=description,
            lines=lines,
            source_type='manual',
            source_id=None,
            entry_date=entry_date,
            created_by=current_user.id,
        )

        if entry:
            return jsonify({'success': True, 'entry_number': entry.entry_number})
        else:
            return jsonify({'success': False, 'error': 'Error al guardar el asiento. Revisa los logs.'}), 500

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@contabilidad_bp.route('/diario/<int:entry_id>/anular', methods=['POST'])
@login_required
@require_role('Master')
def anular_asiento(entry_id):
    """Anular un asiento: cambia status a 'anulado' y crea asiento inverso."""
    from app.models.journal_entry import JournalEntry
    from app.services.accounting.journal_service import JournalService

    data = request.get_json() or {}
    motivo = data.get('motivo', '').strip()
    if not motivo:
        return jsonify({'success': False, 'error': 'Se requiere motivo de anulación'}), 400

    entry = JournalEntry.query.get_or_404(entry_id)

    if entry.status == 'anulado':
        return jsonify({'success': False, 'error': 'El asiento ya está anulado'}), 400

    try:
        # Validar que el período del asiento original esté abierto (M-05)
        period = JournalService.get_or_create_period(entry.entry_date)
        if period.status == 'cerrado':
            return jsonify({
                'success': False,
                'error': f'El período {entry.entry_date.strftime("%m/%Y")} está cerrado. '
                         'Reabra el período antes de anular este asiento.'
            }), 409

        # Marcar original como anulado
        entry.status          = 'anulado'
        entry.annulled_at     = datetime.utcnow()
        entry.annulled_by     = current_user.id
        entry.annulled_reason = motivo
        db.session.flush()

        # Crear asiento inverso en la misma fecha del original (M-05)
        lines = entry.lines.all() if hasattr(entry.lines, 'all') else list(entry.lines)
        inverse_lines = [{
            'account_code': l.account_code,
            'description':  f'Reversión: {l.description or entry.description}',
            'debe':         l.haber,   # invertir
            'haber':        l.debe,
            'currency':     l.currency or 'PEN',
        } for l in lines]

        reversal = JournalService.create_entry(
            entry_type='manual',
            description=f'ANULACIÓN: {entry.entry_number} — {motivo}',
            lines=inverse_lines,
            source_type='anulacion',
            source_id=entry.id,
            entry_date=entry.entry_date,   # misma fecha del asiento original
            created_by=current_user.id,
        )

        if not reversal:
            db.session.rollback()
            return jsonify({
                'success': False,
                'error': 'No se pudo crear el asiento de reversión. Revisa los logs.'
            }), 500

        db.session.commit()
        return jsonify({'success': True, 'message': f'Asiento {entry.entry_number} anulado'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@contabilidad_bp.route('/diario/<int:entry_id>/lines')
@login_required
@require_role('Master')
def entry_lines(entry_id):
    """API: líneas de un asiento (para el acordeón del diario)."""
    from app.models.journal_entry import JournalEntry

    entry = JournalEntry.query.get_or_404(entry_id)
    lines = entry.lines.all() if hasattr(entry.lines, 'all') else list(entry.lines)

    return jsonify([{
        'account_code': l.account_code,
        'description':  l.description or '',
        'debe':         float(l.debe or 0),
        'haber':        float(l.haber or 0),
        'currency':     l.currency or 'PEN',
        'amount_usd':   float(l.amount_usd) if l.amount_usd else None,
        'exchange_rate': float(l.exchange_rate) if l.exchange_rate else None,
    } for l in lines])


@contabilidad_bp.route('/diario/export')
@login_required
@require_role('Master')
def export_diario():
    """Exportar Libro Diario en Excel."""
    from app.models.journal_entry import JournalEntry
    from sqlalchemy import extract

    year  = request.args.get('year',  type=int, default=date.today().year)
    month = request.args.get('month', type=int, default=date.today().month)

    entries = JournalEntry.query.filter(
        extract('year',  JournalEntry.entry_date) == year,
        extract('month', JournalEntry.entry_date) == month,
        JournalEntry.status == 'activo',
    ).order_by(JournalEntry.entry_date.asc(), JournalEntry.id.asc()).all()

    co = _company_info()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Libro Diario'

    # Encabezado SUNAT (M-03)
    ws.merge_cells('A1:H1')
    ws['A1'] = co['razon_social']
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:H2')
    ws['A2'] = f'RUC: {co["ruc"]}'
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A3:H3')
    ws['A3'] = f'LIBRO DIARIO — {_month_name(month)} {year} — Folio 0001'
    ws['A3'].font = Font(bold=True)
    ws['A3'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A4:H4')
    ws['A4'] = 'Régimen MYPE Tributario — PCGE'
    ws['A4'].alignment = Alignment(horizontal='center')

    # Headers tabla
    header_fill = PatternFill(start_color='1B3A6B', end_color='1B3A6B', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    headers = ['N° Asiento', 'Fecha', 'Tipo', 'Descripción', 'Cta.', 'Glosa línea', 'DEBE (S/)', 'HABER (S/)']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=6, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center')

    row = 7
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for entry in entries:
        lines = entry.lines.all() if hasattr(entry.lines, 'all') else list(entry.lines)
        first_line = True
        for line in lines:
            ws.cell(row=row, column=1, value=entry.entry_number if first_line else '').border = border
            ws.cell(row=row, column=2, value=entry.entry_date.strftime('%d/%m/%Y') if first_line else '').border = border
            ws.cell(row=row, column=3, value=entry.entry_type if first_line else '').border = border
            ws.cell(row=row, column=4, value=entry.description if first_line else '').border = border
            ws.cell(row=row, column=5, value=line.account_code).border = border
            ws.cell(row=row, column=6, value=line.description or '').border = border
            ws.cell(row=row, column=7, value=float(line.debe or 0)).border = border
            ws.cell(row=row, column=8, value=float(line.haber or 0)).border = border
            first_line = False
            row += 1
        # Fila de totales del asiento
        ws.cell(row=row, column=5, value='TOTAL ASIENTO').font = Font(bold=True)
        ws.cell(row=row, column=7, value=float(entry.total_debe)).font = Font(bold=True)
        ws.cell(row=row, column=8, value=float(entry.total_haber)).font = Font(bold=True)
        row += 1

    # Ajuste anchos
    widths = [16, 12, 22, 48, 8, 40, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'libro_diario_qoricash_{year}{month:02d}.xlsx'
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=filename)


# ── Gastos ─────────────────────────────────────────────────────────────────────

@contabilidad_bp.route('/gastos')
@login_required
@require_role('Master')
def gastos():
    from app.models.expense_record import ExpenseRecord
    from app.models.accounting_period import AccountingPeriod
    from app.models.journal_entry import JournalEntry
    from sqlalchemy import extract

    year  = request.args.get('year',  type=int, default=date.today().year)
    month = request.args.get('month', type=int, default=date.today().month)

    records = ExpenseRecord.query.filter(
        extract('year',  ExpenseRecord.expense_date) == year,
        extract('month', ExpenseRecord.expense_date) == month,
    ).order_by(ExpenseRecord.expense_date.desc()).all()

    # IDs de gastos que ya tienen asiento de pago registrado
    pagados_ids = set(
        row[0] for row in db.session.query(JournalEntry.source_id).filter(
            JournalEntry.entry_type == 'pago_proveedor',
            JournalEntry.source_type == 'expense_record',
            JournalEntry.source_id.in_([r.id for r in records]) if records else [],
        ).all()
    ) if records else set()

    periods = AccountingPeriod.query.order_by(
        AccountingPeriod.year.desc(), AccountingPeriod.month.desc()
    ).all()

    return render_template(
        'contabilidad/gastos.html',
        records=records,
        pagados_ids=pagados_ids,
        periods=periods,
        selected_year=year,
        selected_month=month,
        user=current_user,
    )


@contabilidad_bp.route('/gastos/nuevo', methods=['POST'])
@login_required
@require_role('Master')
def nuevo_gasto():
    """
    Registrar un gasto / desembolso y generar el asiento contable correcto.

    Reglas de asiento (PCGE — casa de cambio exonerada de IGV):

    CASO A — Factura con IGV (casa de cambio NO usa crédito fiscal por defecto):
        DEBE  63xx / 33xx   amount_pen  (total incl. IGV — IGV es costo)
        HABER 4211           amount_pen  (Facturas por pagar)

    CASO B — Factura con IGV + crédito_fiscal=True (prorrata o actividad gravada):
        DEBE  63xx / 33xx   base_pen
        DEBE  4011           igv_pen    (IGV recuperable)
        HABER 4211           amount_pen

    CASO C — Boleta / recibo / sin comprobante:
        DEBE  63xx           amount_pen
        HABER 4699           amount_pen  (Otras cuentas por pagar)

    CASO D — Planilla:
        DEBE  62x            amount_pen
        HABER 4111           amount_pen  (Sueldos por pagar)
    """
    from app.models.expense_record import ExpenseRecord
    from app.services.accounting.journal_service import JournalService

    data = request.get_json() or request.form

    try:
        expense_date   = date.fromisoformat(data.get('expense_date', str(date.today())))
        amount_pen     = Decimal(str(data.get('amount_pen', 0)))
        igv_raw        = data.get('igv_pen')
        base_raw       = data.get('base_pen')
        igv_pen        = Decimal(str(igv_raw))   if igv_raw  else None
        base_pen       = Decimal(str(base_raw))  if base_raw else None
        credito_fiscal = bool(data.get('credito_fiscal', False))
        expense_type   = data.get('expense_type', 'servicio')
        voucher_type   = data.get('voucher_type') or None

        # Auto-calcular base/IGV si no se proveen pero hay factura
        if voucher_type == 'factura' and igv_pen is None and amount_pen > 0:
            igv_pen  = (amount_pen / Decimal('1.18') * Decimal('0.18')).quantize(Decimal('0.01'))
            base_pen = amount_pen - igv_pen

        record = ExpenseRecord(
            expense_date=expense_date,
            category=data.get('category', '6391'),
            description=data.get('description', ''),
            amount_pen=amount_pen,
            base_pen=base_pen,
            igv_pen=igv_pen,
            credito_fiscal=credito_fiscal,
            expense_type=expense_type,
            voucher_type=voucher_type,
            voucher_number=data.get('voucher_number') or None,
            supplier_ruc=data.get('supplier_ruc') or None,
            supplier_name=data.get('supplier_name') or None,
            created_by=current_user.id,
        )

        period = JournalService.get_or_create_period(expense_date)
        if period.status == 'cerrado':
            return jsonify({
                'success': False,
                'error': f'El período {expense_date.strftime("%m/%Y")} está cerrado.'
            }), 409
        record.period_id = period.id
        db.session.add(record)
        db.session.flush()

        # ── Determinar cuentas del asiento ────────────────────────────────────
        account_code = data.get('category', '6391')
        is_planilla  = expense_type == 'planilla'

        if is_planilla:
            # Planilla: cargo a remuneraciones, por pagar a 4111
            lines = [
                {'account_code': account_code, 'description': record.description,
                 'debe': amount_pen, 'haber': Decimal('0'), 'currency': 'PEN'},
                {'account_code': '4111', 'description': f'Por pagar planilla: {record.description}',
                 'debe': Decimal('0'), 'haber': amount_pen, 'currency': 'PEN'},
            ]
        elif voucher_type == 'factura' and credito_fiscal and igv_pen:
            # Factura con crédito fiscal: separar IGV en 4011
            lines = [
                {'account_code': account_code, 'description': record.description,
                 'debe': base_pen, 'haber': Decimal('0'), 'currency': 'PEN'},
                {'account_code': '4011', 'description': f'IGV compras – {record.description}',
                 'debe': igv_pen, 'haber': Decimal('0'), 'currency': 'PEN'},
                {'account_code': '4211', 'description': f'Factura por pagar: {record.description}',
                 'debe': Decimal('0'), 'haber': amount_pen, 'currency': 'PEN'},
            ]
        elif voucher_type == 'factura':
            # Factura SIN crédito fiscal: IGV es costo (cargo total a la cuenta de gasto)
            lines = [
                {'account_code': account_code, 'description': record.description,
                 'debe': amount_pen, 'haber': Decimal('0'), 'currency': 'PEN'},
                {'account_code': '4211', 'description': f'Factura por pagar: {record.description}',
                 'debe': Decimal('0'), 'haber': amount_pen, 'currency': 'PEN'},
            ]
        else:
            # Boleta / recibo / sin comprobante
            lines = [
                {'account_code': account_code, 'description': record.description,
                 'debe': amount_pen, 'haber': Decimal('0'), 'currency': 'PEN'},
                {'account_code': '4699', 'description': f'Por pagar: {record.description}',
                 'debe': Decimal('0'), 'haber': amount_pen, 'currency': 'PEN'},
            ]

        entry_type = 'activo_fijo' if expense_type == 'activo_fijo' else 'gasto'
        entry = JournalService.create_entry(
            entry_type=entry_type,
            description=f"{'Adq. activo' if entry_type == 'activo_fijo' else 'Gasto'}: {record.description}",
            lines=lines,
            source_type='expense',
            source_id=record.id,
            entry_date=expense_date,
            created_by=current_user.id,
        )
        if entry:
            record.journal_entry_id = entry.id

        # ── Si es activo fijo, crear el registro en FixedAsset ────────────────
        if expense_type == 'activo_fijo':
            from app.models.fixed_asset import FixedAsset
            fa_category     = data.get('fa_category', 'equipo_oficina')
            fa_name         = data.get('fa_name') or record.description
            fa_life_months  = data.get('fa_life_months')
            fa_residual     = data.get('fa_residual_value')
            fixed_asset = FixedAsset.from_expense(
                expense_record=record,
                category=fa_category,
                name=fa_name,
                useful_life_months=int(fa_life_months) if fa_life_months else None,
                residual_value=Decimal(str(fa_residual)) if fa_residual else None,
                created_by=current_user.id,
            )
            db.session.add(fixed_asset)

        db.session.commit()
        return jsonify({'success': True, 'message': 'Gasto registrado correctamente'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


# ── Pago de Proveedores ────────────────────────────────────────────────────────

@contabilidad_bp.route('/gastos/<int:record_id>/pagar', methods=['POST'])
@login_required
@require_role('Master')
def pagar_proveedor(record_id):
    """
    Registra el pago de una factura o planilla pendiente.
    Asiento: DEBE 4211 (facturas) o 4111 (planilla) / HABER cuenta_banco seleccionada
    Solo aplica a gastos con comprobante tipo Factura, o expense_type planilla.
    """
    from app.models.expense_record import ExpenseRecord
    from app.models.journal_entry import JournalEntry
    from app.services.accounting.journal_service import JournalService

    data          = request.get_json() or {}
    payment_date  = date.fromisoformat(data.get('payment_date', str(date.today())))
    bank_account  = data.get('bank_account', '1041')   # cuenta desde donde se paga
    bank_label    = data.get('bank_label', bank_account)

    record = ExpenseRecord.query.get_or_404(record_id)

    # Verificar que sea una factura o planilla (que generó un pasivo)
    es_factura = (record.voucher_type or '').lower() in ('factura', 'factura electrónica')
    es_planilla = (record.expense_type or '') == 'planilla'

    if not es_factura and not es_planilla:
        return jsonify({'success': False, 'error': 'Solo se pueden pagar facturas y planillas.'}), 400

    # Verificar que no tenga ya un asiento de pago
    existing = JournalEntry.query.filter_by(
        entry_type='pago_proveedor',
        source_type='expense_record',
        source_id=record.id,
    ).first()
    if existing:
        return jsonify({'success': False, 'error': 'Esta factura/planilla ya fue pagada.'}), 409

    contra_debe = '4111' if es_planilla else '4211'
    monto = Decimal(str(record.amount_pen))

    entry = JournalService.create_entry(
        entry_type='pago_proveedor',
        description=(
            f'Pago {record.voucher_type or "planilla"} '
            f'{record.voucher_number or ""} – '
            f'{record.supplier_name or record.description[:40]}'
        ).strip(),
        lines=[
            {
                'account_code': contra_debe,
                'description': f'Cancelación {contra_debe} – {record.description[:40]}',
                'debe': monto,
                'haber': Decimal('0'),
                'currency': 'PEN',
            },
            {
                'account_code': bank_account,
                'description': f'Pago desde {bank_label}',
                'debe': Decimal('0'),
                'haber': monto,
                'currency': 'PEN',
            },
        ],
        source_type='expense_record',
        source_id=record.id,
        entry_date=payment_date,
        created_by=current_user.id,
    )

    if not entry:
        return jsonify({'success': False, 'error': 'Período cerrado o error al crear asiento.'}), 409

    return jsonify({
        'success': True,
        'message': f'Pago registrado. Asiento {entry.entry_number}',
        'entry_number': entry.entry_number,
    })


# ── Ajuste por diferencia de tipo de cambio ────────────────────────────────────

@contabilidad_bp.route('/ajuste-fx', methods=['POST'])
@login_required
@require_role('Master')
def ajuste_fx_cierre():
    """
    Ajuste monetario por diferencia de tipo de cambio al cierre de período (NIC 21 / CPC 14).
    Revalúa los saldos en USD al tipo de cambio de cierre.

    Si USD apreciado : DEBE 1012 (Caja ME) / HABER 7761 (Ganancia diferencia TC)
    Si USD depreciado: DEBE 6762 (Pérdida diferencia TC) / HABER 1012 (Caja ME)

    La base de USD se toma de BankBalance.balance_usd (saldo operativo actual).
    El valor libro PEN se toma de los saldos acumulados de cuentas USD en el diario.
    """
    from app.models.bank_balance import BankBalance
    from app.models.journal_entry import JournalEntry
    from app.services.accounting.journal_service import JournalService
    import calendar as cal

    data         = request.get_json() or {}
    year         = int(data.get('year',  date.today().year))
    month        = int(data.get('month', date.today().month))
    closing_rate = Decimal(str(data.get('closing_rate', '0')))

    if closing_rate <= 0:
        return jsonify({'success': False, 'error': 'Ingresa un tipo de cambio de cierre válido.'}), 400

    _, last_day = cal.monthrange(year, month)
    corte = date(year, month, last_day)

    # ── Saldo USD por cuenta (BankBalance + mapeo PCGE) ─────────────────────
    # Mapeo: nombre de banco → código PCGE cuenta USD
    _BANK_USD_PCGE = {
        'BCP':       '1044',
        'INTERBANK': '1047',
        'BANBIF':    '1050',
        'PICHINCHA': '1052',
    }
    banks = BankBalance.query.all()
    total_usd = sum(Decimal(str(b.balance_usd)) for b in banks)

    if total_usd <= 0:
        return jsonify({'success': False, 'error': 'No hay saldos en USD registrados en BankBalance.'})

    # ── Valor libro PEN de cuentas USD en el diario ───────────────────────────
    usd_accounts = ('1012', '1044', '1047', '1050', '1052')
    pen_libro = sum(_saldo_acumulado_hasta(c, corte, 'deudora') for c in usd_accounts)

    # ── Diferencia total ──────────────────────────────────────────────────────
    pen_revaluado = total_usd * closing_rate
    diferencia    = pen_revaluado - pen_libro

    if abs(diferencia) < Decimal('0.01'):
        return jsonify({
            'success': False,
            'error': f'Diferencia ({diferencia:.2f}) menor a S/ 0.01. No requiere ajuste.',
        })

    # ── Verificar duplicado en el período ─────────────────────────────────────
    period = JournalService.get_or_create_period(date(year, month, 1))
    if period.status == 'cerrado':
        return jsonify({'success': False, 'error': f'El período {period.label} está cerrado.'}), 409

    existing = JournalEntry.query.filter_by(
        entry_type='ajuste_fx',
        period_id=period.id,
    ).first()
    if existing:
        return jsonify({
            'success': False,
            'error': f'Ya existe un ajuste FX para {period.label} (asiento {existing.entry_number}).',
        }), 409

    # ── Construir líneas del asiento por cuenta USD ───────────────────────────
    # Distribuye el ajuste proporcionalmente entre cada cuenta USD con saldo > 0
    descripcion = (
        f'Ajuste diferencia de cambio – cierre {period.label} '
        f'(USD {total_usd:.2f} × TC {closing_rate} = S/ {pen_revaluado:.2f}; libro S/ {pen_libro:.2f})'
    )

    # Calcular diferencia por cuenta individual
    cuenta_diferencias = []
    for code in usd_accounts:
        pen_lib_cuenta  = _saldo_acumulado_hasta(code, corte, 'deudora')
        # USD de esta cuenta: usar BankBalance si hay match, sino proporcional
        usd_cuenta = Decimal('0')
        if code == '1012':
            # Caja ME — no mapea a BankBalance directamente; usa el residual
            usd_bancos = sum(
                Decimal(str(b.balance_usd))
                for b in banks
                if b.bank_name.upper() in _BANK_USD_PCGE
            )
            usd_cuenta = total_usd - usd_bancos
        else:
            bank_name_match = next(
                (k for k, v in _BANK_USD_PCGE.items() if v == code), None
            )
            if bank_name_match:
                bank_row = next(
                    (b for b in banks if b.bank_name.upper() == bank_name_match), None
                )
                usd_cuenta = Decimal(str(bank_row.balance_usd)) if bank_row else Decimal('0')

        if usd_cuenta <= 0 and pen_lib_cuenta <= 0:
            continue

        pen_rev_cuenta = usd_cuenta * closing_rate
        diff_cuenta    = pen_rev_cuenta - pen_lib_cuenta
        if abs(diff_cuenta) >= Decimal('0.01'):
            cuenta_diferencias.append((code, usd_cuenta, diff_cuenta))

    if not cuenta_diferencias:
        return jsonify({
            'success': False,
            'error': 'No se encontraron cuentas USD con diferencia ajustable.',
        })

    lines = []
    if diferencia > 0:
        # Ganancia: DEBE cuentas USD / HABER 7761
        for code, usd_c, diff_c in cuenta_diferencias:
            if diff_c > 0:
                lines.append({
                    'account_code': code,
                    'description': f'Revaluación {code} – ajuste TC cierre',
                    'debe': diff_c, 'haber': Decimal('0'),
                    'currency': 'USD', 'amount_usd': usd_c, 'exchange_rate': closing_rate,
                })
        lines.append({
            'account_code': '7761',
            'description': 'Ganancia diferencia TC – ajuste cierre',
            'debe': Decimal('0'), 'haber': diferencia, 'currency': 'PEN',
        })
    else:
        monto = abs(diferencia)
        lines.append({
            'account_code': '6762',
            'description': 'Pérdida diferencia TC – ajuste cierre',
            'debe': monto, 'haber': Decimal('0'), 'currency': 'PEN',
        })
        for code, usd_c, diff_c in cuenta_diferencias:
            if diff_c < 0:
                lines.append({
                    'account_code': code,
                    'description': f'Revaluación {code} – ajuste TC cierre',
                    'debe': Decimal('0'), 'haber': abs(diff_c),
                    'currency': 'USD', 'amount_usd': usd_c, 'exchange_rate': closing_rate,
                })

    entry = JournalService.create_entry(
        entry_type='ajuste_fx',
        description=descripcion,
        lines=lines,
        source_type='ajuste_fx',
        source_id=period.id,
        entry_date=corte,
        created_by=current_user.id,
    )

    if not entry:
        return jsonify({'success': False, 'error': 'Error al crear asiento.'}), 500

    signo = 'Ganancia' if diferencia > 0 else 'Pérdida'
    return jsonify({
        'success': True,
        'message': f'Ajuste FX registrado. {signo} S/ {abs(diferencia):.2f}. Asiento {entry.entry_number}',
        'diferencia': float(diferencia),
        'total_usd': float(total_usd),
        'pen_libro': float(pen_libro),
        'pen_revaluado': float(pen_revaluado),
        'entry_number': entry.entry_number,
    })


# ── Asiento de Apertura ────────────────────────────────────────────────────────

# Cuentas disponibles para el asiento de apertura (banco + caja)
_APERTURA_ACCOUNTS = [
    # PEN
    ('1011', 'Caja Moneda Nacional',                  'PEN'),
    ('1041', 'BCP — Cta. Cte. PEN',                   'PEN'),
    ('1048', 'Interbank — Cta. Cte. PEN',             'PEN'),
    ('1049', 'BanBif — Cta. Cte. PEN',                'PEN'),
    ('1051', 'Pichincha — Cta. Cte. PEN (en apertura)', 'PEN'),
    # USD
    ('1012', 'Caja Moneda Extranjera',                'USD'),
    ('1044', 'BCP — Cta. Cte. USD',                   'USD'),
    ('1047', 'Interbank — Cta. Cte. USD',             'USD'),
    ('1050', 'BanBif — Cta. Cte. USD',                'USD'),
    ('1052', 'Pichincha — Cta. Cte. USD (en apertura)', 'USD'),
]

_CONTRAPARTIDA_OPTIONS = [
    ('3111', 'Capital Social'),
    ('3511', 'Utilidades Acumuladas'),
    ('4511', 'Préstamos de accionistas'),
]


@contabilidad_bp.route('/apertura')
@login_required
@require_role('Master')
def apertura():
    from app.models.journal_entry import JournalEntry
    from sqlalchemy import extract

    year = request.args.get('year', type=int, default=date.today().year)

    # Verificar si ya existe un asiento de apertura para este año
    existing = JournalEntry.query.filter(
        extract('year', JournalEntry.entry_date) == year,
        JournalEntry.entry_type == 'apertura',
        JournalEntry.status == 'activo',
    ).first()

    return render_template(
        'contabilidad/apertura.html',
        accounts=_APERTURA_ACCOUNTS,
        contrapartida_options=_CONTRAPARTIDA_OPTIONS,
        existing=existing,
        selected_year=year,
        user=current_user,
    )


@contabilidad_bp.route('/apertura', methods=['POST'])
@login_required
@require_role('Master')
def crear_apertura():
    from app.models.journal_entry import JournalEntry
    from app.services.accounting.journal_service import JournalService
    from sqlalchemy import extract

    data = request.get_json() or {}

    try:
        apertura_date = date.fromisoformat(data.get('apertura_date', str(date.today())))
        contrapartida = data.get('contrapartida', '3111')
        forzar        = data.get('forzar', False)
        saldos        = data.get('saldos', {})  # {account_code: importe_str}

        year = apertura_date.year

        # Verificar duplicado
        existing = JournalEntry.query.filter(
            extract('year', JournalEntry.entry_date) == year,
            JournalEntry.entry_type == 'apertura',
            JournalEntry.status == 'activo',
        ).first()

        if existing and not forzar:
            return jsonify({
                'success': False,
                'duplicate': True,
                'entry_number': existing.entry_number,
                'message': f'Ya existe el asiento de apertura {existing.entry_number} para {year}. '
                           '¿Desea reemplazarlo?',
            }), 409

        if existing and forzar:
            # Anular el asiento previo antes de crear el nuevo
            existing.status        = 'anulado'
            existing.annulled_at   = datetime.utcnow()
            existing.annulled_by   = current_user.id
            existing.annulled_reason = 'Reemplazado por nuevo asiento de apertura'
            db.session.flush()

        # Construir líneas DEBE
        lines = []
        total = Decimal('0')
        for code, name, currency in _APERTURA_ACCOUNTS:
            raw = saldos.get(code, '').strip()
            if not raw:
                continue
            try:
                importe = Decimal(raw.replace(',', '.'))
            except Exception:
                continue
            if importe <= 0:
                continue

            lines.append({
                'account_code': code,
                'description':  f'Saldo inicial {name}',
                'debe':         importe,
                'haber':        Decimal('0'),
                'currency':     currency,
            })
            total += importe

        if not lines:
            return jsonify({'success': False, 'error': 'Debe ingresar al menos un saldo mayor a 0'}), 400

        # Línea HABER: contrapartida (capital / patrimonio)
        lines.append({
            'account_code': contrapartida,
            'description':  'Saldo de apertura — contrapartida patrimonial',
            'debe':         Decimal('0'),
            'haber':        total,
            'currency':     'PEN',
        })

        entry = JournalService.create_entry(
            entry_type='apertura',
            description=f'Asiento de apertura {year} — saldos iniciales',
            lines=lines,
            source_type='manual',
            source_id=None,
            entry_date=apertura_date,
            created_by=current_user.id,
        )

        if not entry:
            return jsonify({'success': False, 'error': 'No se pudo crear el asiento. Revisa que el período esté abierto.'}), 500

        return jsonify({
            'success': True,
            'entry_number': entry.entry_number,
            'message': f'Asiento de apertura {entry.entry_number} creado correctamente',
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400


# ── Períodos ───────────────────────────────────────────────────────────────────

@contabilidad_bp.route('/periodos')
@login_required
@require_role('Master')
def periodos():
    from app.models.accounting_period import AccountingPeriod
    periods = AccountingPeriod.query.order_by(
        AccountingPeriod.year.desc(), AccountingPeriod.month.desc()
    ).all()
    return render_template(
        'contabilidad/periodos.html',
        periods=periods,
        user=current_user,
    )


@contabilidad_bp.route('/periodos/cerrar', methods=['POST'])
@login_required
@require_role('Master')
def cerrar_periodo():
    from app.services.accounting.journal_service import JournalService
    data = request.get_json() or request.form
    year  = int(data.get('year'))
    month = int(data.get('month'))
    success, message = JournalService.close_period(year, month, current_user.id)
    return jsonify({'success': success, 'message': message})


@contabilidad_bp.route('/periodos/reabrir', methods=['POST'])
@login_required
@require_role('Master')
def reabrir_periodo():
    """
    Reabre un período contable cerrado.
    Requiere motivo obligatorio — queda registrado en AuditLog.
    """
    from app.models.accounting_period import AccountingPeriod
    from app.models.audit_log import AuditLog

    data   = request.get_json() or {}
    year   = int(data.get('year'))
    month  = int(data.get('month'))
    motivo = (data.get('motivo') or '').strip()

    if not motivo:
        return jsonify({'success': False, 'message': 'El motivo de reapertura es obligatorio.'}), 400

    period = AccountingPeriod.query.filter_by(year=year, month=month).first()
    if not period:
        return jsonify({'success': False, 'message': 'Período no encontrado.'}), 404
    if period.status == 'abierto':
        return jsonify({'success': False, 'message': f'El período {period.label} ya está abierto.'}), 409

    try:
        period.status    = 'abierto'
        period.closed_at = None
        period.closed_by = None

        AuditLog.log_action(
            user_id    = current_user.id,
            action     = 'REOPEN_PERIOD',
            entity     = 'AccountingPeriod',
            entity_id  = period.id,
            details    = f'Reapertura del período {period.label}. Motivo: {motivo}',
            ip_address = request.remote_addr,
            user_agent = request.headers.get('User-Agent', '')[:200],
        )

        db.session.commit()
        return jsonify({
            'success': True,
            'message': f'Período {period.label} reabierto correctamente.',
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@contabilidad_bp.route('/periodos/generar-asientos', methods=['POST'])
@login_required
@require_role('Master')
def generar_asientos_retroactivos():
    """
    Genera asientos contables retroactivos para operaciones Completadas que
    no tienen asiento registrado. Útil cuando el módulo contable se desplegó
    después de que las operaciones ya estaban completadas.
    """
    from app.models import Operation
    from app.models.journal_entry import JournalEntry
    from app.services.accounting.journal_service import JournalService
    from sqlalchemy import extract

    data  = request.get_json() or {}
    year  = int(data.get('year',  date.today().year))
    month = int(data.get('month', date.today().month))

    # IDs de operaciones que ya tienen asiento activo
    existing_ids = {
        r[0] for r in db.session.query(JournalEntry.source_id).filter(
            JournalEntry.source_type == 'operation',
            JournalEntry.status == 'activo',
        ).all() if r[0] is not None
    }

    # Operaciones completadas en el período sin asiento
    ops = Operation.query.filter(
        Operation.status == 'Completada',
        extract('year',  Operation.completed_at) == year,
        extract('month', Operation.completed_at) == month,
        ~Operation.id.in_(existing_ids) if existing_ids else True,
    ).order_by(Operation.completed_at.asc()).all()

    if not ops:
        return jsonify({
            'success': True,
            'message': 'Todas las operaciones del período ya tienen asiento contable.',
            'generados': 0,
            'errores': 0,
        })

    generados = 0
    errores   = []

    for op in ops:
        entry = JournalService.create_entry_for_completed_operation(
            op, created_by_id=current_user.id
        )
        if entry:
            generados += 1
        else:
            errores.append(op.operation_id)

    return jsonify({
        'success': True,
        'message': f'{generados} asiento(s) generado(s). {len(errores)} error(es).',
        'generados': generados,
        'errores': len(errores),
        'operaciones_con_error': errores,
        'total_sin_asiento': len(ops),
    })


# ── Libro Caja y Bancos ────────────────────────────────────────────────────────

# Mapa código PCGE → etiqueta legible
_ACCOUNT_LABELS = {
    '1011': ('Caja MN',              'PEN', 'efectivo'),
    '1012': ('Caja ME',              'USD', 'efectivo'),
    '1041': ('BCP PEN',              'PEN', 'banco'),
    '1044': ('BCP USD',              'USD', 'banco'),
    '1047': ('Interbank USD',        'USD', 'banco'),
    '1048': ('Interbank PEN',        'PEN', 'banco'),
    '1049': ('BanBif PEN',           'PEN', 'banco'),
    '1050': ('BanBif USD',           'USD', 'banco'),
    '1051': ('Pichincha PEN',        'PEN', 'banco'),
    '1052': ('Pichincha USD',        'USD', 'banco'),
}

_CODE_TO_BANK = {
    '1041': 'BCP',       '1044': 'BCP',
    '1047': 'INTERBANK', '1048': 'INTERBANK',
    '1049': 'BANBIF',    '1050': 'BANBIF',
    '1051': 'PICHINCHA', '1052': 'PICHINCHA',
}


def _caja_movimientos(year: int, month: int, account_code: str):
    """
    Retorna lista de movimientos de una cuenta en el período, ordenados por fecha.
    Cada item: {date, entry_number, description, debe, haber, saldo_acum}
    También retorna saldo_anterior (suma hasta fin del mes previo).
    """
    from app.models.journal_entry import JournalEntry
    from app.models.journal_entry_line import JournalEntryLine
    from sqlalchemy import func, extract
    import calendar

    # Saldo anterior = todo lo contabilizado antes del período actual
    first_day = date(year, month, 1)
    prev = db.session.query(
        func.sum(JournalEntryLine.debe).label('d'),
        func.sum(JournalEntryLine.haber).label('h'),
    ).join(JournalEntry, JournalEntryLine.journal_entry_id == JournalEntry.id
    ).filter(
        JournalEntryLine.account_code == account_code,
        JournalEntry.entry_date < first_day,
        JournalEntry.status == 'activo',
    ).first()

    saldo_ant = Decimal(str(prev.d or 0)) - Decimal(str(prev.h or 0))

    # Movimientos del período
    rows = db.session.query(
        JournalEntry.entry_date,
        JournalEntry.entry_number,
        JournalEntry.description.label('entry_desc'),
        JournalEntryLine.description.label('line_desc'),
        JournalEntryLine.debe,
        JournalEntryLine.haber,
    ).join(JournalEntry, JournalEntryLine.journal_entry_id == JournalEntry.id
    ).filter(
        JournalEntryLine.account_code == account_code,
        extract('year',  JournalEntry.entry_date) == year,
        extract('month', JournalEntry.entry_date) == month,
        JournalEntry.status == 'activo',
    ).order_by(
        JournalEntry.entry_date.asc(), JournalEntry.id.asc()
    ).all()

    movs = []
    saldo = saldo_ant
    for r in rows:
        debe  = Decimal(str(r.debe  or 0))
        haber = Decimal(str(r.haber or 0))
        saldo += debe - haber
        movs.append({
            'date':         r.entry_date,
            'entry_number': r.entry_number,
            'description':  r.line_desc or r.entry_desc,
            'debe':         debe,
            'haber':        haber,
            'saldo':        saldo,
        })

    total_debe  = sum(m['debe']  for m in movs)
    total_haber = sum(m['haber'] for m in movs)

    return {
        'code':        account_code,
        'label':       _ACCOUNT_LABELS.get(account_code, (f'Cta. {account_code}', 'PEN', 'otro'))[0],
        'currency':    _ACCOUNT_LABELS.get(account_code, ('', 'PEN', ''))[1],
        'kind':        _ACCOUNT_LABELS.get(account_code, ('', '', 'otro'))[2],
        'saldo_ant':   saldo_ant,
        'movimientos': movs,
        'total_debe':  total_debe,
        'total_haber': total_haber,
        'saldo_final': saldo_ant + total_debe - total_haber,
    }


@contabilidad_bp.route('/caja')
@login_required
@require_role('Master')
def caja():
    from app.models.bank_balance import BankBalance
    from app.models.accounting_period import AccountingPeriod
    from app.models.journal_entry_line import JournalEntryLine
    from app.models.journal_entry import JournalEntry
    from sqlalchemy import func, extract

    year  = request.args.get('year',  type=int, default=date.today().year)
    month = request.args.get('month', type=int, default=date.today().month)

    # Determinar qué cuentas tienen movimiento en el período o saldo anterior
    used_codes = {r[0] for r in db.session.query(JournalEntryLine.account_code).join(
        JournalEntry, JournalEntryLine.journal_entry_id == JournalEntry.id
    ).filter(
        JournalEntry.status == 'activo',
    ).distinct().all()} & set(_ACCOUNT_LABELS.keys())

    # Si no hay nada en la DB, mostrar todas de todas formas
    codes_to_show = sorted(used_codes) if used_codes else sorted(_ACCOUNT_LABELS.keys())

    accounts = [_caja_movimientos(year, month, code) for code in codes_to_show]
    # Solo mostrar cuentas con saldo anterior != 0 o con movimientos en el período
    accounts = [a for a in accounts if a['movimientos'] or a['saldo_ant'] != 0]

    # Saldos reales del módulo de posición (para conciliación)
    bank_balances = {b.bank_name.upper(): b for b in BankBalance.query.all()}

    periods = AccountingPeriod.query.order_by(
        AccountingPeriod.year.desc(), AccountingPeriod.month.desc()
    ).all()

    return render_template(
        'contabilidad/caja.html',
        accounts=accounts,
        bank_balances=bank_balances,
        CODE_TO_BANK=_CODE_TO_BANK,
        periods=periods,
        selected_year=year,
        selected_month=month,
        user=current_user,
    )


@contabilidad_bp.route('/caja/export')
@login_required
@require_role('Master')
def export_caja():
    from app.models.bank_balance import BankBalance
    from app.models.journal_entry_line import JournalEntryLine
    from app.models.journal_entry import JournalEntry

    year  = request.args.get('year',  type=int, default=date.today().year)
    month = request.args.get('month', type=int, default=date.today().month)

    used_codes = {r[0] for r in db.session.query(JournalEntryLine.account_code).join(
        JournalEntry, JournalEntryLine.journal_entry_id == JournalEntry.id
    ).filter(JournalEntry.status == 'activo').distinct().all()} & set(_ACCOUNT_LABELS.keys())

    codes = sorted(used_codes) if used_codes else sorted(_ACCOUNT_LABELS.keys())
    accounts = [_caja_movimientos(year, month, c) for c in codes]
    accounts = [a for a in accounts if a['movimientos'] or a['saldo_ant'] != 0]

    co = _company_info()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # quitamos hoja vacía default

    hfill  = PatternFill(start_color='1B3A6B', end_color='1B3A6B', fill_type='solid')
    sfill  = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    tfill  = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    thin   = Side(style='thin')
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)

    for folio, acc in enumerate(accounts, 1):
        ws = wb.create_sheet(f"{acc['code']} {acc['label']}"[:31])

        # Encabezado SUNAT (M-03)
        ws.merge_cells('A1:G1')
        ws['A1'] = co['razon_social']
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:G2')
        ws['A2'] = f'RUC: {co["ruc"]}'
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A3:G3')
        ws['A3'] = (f"LIBRO CAJA Y BANCOS — {acc['code']} {acc['label']} "
                    f"({acc['currency']}) — {_month_name(month)} {year} — Folio {folio:04d}")
        ws['A3'].font = Font(bold=True)
        ws['A3'].alignment = Alignment(horizontal='center')

        # Saldo anterior
        ws.merge_cells('A5:D5')
        ws['A5'] = 'SALDO ANTERIOR (al inicio del período)'
        ws['A5'].font = Font(bold=True)
        ws['A5'].fill = sfill
        ws['E5'] = float(max(acc['saldo_ant'], Decimal('0')))
        ws['F5'] = float(max(-acc['saldo_ant'], Decimal('0')))
        for c in 'ABCDEF':
            ws[f'{c}5'].border = bdr

        # Headers
        hdrs = ['N°', 'Fecha', 'N° Asiento', 'Descripción', 'DEBE', 'HABER', 'SALDO']
        for col, h in enumerate(hdrs, 1):
            cell = ws.cell(row=6, column=col, value=h)
            cell.fill = hfill
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')
            cell.border = bdr

        row = 7
        for n, m in enumerate(acc['movimientos'], 1):
            ws.cell(row=row, column=1, value=n).border = bdr
            ws.cell(row=row, column=2, value=m['date'].strftime('%d/%m/%Y')).border = bdr
            ws.cell(row=row, column=3, value=m['entry_number']).border = bdr
            ws.cell(row=row, column=4, value=m['description']).border = bdr
            ws.cell(row=row, column=5, value=float(m['debe'])).border = bdr
            ws.cell(row=row, column=6, value=float(m['haber'])).border = bdr
            ws.cell(row=row, column=7, value=float(m['saldo'])).border = bdr
            row += 1

        # Totales
        ws.cell(row=row, column=4, value='TOTAL DEL PERÍODO').font = Font(bold=True)
        ws.cell(row=row, column=5, value=float(acc['total_debe'])).font  = Font(bold=True)
        ws.cell(row=row, column=6, value=float(acc['total_haber'])).font = Font(bold=True)
        for c in range(1, 8):
            ws.cell(row=row, column=c).fill   = tfill
            ws.cell(row=row, column=c).border = bdr
        row += 1

        ws.cell(row=row, column=4, value='SALDO FINAL').font = Font(bold=True)
        ws.cell(row=row, column=7, value=float(acc['saldo_final'])).font = Font(bold=True)
        for c in range(1, 8):
            ws.cell(row=row, column=c).border = bdr

        widths = [5, 12, 16, 54, 14, 14, 14]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    if not wb.worksheets:
        ws = wb.create_sheet('Sin movimientos')
        ws['A1'] = 'No hay movimientos en cuentas de caja/bancos para el período seleccionado.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'libro_caja_bancos_{year}{month:02d}.xlsx')


# ── Libro de Ingresos y Gastos (LIG) ─────────────────────────────────────────

_INGRESO_LABELS = {
    '70': 'Ventas',
    '71': 'Variación de inventarios',
    '72': 'Producción de activo inmovilizado',
    '73': 'Descuentos, rebajas y bonificaciones obtenidas',
    '74': 'Descuentos, rebajas y bonificaciones concedidas',
    '75': 'Otros ingresos de gestión',
    '76': 'Ingresos excepcionales',
    '77': 'Diferencial cambiario',
    '78': 'Cargas cubiertas por provisiones',
    '79': 'Cargas imputables a cuentas de costos y gastos',
}


@contabilidad_bp.route('/lig')
@login_required
@require_role('Master')
def lig():
    """
    Libro de Ingresos y Gastos — formato SUNAT (RS 234-2006).
    Ingresos: líneas de asientos con cuentas 7xxx (ingresos PCGE).
    Gastos:   expense_records del período.
    """
    from app.models.journal_entry import JournalEntry
    from app.models.journal_entry_line import JournalEntryLine
    from app.models.expense_record import ExpenseRecord
    from app.models.accounting_period import AccountingPeriod
    from sqlalchemy import extract, func

    year  = request.args.get('year',  type=int, default=date.today().year)
    month = request.args.get('month', type=int, default=date.today().month)

    # Ingresos: líneas de cuentas 7xxx activas en el período
    rows = db.session.query(
        JournalEntry.entry_date,
        JournalEntry.entry_number,
        JournalEntry.description.label('entry_desc'),
        JournalEntry.entry_type,
        JournalEntryLine.account_code,
        JournalEntryLine.description.label('line_desc'),
        JournalEntryLine.haber,
    ).join(
        JournalEntry, JournalEntryLine.journal_entry_id == JournalEntry.id
    ).filter(
        JournalEntryLine.account_code.like('7%'),
        JournalEntryLine.haber > 0,
        extract('year',  JournalEntry.entry_date) == year,
        extract('month', JournalEntry.entry_date) == month,
        JournalEntry.status == 'activo',
    ).order_by(
        JournalEntry.entry_date.asc(), JournalEntry.id.asc()
    ).all()

    ingresos = [{
        'fecha':        r.entry_date,
        'entry_number': r.entry_number,
        'entry_type':   r.entry_type,
        'account_code': r.account_code,
        'tipo_ingreso': _INGRESO_LABELS.get(r.account_code[:2], r.account_code),
        'description':  r.line_desc or r.entry_desc,
        'importe':      Decimal(str(r.haber or 0)),
    } for r in rows]

    # Gastos: expense_records del período
    gastos = ExpenseRecord.query.filter(
        extract('year',  ExpenseRecord.expense_date) == year,
        extract('month', ExpenseRecord.expense_date) == month,
    ).order_by(ExpenseRecord.expense_date.asc()).all()

    total_ingresos = sum(float(i['importe']) for i in ingresos)
    total_gastos   = sum(float(g.amount_pen or 0) for g in gastos)

    periods = AccountingPeriod.query.order_by(
        AccountingPeriod.year.desc(), AccountingPeriod.month.desc()
    ).all()

    return render_template(
        'contabilidad/lig.html',
        ingresos=ingresos,
        gastos=gastos,
        total_ingresos=total_ingresos,
        total_gastos=total_gastos,
        periods=periods,
        selected_year=year,
        selected_month=month,
        user=current_user,
    )


@contabilidad_bp.route('/lig/export')
@login_required
@require_role('Master')
def export_lig():
    """Exporta el LIG en Excel con dos hojas: Ingresos y Gastos."""
    from app.models.journal_entry import JournalEntry
    from app.models.journal_entry_line import JournalEntryLine
    from app.models.expense_record import ExpenseRecord
    from sqlalchemy import extract as sa_extract

    year  = request.args.get('year',  type=int, default=date.today().year)
    month = request.args.get('month', type=int, default=date.today().month)

    rows = db.session.query(
        JournalEntry.entry_date,
        JournalEntry.entry_number,
        JournalEntry.description.label('entry_desc'),
        JournalEntryLine.account_code,
        JournalEntryLine.description.label('line_desc'),
        JournalEntryLine.haber,
    ).join(
        JournalEntry, JournalEntryLine.journal_entry_id == JournalEntry.id
    ).filter(
        JournalEntryLine.account_code.like('7%'),
        JournalEntryLine.haber > 0,
        sa_extract('year',  JournalEntry.entry_date) == year,
        sa_extract('month', JournalEntry.entry_date) == month,
        JournalEntry.status == 'activo',
    ).order_by(JournalEntry.entry_date.asc(), JournalEntry.id.asc()).all()

    ingresos = [{
        'fecha':        r.entry_date,
        'entry_number': r.entry_number,
        'account_code': r.account_code,
        'tipo_ingreso': _INGRESO_LABELS.get(r.account_code[:2], r.account_code),
        'description':  r.line_desc or r.entry_desc,
        'importe':      float(r.haber or 0),
    } for r in rows]

    gastos = ExpenseRecord.query.filter(
        sa_extract('year',  ExpenseRecord.expense_date) == year,
        sa_extract('month', ExpenseRecord.expense_date) == month,
    ).order_by(ExpenseRecord.expense_date.asc()).all()

    wb = openpyxl.Workbook()

    header_fill = PatternFill(start_color='1B3A6B', end_color='1B3A6B', fill_type='solid')
    hfont = Font(bold=True, color='FFFFFF')
    center = Alignment(horizontal='center')
    thin  = Side(style='thin')
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)

    co = _company_info()

    def title_rows(ws, titulo, ncols):
        col_letter = openpyxl.utils.get_column_letter(ncols)
        ws.merge_cells(f'A1:{col_letter}1')
        ws['A1'] = co['razon_social']
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].alignment = center
        ws.merge_cells(f'A2:{col_letter}2')
        ws['A2'] = f'RUC: {co["ruc"]}'
        ws['A2'].alignment = center
        ws.merge_cells(f'A3:{col_letter}3')
        ws['A3'] = titulo
        ws['A3'].font = Font(bold=True)
        ws['A3'].alignment = center
        ws.merge_cells(f'A4:{col_letter}4')
        ws['A4'] = f'Período: {_month_name(month)} {year}  —  R.M. MYPE Tributario'
        ws['A4'].alignment = center

    # ── Hoja 1: INGRESOS ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Ingresos'
    title_rows(ws1, f'LIBRO DE INGRESOS — {_month_name(month).upper()} {year}', 7)

    hdrs = ['N°', 'Fecha', 'N° Asiento', 'Tipo de Ingreso', 'Cuenta PCGE', 'Descripción', 'Importe S/']
    for c, h in enumerate(hdrs, 1):
        cell = ws1.cell(row=5, column=c, value=h)
        cell.fill = header_fill
        cell.font = hfont
        cell.alignment = center
        cell.border = bdr

    row = 6
    for n, ing in enumerate(ingresos, 1):
        ws1.cell(row=row, column=1, value=n).border = bdr
        ws1.cell(row=row, column=2, value=ing['fecha'].strftime('%d/%m/%Y')).border = bdr
        ws1.cell(row=row, column=3, value=ing['entry_number']).border = bdr
        ws1.cell(row=row, column=4, value=ing['tipo_ingreso']).border = bdr
        ws1.cell(row=row, column=5, value=ing['account_code']).border = bdr
        ws1.cell(row=row, column=6, value=ing['description']).border = bdr
        ws1.cell(row=row, column=7, value=ing['importe']).border = bdr
        row += 1

    # Fila de total
    ws1.cell(row=row, column=6, value='TOTAL').font = Font(bold=True)
    total_i = sum(i['importe'] for i in ingresos)
    ws1.cell(row=row, column=7, value=total_i).font = Font(bold=True)
    ws1.cell(row=row, column=7).fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')

    widths1 = [5, 12, 16, 28, 12, 54, 14]
    for i, w in enumerate(widths1, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Hoja 2: GASTOS ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Gastos')
    title_rows(ws2, f'LIBRO DE GASTOS — {_month_name(month).upper()} {year}', 9)

    hdrs2 = ['N°', 'Fecha', 'Proveedor', 'RUC', 'Tipo Comprobante',
             'N° Comprobante', 'Cta. PCGE', 'Importe S/', 'Descripción']
    for c, h in enumerate(hdrs2, 1):
        cell = ws2.cell(row=5, column=c, value=h)
        cell.fill = header_fill
        cell.font = hfont
        cell.alignment = center
        cell.border = bdr

    row2 = 6
    for n, g in enumerate(gastos, 1):
        ws2.cell(row=row2, column=1, value=n).border = bdr
        ws2.cell(row=row2, column=2, value=g.expense_date.strftime('%d/%m/%Y')).border = bdr
        ws2.cell(row=row2, column=3, value=g.supplier_name or '—').border = bdr
        ws2.cell(row=row2, column=4, value=g.supplier_ruc  or '—').border = bdr
        ws2.cell(row=row2, column=5, value=g.voucher_type  or '—').border = bdr
        ws2.cell(row=row2, column=6, value=g.voucher_number or '—').border = bdr
        ws2.cell(row=row2, column=7, value=g.category).border = bdr
        ws2.cell(row=row2, column=8, value=float(g.amount_pen)).border = bdr
        ws2.cell(row=row2, column=9, value=g.description).border = bdr
        row2 += 1

    ws2.cell(row=row2, column=7, value='TOTAL').font = Font(bold=True)
    total_g = sum(float(g.amount_pen) for g in gastos)
    ws2.cell(row=row2, column=8, value=total_g).font = Font(bold=True)
    ws2.cell(row=row2, column=8).fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')

    widths2 = [5, 12, 36, 14, 18, 16, 10, 14, 48]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'LIG_qoricash_{year}{month:02d}.xlsx')


@contabilidad_bp.route('/lig/export_ple')
@login_required
@require_role('Master')
def export_lig_ple():
    """
    Exporta el Libro de Ingresos y Gastos en formato PLE SUNAT (M-03).
    Genera dos archivos .txt pipe-delimited en un zip:
      - LE{RUC}AAAAMM00080100001.txt  (Ingresos)
      - LE{RUC}AAAAMM00080200001.txt  (Gastos)

    Estructura basada en el Formato 8.1 / 8.2 del PLE SUNAT para
    Libro de Ingresos y Gastos (Régimen MYPE Tributario).
    """
    import zipfile
    from app.models.journal_entry import JournalEntry
    from app.models.journal_entry_line import JournalEntryLine
    from app.models.expense_record import ExpenseRecord
    from app.models.system_config import SystemConfig
    from sqlalchemy import extract

    year  = request.args.get('year',  type=int, default=date.today().year)
    month = request.args.get('month', type=int, default=date.today().month)

    ruc          = SystemConfig.get('RUC', '20615113698')
    periodo      = f'{year}{month:02d}00'
    ruc_clean    = ruc.replace('-', '').replace(' ', '')

    # ── Ingresos: líneas JournalEntryLine con cuentas 7xxx ───────────────────
    from app.models.journal_entry_line import JournalEntryLine as JEL
    lines_7 = db.session.query(
        JournalEntry, JEL
    ).join(JEL, JEL.journal_entry_id == JournalEntry.id
    ).filter(
        extract('year',  JournalEntry.entry_date) == year,
        extract('month', JournalEntry.entry_date) == month,
        JournalEntry.status == 'activo',
        JEL.account_code.like('7%'),
    ).order_by(JournalEntry.entry_date.asc(), JournalEntry.id.asc()).all()

    ing_lines = []
    for corr, (entry, line) in enumerate(lines_7, 1):
        # Formato 8.1 simplificado: Período|Correlativo|Fecha|TipoComp|Serie|Num|
        #   TipoDoc|NumDoc|RazonSocial|ValExport|BaseImpon|Descuento|IGV|
        #   Inafecto|Exonerado|Total|Moneda|TC|FechaRef|TipoRef|SerieRef|
        #   NumRef|MedPago|Estado
        haber = float(line.haber or 0)
        fields = [
            periodo,                              # 1 Período
            f'{corr:05d}',                       # 2 Correlativo
            entry.entry_date.strftime('%d/%m/%Y'),# 3 Fecha
            '00',                                 # 4 Tipo comprobante (00=sin comprobante)
            '',                                   # 5 Serie
            entry.entry_number,                   # 6 Número correlativo comprobante
            '',                                   # 7 Tipo doc identidad
            '',                                   # 8 Número documento
            '',                                   # 9 Razón social
            '0.00',                               # 10 Valor exportación
            f'{haber:.2f}',                       # 11 Base imponible
            '0.00',                               # 12 Descuento
            '0.00',                               # 13 IGV
            '0.00',                               # 14 Inafecto
            '0.00',                               # 15 Exonerado
            f'{haber:.2f}',                       # 16 Total
            'PEN',                                # 17 Moneda
            '1.000',                              # 18 Tipo de cambio
            '',                                   # 19 Fecha comprobante referenciado
            '',                                   # 20 Tipo comprobante ref
            '',                                   # 21 Serie ref
            '',                                   # 22 Num ref
            '',                                   # 23 Indicador medio pago
            '1',                                  # 24 Estado (1=activo)
        ]
        ing_lines.append('|'.join(fields) + '|\n')

    # ── Gastos: expense_records del período ───────────────────────────────────
    gastos = ExpenseRecord.query.filter(
        extract('year',  ExpenseRecord.expense_date) == year,
        extract('month', ExpenseRecord.expense_date) == month,
    ).order_by(ExpenseRecord.expense_date.asc()).all()

    gasto_lines = []
    for corr, g in enumerate(gastos, 1):
        # Formato 8.2 simplificado
        fields = [
            periodo,
            f'{corr:05d}',
            g.expense_date.strftime('%d/%m/%Y'),
            g.voucher_type or '00',
            '',
            g.voucher_number or '',
            '6' if g.supplier_ruc and len(g.supplier_ruc) == 11 else '1',
            g.supplier_ruc or '',
            g.supplier_name or '',
            f'{float(g.amount_pen):.2f}',  # Base imponible
            '0.00',                         # IGV (casa de cambio exonerada)
            f'{float(g.amount_pen):.2f}',   # Total
            'PEN',
            '1.000',
            '1',
        ]
        gasto_lines.append('|'.join(fields) + '|\n')

    # ── Empaquetar en ZIP ─────────────────────────────────────────────────────
    ing_filename   = f'LE{ruc_clean}{periodo}080100001.txt'
    gasto_filename = f'LE{ruc_clean}{periodo}080200001.txt'

    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(ing_filename,   ''.join(ing_lines)   or f'{periodo}|||||||||||||||||||||||\n')
        zf.writestr(gasto_filename, ''.join(gasto_lines) or f'{periodo}||||||||||||||||\n')
    zip_buffer.seek(0)

    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'PLE_LIG_{ruc_clean}_{year}{month:02d}.zip',
    )


# ── Libro Mayor ────────────────────────────────────────────────────────────────

@contabilidad_bp.route('/mayor')
@login_required
@require_role('Master')
def mayor():
    """
    Libro Mayor — agrupa movimientos del Libro Diario por cuenta PCGE,
    con saldo acumulado (saldo anterior + movimientos del período).
    Requerido SUNAT para presentación de libros contables (M-04).
    """
    from app.models.journal_entry import JournalEntry
    from app.models.journal_entry_line import JournalEntryLine
    from app.models.accounting_account import AccountingAccount
    from app.models.accounting_period import AccountingPeriod
    from sqlalchemy import func, extract

    year  = request.args.get('year',  type=int, default=date.today().year)
    month = request.args.get('month', type=int, default=date.today().month)

    current_period = AccountingPeriod.query.filter_by(year=year, month=month).first()

    # Todas las cuentas con movimientos en el período
    active_codes = [
        r.account_code for r in
        db.session.query(JournalEntryLine.account_code).join(
            JournalEntry, JournalEntryLine.journal_entry_id == JournalEntry.id
        ).filter(
            extract('year',  JournalEntry.entry_date) == year,
            extract('month', JournalEntry.entry_date) == month,
            JournalEntry.status == 'activo',
        ).distinct().order_by(JournalEntryLine.account_code).all()
    ]

    catalog = {a.code: a for a in AccountingAccount.query.all()}
    first_day = date(year, month, 1)

    cuentas = []
    for code in active_codes:
        # Saldo anterior (todo lo contabilizado antes del período)
        prev = db.session.query(
            func.sum(JournalEntryLine.debe).label('d'),
            func.sum(JournalEntryLine.haber).label('h'),
        ).join(JournalEntry, JournalEntryLine.journal_entry_id == JournalEntry.id
        ).filter(
            JournalEntryLine.account_code == code,
            JournalEntry.entry_date < first_day,
            JournalEntry.status == 'activo',
        ).first()
        saldo_ant = Decimal(str(prev.d or 0)) - Decimal(str(prev.h or 0))

        # Movimientos del período
        rows = db.session.query(
            JournalEntry.entry_date,
            JournalEntry.entry_number,
            JournalEntry.description.label('entry_desc'),
            JournalEntryLine.description.label('line_desc'),
            JournalEntryLine.debe,
            JournalEntryLine.haber,
        ).join(JournalEntry, JournalEntryLine.journal_entry_id == JournalEntry.id
        ).filter(
            JournalEntryLine.account_code == code,
            extract('year',  JournalEntry.entry_date) == year,
            extract('month', JournalEntry.entry_date) == month,
            JournalEntry.status == 'activo',
        ).order_by(JournalEntry.entry_date.asc(), JournalEntry.id.asc()).all()

        movs  = []
        saldo = saldo_ant
        for r in rows:
            d = Decimal(str(r.debe  or 0))
            h = Decimal(str(r.haber or 0))
            saldo += d - h
            movs.append({
                'date':         r.entry_date,
                'entry_number': r.entry_number,
                'description':  r.line_desc or r.entry_desc,
                'debe':         d,
                'haber':        h,
                'saldo':        saldo,
            })

        total_debe  = sum(m['debe']  for m in movs)
        total_haber = sum(m['haber'] for m in movs)
        acc = catalog.get(code)

        cuentas.append({
            'code':        code,
            'name':        acc.name if acc else f'Cuenta {code}',
            'type':        acc.type if acc else _infer_type(code),
            'saldo_ant':   saldo_ant,
            'movimientos': movs,
            'total_debe':  total_debe,
            'total_haber': total_haber,
            'saldo_final': saldo_ant + total_debe - total_haber,
        })

    return render_template(
        'contabilidad/mayor.html',
        cuentas=cuentas,
        selected_year=year,
        selected_month=month,
        current_period=current_period,
        user=current_user,
    )


@contabilidad_bp.route('/mayor/export')
@login_required
@require_role('Master')
def export_mayor():
    """Exporta el Libro Mayor del período a Excel."""
    from app.models.journal_entry import JournalEntry
    from app.models.journal_entry_line import JournalEntryLine
    from app.models.accounting_account import AccountingAccount
    from app.models.system_config import SystemConfig
    from sqlalchemy import func, extract

    year  = request.args.get('year',  type=int, default=date.today().year)
    month = request.args.get('month', type=int, default=date.today().month)

    razon_social = SystemConfig.get('RAZON_SOCIAL', 'QORICASH SAC')
    ruc          = SystemConfig.get('RUC', '20615113698')

    active_codes = [
        r.account_code for r in
        db.session.query(JournalEntryLine.account_code).join(
            JournalEntry, JournalEntryLine.journal_entry_id == JournalEntry.id
        ).filter(
            extract('year',  JournalEntry.entry_date) == year,
            extract('month', JournalEntry.entry_date) == month,
            JournalEntry.status == 'activo',
        ).distinct().order_by(JournalEntryLine.account_code).all()
    ]

    catalog   = {a.code: a for a in AccountingAccount.query.all()}
    first_day = date(year, month, 1)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # eliminar hoja por defecto

    header_fill  = PatternFill(start_color='1B3A6B', end_color='1B3A6B', fill_type='solid')
    sub_fill     = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    folio = 1

    for code in active_codes:
        acc  = catalog.get(code)
        name = acc.name if acc else f'Cuenta {code}'
        ws   = wb.create_sheet(title=f'{code}')

        # Encabezado SUNAT
        ws.merge_cells('A1:G1')
        ws['A1'] = razon_social
        ws['A1'].font = Font(bold=True, size=11)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:G2')
        ws['A2'] = f'RUC: {ruc}'
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A3:G3')
        ws['A3'] = f'LIBRO MAYOR — {_month_name(month)} {year} — Folio {folio:04d}'
        ws['A3'].font = Font(bold=True)
        ws['A3'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A4:G4')
        ws['A4'] = f'Cuenta: {code} — {name}'
        ws['A4'].font = Font(bold=True)
        ws['A4'].fill = sub_fill

        headers = ['Fecha', 'N° Asiento', 'Descripción', 'DEBE S/', 'HABER S/', 'Saldo S/', 'D/A']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=6, column=col, value=h)
            c.fill = header_fill
            c.font = Font(bold=True, color='FFFFFF')
            c.alignment = Alignment(horizontal='center')

        # Saldo anterior
        prev = db.session.query(
            func.sum(JournalEntryLine.debe).label('d'),
            func.sum(JournalEntryLine.haber).label('h'),
        ).join(JournalEntry, JournalEntryLine.journal_entry_id == JournalEntry.id
        ).filter(
            JournalEntryLine.account_code == code,
            JournalEntry.entry_date < first_day,
            JournalEntry.status == 'activo',
        ).first()
        saldo_ant = Decimal(str(prev.d or 0)) - Decimal(str(prev.h or 0))

        row_num = 7
        ws.cell(row=row_num, column=1, value='Saldo anterior')
        ws.cell(row=row_num, column=6, value=float(saldo_ant))
        row_num += 1

        rows_db = db.session.query(
            JournalEntry.entry_date,
            JournalEntry.entry_number,
            JournalEntry.description.label('entry_desc'),
            JournalEntryLine.description.label('line_desc'),
            JournalEntryLine.debe,
            JournalEntryLine.haber,
        ).join(JournalEntry, JournalEntryLine.journal_entry_id == JournalEntry.id
        ).filter(
            JournalEntryLine.account_code == code,
            extract('year',  JournalEntry.entry_date) == year,
            extract('month', JournalEntry.entry_date) == month,
            JournalEntry.status == 'activo',
        ).order_by(JournalEntry.entry_date.asc(), JournalEntry.id.asc()).all()

        saldo = saldo_ant
        for r in rows_db:
            d = Decimal(str(r.debe  or 0))
            h = Decimal(str(r.haber or 0))
            saldo += d - h
            ws.cell(row=row_num, column=1, value=r.entry_date.strftime('%d/%m/%Y'))
            ws.cell(row=row_num, column=2, value=r.entry_number)
            ws.cell(row=row_num, column=3, value=r.line_desc or r.entry_desc)
            ws.cell(row=row_num, column=4, value=float(d))
            ws.cell(row=row_num, column=5, value=float(h))
            ws.cell(row=row_num, column=6, value=float(saldo))
            ws.cell(row=row_num, column=7, value='D' if saldo >= 0 else 'A')
            row_num += 1

        for i, w in enumerate([12, 14, 42, 12, 12, 12, 4], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        folio += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'libro_mayor_{year}{month:02d}.xlsx')


# ── Balance de Comprobación ────────────────────────────────────────────────────

@contabilidad_bp.route('/balance')
@login_required
@require_role('Master')
def balance():
    from app.models.journal_entry import JournalEntry
    from app.models.journal_entry_line import JournalEntryLine
    from app.models.accounting_account import AccountingAccount
    from app.models.accounting_period import AccountingPeriod
    from sqlalchemy import func, extract

    year  = request.args.get('year',  type=int, default=date.today().year)
    month = request.args.get('month', type=int, default=date.today().month)

    # Totales acumulados por cuenta en el período
    rows = db.session.query(
        JournalEntryLine.account_code,
        func.sum(JournalEntryLine.debe).label('total_debe'),
        func.sum(JournalEntryLine.haber).label('total_haber'),
    ).join(
        JournalEntry, JournalEntryLine.journal_entry_id == JournalEntry.id
    ).filter(
        extract('year',  JournalEntry.entry_date) == year,
        extract('month', JournalEntry.entry_date) == month,
        JournalEntry.status == 'activo',
    ).group_by(
        JournalEntryLine.account_code
    ).order_by(
        JournalEntryLine.account_code
    ).all()

    # Catálogo para nombres y tipo
    catalog = {a.code: a for a in AccountingAccount.query.all()}

    accounts = []
    grand_debe = Decimal('0')
    grand_haber = Decimal('0')

    for r in rows:
        code = r.account_code
        acc  = catalog.get(code)
        td   = Decimal(str(r.total_debe  or 0))
        th   = Decimal(str(r.total_haber or 0))
        # Saldo según naturaleza de la cuenta
        if acc:
            if acc.nature == 'deudora':
                saldo_deudor  = max(td - th, Decimal('0'))
                saldo_acreedor = max(th - td, Decimal('0'))
            else:
                saldo_acreedor = max(th - td, Decimal('0'))
                saldo_deudor   = max(td - th, Decimal('0'))
            acc_type = acc.type
            acc_name = acc.name
        else:
            saldo_deudor   = max(td - th, Decimal('0'))
            saldo_acreedor = max(th - td, Decimal('0'))
            acc_type = _infer_type(code)
            acc_name = '(sin descripción)'

        accounts.append({
            'code':            code,
            'name':            acc_name,
            'type':            acc_type,
            'total_debe':      td,
            'total_haber':     th,
            'saldo_deudor':    saldo_deudor,
            'saldo_acreedor':  saldo_acreedor,
        })
        grand_debe  += td
        grand_haber += th

    periods = AccountingPeriod.query.order_by(
        AccountingPeriod.year.desc(), AccountingPeriod.month.desc()
    ).all()

    return render_template(
        'contabilidad/balance.html',
        accounts=accounts,
        grand_debe=grand_debe,
        grand_haber=grand_haber,
        grand_saldo_deudor=sum(a['saldo_deudor'] for a in accounts),
        grand_saldo_acreedor=sum(a['saldo_acreedor'] for a in accounts),
        periods=periods,
        selected_year=year,
        selected_month=month,
        user=current_user,
    )


@contabilidad_bp.route('/balance/export')
@login_required
@require_role('Master')
def export_balance():
    from app.models.journal_entry import JournalEntry
    from app.models.journal_entry_line import JournalEntryLine
    from app.models.accounting_account import AccountingAccount
    from sqlalchemy import func, extract

    year  = request.args.get('year',  type=int, default=date.today().year)
    month = request.args.get('month', type=int, default=date.today().month)

    rows = db.session.query(
        JournalEntryLine.account_code,
        func.sum(JournalEntryLine.debe).label('total_debe'),
        func.sum(JournalEntryLine.haber).label('total_haber'),
    ).join(JournalEntry, JournalEntryLine.journal_entry_id == JournalEntry.id
    ).filter(
        extract('year',  JournalEntry.entry_date) == year,
        extract('month', JournalEntry.entry_date) == month,
        JournalEntry.status == 'activo',
    ).group_by(JournalEntryLine.account_code
    ).order_by(JournalEntryLine.account_code).all()

    catalog = {a.code: a for a in AccountingAccount.query.all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Balance Comprobación'

    co = _company_info()
    # Encabezado SUNAT (M-03)
    ws.merge_cells('A1:G1')
    ws['A1'] = co['razon_social']
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:G2')
    ws['A2'] = f'RUC: {co["ruc"]}'
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A3:G3')
    ws['A3'] = f'BALANCE DE COMPROBACIÓN — {_month_name(month)} {year}'
    ws['A3'].font = Font(bold=True)
    ws['A3'].alignment = Alignment(horizontal='center')

    header_fill = PatternFill(start_color='1B3A6B', end_color='1B3A6B', fill_type='solid')
    headers = ['Código', 'Cuenta', 'Tipo', 'DEBE (S/)', 'HABER (S/)', 'Saldo Deudor', 'Saldo Acreedor']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=col, value=h)
        c.fill = header_fill
        c.font = Font(bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='center')

    row = 6
    for r in rows:
        code = r.account_code
        acc  = catalog.get(code)
        td   = float(r.total_debe  or 0)
        th   = float(r.total_haber or 0)
        if acc and acc.nature == 'deudora':
            sd, sa = max(td - th, 0), max(th - td, 0)
        else:
            sa, sd = max(th - td, 0), max(td - th, 0)

        ws.cell(row=row, column=1, value=code)
        ws.cell(row=row, column=2, value=acc.name if acc else '(sin descripción)')
        ws.cell(row=row, column=3, value=acc.type if acc else _infer_type(code))
        ws.cell(row=row, column=4, value=td)
        ws.cell(row=row, column=5, value=th)
        ws.cell(row=row, column=6, value=sd)
        ws.cell(row=row, column=7, value=sa)
        row += 1

    for i, w in enumerate([10, 42, 12, 14, 14, 14, 14], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'balance_comprobacion_{year}{month:02d}.xlsx')


# ── Estado de Resultados ───────────────────────────────────────────────────────

_IR_TRAMO_1_TASA = Decimal('0.10')
_IR_TRAMO_2_TASA = Decimal('0.295')
_UIT_DEFAULT     = Decimal('5350')   # fallback si SystemConfig no tiene valor


def _get_uit() -> Decimal:
    """Lee la UIT vigente desde SystemConfig (M-06). Fallback: S/ 5,350."""
    from app.models.system_config import SystemConfig
    try:
        val = SystemConfig.get('UIT')
        return Decimal(val) if val else _UIT_DEFAULT
    except Exception:
        return _UIT_DEFAULT


def _ir_mype(utilidad: Decimal) -> Decimal:
    """Cálculo IR Régimen MYPE Tributario (anual estimado sobre utilidad del período)."""
    if utilidad <= 0:
        return Decimal('0')
    uit   = _get_uit()
    limit = uit * 15   # S/ 80,250 con UIT 5,350
    if utilidad <= limit:
        return (utilidad * _IR_TRAMO_1_TASA).quantize(Decimal('0.01'))
    else:
        t1 = limit * _IR_TRAMO_1_TASA
        t2 = (utilidad - limit) * _IR_TRAMO_2_TASA
        return (t1 + t2).quantize(Decimal('0.01'))


@contabilidad_bp.route('/resultados')
@login_required
@require_role('Master')
def resultados():
    from app.models.journal_entry import JournalEntry
    from app.models.journal_entry_line import JournalEntryLine
    from app.models.accounting_account import AccountingAccount
    from app.models.accounting_period import AccountingPeriod
    from sqlalchemy import func, extract

    year  = request.args.get('year',  type=int, default=date.today().year)
    month = request.args.get('month', type=int, default=date.today().month)

    rows = db.session.query(
        JournalEntryLine.account_code,
        func.sum(JournalEntryLine.debe).label('total_debe'),
        func.sum(JournalEntryLine.haber).label('total_haber'),
    ).join(JournalEntry, JournalEntryLine.journal_entry_id == JournalEntry.id
    ).filter(
        extract('year',  JournalEntry.entry_date) == year,
        extract('month', JournalEntry.entry_date) == month,
        JournalEntry.status == 'activo',
    ).group_by(JournalEntryLine.account_code
    ).order_by(JournalEntryLine.account_code).all()

    catalog = {a.code: a for a in AccountingAccount.query.all()}

    # Clasificar por tipo para P&L
    ingresos = []
    gastos   = []

    for r in rows:
        code = r.account_code
        acc  = catalog.get(code)
        acc_type = acc.type if acc else _infer_type(code)
        acc_name = acc.name if acc else f'Cuenta {code}'
        td = Decimal(str(r.total_debe  or 0))
        th = Decimal(str(r.total_haber or 0))

        if acc_type == 'ingreso':
            # Cuentas de ingreso: naturaleza acreedora → saldo neto = haber - debe
            neto = max(th - td, Decimal('0'))
            if neto > 0:
                ingresos.append({'code': code, 'name': acc_name, 'monto': neto})

        elif acc_type == 'gasto':
            # Cuentas de gasto: naturaleza deudora → saldo neto = debe - haber
            neto = max(td - th, Decimal('0'))
            if neto > 0:
                gastos.append({'code': code, 'name': acc_name, 'monto': neto})

    total_ingresos = sum(i['monto'] for i in ingresos)
    total_gastos   = sum(g['monto'] for g in gastos)
    utilidad_antes_ir = total_ingresos - total_gastos
    ir_estimado = _ir_mype(utilidad_antes_ir)
    utilidad_neta = utilidad_antes_ir - ir_estimado

    periods = AccountingPeriod.query.order_by(
        AccountingPeriod.year.desc(), AccountingPeriod.month.desc()
    ).all()

    return render_template(
        'contabilidad/resultados.html',
        ingresos=ingresos,
        gastos=gastos,
        total_ingresos=total_ingresos,
        total_gastos=total_gastos,
        utilidad_antes_ir=utilidad_antes_ir,
        ir_estimado=ir_estimado,
        utilidad_neta=utilidad_neta,
        periods=periods,
        selected_year=year,
        selected_month=month,
        UIT=_get_uit(),
        IR_LIMITE=_get_uit() * 15,
        user=current_user,
    )


@contabilidad_bp.route('/resultados/export')
@login_required
@require_role('Master')
def export_resultados():
    from app.models.journal_entry import JournalEntry
    from app.models.journal_entry_line import JournalEntryLine
    from app.models.accounting_account import AccountingAccount
    from sqlalchemy import func, extract

    year  = request.args.get('year',  type=int, default=date.today().year)
    month = request.args.get('month', type=int, default=date.today().month)

    rows = db.session.query(
        JournalEntryLine.account_code,
        func.sum(JournalEntryLine.debe).label('total_debe'),
        func.sum(JournalEntryLine.haber).label('total_haber'),
    ).join(JournalEntry, JournalEntryLine.journal_entry_id == JournalEntry.id
    ).filter(
        extract('year',  JournalEntry.entry_date) == year,
        extract('month', JournalEntry.entry_date) == month,
        JournalEntry.status == 'activo',
    ).group_by(JournalEntryLine.account_code
    ).order_by(JournalEntryLine.account_code).all()

    catalog = {a.code: a for a in AccountingAccount.query.all()}

    ingresos, gastos = [], []
    for r in rows:
        code = r.account_code
        acc  = catalog.get(code)
        acc_type = acc.type if acc else _infer_type(code)
        acc_name = acc.name if acc else f'Cuenta {code}'
        td = float(r.total_debe or 0)
        th = float(r.total_haber or 0)
        if acc_type == 'ingreso':
            neto = max(th - td, 0)
            if neto > 0:
                ingresos.append({'code': code, 'name': acc_name, 'monto': neto})
        elif acc_type == 'gasto':
            neto = max(td - th, 0)
            if neto > 0:
                gastos.append({'code': code, 'name': acc_name, 'monto': neto})

    total_i = sum(i['monto'] for i in ingresos)
    total_g = sum(g['monto'] for g in gastos)
    util    = total_i - total_g
    ir      = float(_ir_mype(Decimal(str(util))))
    util_n  = util - ir

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Estado de Resultados'

    header_fill = PatternFill(start_color='1B3A6B', end_color='1B3A6B', fill_type='solid')
    green_fill  = PatternFill(start_color='198754', end_color='198754', fill_type='solid')
    red_fill    = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')

    co = _company_info()
    ws.merge_cells('A1:C1')
    ws['A1'] = co['razon_social']
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:C2')
    ws['A2'] = f'RUC: {co["ruc"]}'
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A3:C3')
    ws['A3'] = f'ESTADO DE RESULTADOS — {_month_name(month)} {year}'
    ws['A3'].font = Font(bold=True)
    ws['A3'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A4:C4')
    ws['A4'] = 'Régimen MYPE Tributario'
    ws['A4'].alignment = Alignment(horizontal='center')

    def hrow(r, text, fill=None):
        ws.merge_cells(f'A{r}:C{r}')
        ws[f'A{r}'] = text
        ws[f'A{r}'].font = Font(bold=True, color='FFFFFF' if fill else '000000')
        if fill:
            for c in ['A', 'B', 'C']:
                ws[f'{c}{r}'].fill = fill

    def drow(r, code, name, monto):
        ws.cell(row=r, column=1, value=code)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=monto)
        ws.cell(row=r, column=3).number_format = '#,##0.00'

    row = 6
    hrow(row, 'INGRESOS', header_fill); row += 1
    for i in ingresos:
        drow(row, i['code'], i['name'], i['monto']); row += 1
    drow(row, '', 'TOTAL INGRESOS', total_i)
    ws.cell(row=row, column=3).font = Font(bold=True)
    ws.cell(row=row, column=3).fill = green_fill
    ws.cell(row=row, column=3).font = Font(bold=True, color='FFFFFF')
    row += 2

    hrow(row, 'GASTOS', header_fill); row += 1
    for g in gastos:
        drow(row, g['code'], g['name'], g['monto']); row += 1
    drow(row, '', 'TOTAL GASTOS', total_g)
    ws.cell(row=row, column=3).font = Font(bold=True)
    row += 2

    drow(row, '', 'UTILIDAD ANTES DE IR', util)
    ws.cell(row=row, column=3).font = Font(bold=True)
    row += 1
    drow(row, '', f'(-) IR MYPE Tributario', ir); row += 1
    drow(row, '', 'UTILIDAD NETA', util_n)
    ws.cell(row=row, column=3).font = Font(bold=True)
    ws.cell(row=row, column=3).fill = green_fill if util_n >= 0 else red_fill
    ws.cell(row=row, column=3).font = Font(bold=True, color='FFFFFF')

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 48
    ws.column_dimensions['C'].width = 16

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'estado_resultados_{year}{month:02d}.xlsx')


# ── Balance General (Estado de Situación Financiera) ──────────────────────────

def _saldo_acumulado_hasta(account_prefix: str, hasta_fecha, saldo_normal: str = 'deudora') -> Decimal:
    """
    Calcula el saldo acumulado de todas las cuentas que empiezan con account_prefix,
    desde el primer asiento hasta hasta_fecha (inclusive).
    saldo_normal: 'deudora' (activos/gastos) o 'acreedora' (pasivos/patrimonio/ingresos).
    """
    from app.models.journal_entry import JournalEntry
    from app.models.journal_entry_line import JournalEntryLine
    from sqlalchemy import func

    q = db.session.query(
        func.sum(JournalEntryLine.debe).label('d'),
        func.sum(JournalEntryLine.haber).label('h'),
    ).join(JournalEntry, JournalEntryLine.journal_entry_id == JournalEntry.id
    ).filter(
        JournalEntryLine.account_code.like(f'{account_prefix}%'),
        JournalEntry.entry_date <= hasta_fecha,
        JournalEntry.status == 'activo',
    ).first()

    d = Decimal(str(q.d or 0))
    h = Decimal(str(q.h or 0))
    return (d - h) if saldo_normal == 'deudora' else (h - d)


@contabilidad_bp.route('/balance-general')
@login_required
@require_role('Master')
def balance_general():
    """
    Balance General (Estado de Situación Financiera) a una fecha de corte.
    Presenta saldos acumulados (no solo del mes) en formato ACTIVO / PASIVO / PATRIMONIO.
    """
    from app.models.accounting_period import AccountingPeriod
    from app.models.fixed_asset import FixedAsset
    import calendar

    year  = request.args.get('year',  type=int, default=date.today().year)
    month = request.args.get('month', type=int, default=date.today().month)
    _, last_day = calendar.monthrange(year, month)
    corte = date(year, month, last_day)

    def saldo_d(prefix): return _saldo_acumulado_hasta(prefix, corte, 'deudora')
    def saldo_a(prefix): return _saldo_acumulado_hasta(prefix, corte, 'acreedora')

    # ── ACTIVO ────────────────────────────────────────────────────────────────
    caja_mn     = saldo_d('1011')
    caja_me     = saldo_d('1012')
    bancos_pen  = sum(saldo_d(c) for c in ('1041','1048','1049','1051'))
    bancos_usd  = sum(saldo_d(c) for c in ('1044','1047','1050','1052'))
    ctas_cobrar = saldo_d('121')
    activo_corriente = caja_mn + caja_me + bancos_pen + bancos_usd + ctas_cobrar

    # Activos fijos netos (cost - deprec. acum.)
    activos_netos = Decimal('0')
    fixed_assets  = FixedAsset.query.filter_by(status='activo').all()
    for fa in fixed_assets:
        activos_netos += fa.net_book_value
    # Fallback desde asientos si no hay fixed_assets registrados
    if not fixed_assets:
        costo_af     = sum(saldo_d(c) for c in ('3321','3351','3361','3362'))
        deprec_acum  = sum(saldo_a(c) for c in ('3921','3951','3961','3962'))
        activos_netos = costo_af - deprec_acum

    total_activo = activo_corriente + activos_netos

    # ── PASIVO ────────────────────────────────────────────────────────────────
    facturas_pagar  = saldo_a('4211')
    otras_ctas_pag  = saldo_a('4699')
    sueldos_pagar   = saldo_a('4111')
    ir_pagar        = saldo_a('4017')
    essalud_pagar   = saldo_a('4031')
    afp_pagar       = saldo_a('4032')
    total_pasivo    = facturas_pagar + otras_ctas_pag + sueldos_pagar + ir_pagar + essalud_pagar + afp_pagar

    # ── PATRIMONIO ────────────────────────────────────────────────────────────
    capital         = saldo_a('501')
    utilidades_acc  = saldo_a('591')
    perdidas_acc    = saldo_d('592')
    # Resultado del ejercicio acumulado (ingresos - gastos hasta la fecha)
    ingresos_ac = saldo_a('77')
    gastos_ac   = saldo_d('6')
    resultado_ejercicio = ingresos_ac - gastos_ac
    total_patrimonio = capital + utilidades_acc - perdidas_acc + resultado_ejercicio

    periods = AccountingPeriod.query.order_by(
        AccountingPeriod.year.desc(), AccountingPeriod.month.desc()
    ).all()

    return render_template(
        'contabilidad/balance_general.html',
        corte=corte,
        caja_mn=caja_mn, caja_me=caja_me,
        bancos_pen=bancos_pen, bancos_usd=bancos_usd,
        ctas_cobrar=ctas_cobrar,
        activo_corriente=activo_corriente,
        activos_netos=activos_netos,
        total_activo=total_activo,
        facturas_pagar=facturas_pagar, otras_ctas_pag=otras_ctas_pag,
        sueldos_pagar=sueldos_pagar, ir_pagar=ir_pagar,
        essalud_pagar=essalud_pagar, afp_pagar=afp_pagar,
        total_pasivo=total_pasivo,
        capital=capital, utilidades_acc=utilidades_acc,
        perdidas_acc=perdidas_acc,
        resultado_ejercicio=resultado_ejercicio,
        total_patrimonio=total_patrimonio,
        periods=periods,
        selected_year=year, selected_month=month,
        user=current_user,
    )


# ── Activos Fijos ──────────────────────────────────────────────────────────────

@contabilidad_bp.route('/activos-fijos')
@login_required
@require_role('Master')
def activos_fijos():
    from app.models.fixed_asset import FixedAsset
    assets = FixedAsset.query.order_by(FixedAsset.acquisition_date.desc()).all()
    return render_template(
        'contabilidad/activos_fijos.html',
        assets=assets,
        user=current_user,
    )


@contabilidad_bp.route('/activos-fijos/api/list')
@login_required
@require_role('Master')
def api_activos_list():
    from app.models.fixed_asset import FixedAsset
    status_filter = request.args.get('status', 'activo')
    q = FixedAsset.query
    if status_filter != 'todos':
        q = q.filter_by(status=status_filter)
    assets = q.order_by(FixedAsset.acquisition_date.desc()).all()
    return jsonify([a.to_dict() for a in assets])


@contabilidad_bp.route('/activos-fijos/depreciar', methods=['POST'])
@login_required
@require_role('Master')
def depreciar_activos():
    """
    Registra la depreciación mensual de todos los activos activos.
    Genera un asiento por activo: DEBE 6814 / HABER 3951/3961/3962.
    Se debe ejecutar una vez por mes (el sistema previene duplicados por period_id + source_id).
    """
    from app.models.fixed_asset import FixedAsset
    from app.services.accounting.journal_service import JournalService
    from app.models.journal_entry import JournalEntry

    data          = request.get_json() or {}
    dep_year      = int(data.get('year',  date.today().year))
    dep_month     = int(data.get('month', date.today().month))
    dep_date      = date(dep_year, dep_month, 1)

    assets = FixedAsset.query.filter_by(status='activo').all()
    if not assets:
        return jsonify({'success': False, 'error': 'No hay activos fijos activos registrados.'})

    period = JournalService.get_or_create_period(dep_date)
    if period.status == 'cerrado':
        return jsonify({
            'success': False,
            'error': f'El período {dep_date.strftime("%m/%Y")} está cerrado.'
        }), 409

    generados = 0
    omitidos  = 0
    errores   = []

    for asset in assets:
        if asset.is_fully_depreciated:
            asset.status = 'depreciado'
            db.session.flush()
            omitidos += 1
            continue

        # Verificar si ya existe asiento de depreciación para este activo en este período
        existing = JournalEntry.query.filter_by(
            entry_type='depreciacion',
            source_type='fixed_asset',
            source_id=asset.id,
            period_id=period.id,
        ).first()
        if existing:
            omitidos += 1
            continue

        cuota = Decimal(str(asset.monthly_depreciation))

        entry = JournalService.create_entry(
            entry_type='depreciacion',
            description=f'Depreciación {asset.asset_code} – {asset.name} '
                        f'({asset.months_depreciated + 1}/{asset.useful_life_months})',
            lines=[
                {
                    'account_code': '6814',
                    'description':  f'Deprec. {asset.asset_code}',
                    'debe':         cuota,
                    'haber':        Decimal('0'),
                    'currency':     'PEN',
                },
                {
                    'account_code': asset.deprec_account,
                    'description':  f'Deprec. acum. {asset.asset_code}',
                    'debe':         Decimal('0'),
                    'haber':        cuota,
                    'currency':     'PEN',
                },
            ],
            source_type='fixed_asset',
            source_id=asset.id,
            entry_date=dep_date,
            created_by=current_user.id,
        )

        if entry:
            asset.months_depreciated    = (asset.months_depreciated or 0) + 1
            asset.accumulated_depreciation = (
                Decimal(str(asset.accumulated_depreciation or 0)) + cuota
            )
            if asset.is_fully_depreciated:
                asset.status = 'depreciado'
            generados += 1
        else:
            errores.append(asset.asset_code)

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

    return jsonify({
        'success': True,
        'message': f'{generados} asiento(s) de depreciación generado(s). '
                   f'{omitidos} activo(s) omitido(s) (ya depreciado o duplicado). '
                   f'{len(errores)} error(es).',
        'generados': generados,
        'omitidos': omitidos,
        'errores': errores,
    })


@contabilidad_bp.route('/activos-fijos/<int:asset_id>/baja', methods=['POST'])
@login_required
@require_role('Master')
def dar_baja_activo(asset_id):
    """Da de baja un activo fijo (status = 'baja')."""
    from app.models.fixed_asset import FixedAsset
    asset = FixedAsset.query.get_or_404(asset_id)
    data  = request.get_json() or {}
    asset.status     = 'baja'
    asset.baja_date  = date.today()
    asset.baja_notes = data.get('notes', '')
    try:
        db.session.commit()
        return jsonify({'success': True, 'message': f'Activo {asset.asset_code} dado de baja.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ── Helper ─────────────────────────────────────────────────────────────────────

def _month_name(m: int) -> str:
    names = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
    return names[m - 1] if 1 <= m <= 12 else str(m)


def _infer_type(code: str) -> str:
    """Inferir tipo de cuenta por el código PCGE cuando no está en el catálogo."""
    if code.startswith('1') or code.startswith('3'):
        return 'activo'
    if code.startswith('4') or code.startswith('46'):
        return 'pasivo'
    if code.startswith('5'):
        return 'patrimonio'
    if code.startswith('7'):
        return 'ingreso'
    if code.startswith('6'):
        return 'gasto'
    return 'activo'


# ── Exportar Excel multi-hoja ───────────────────────────────────────────────────

@contabilidad_bp.route('/exportar-excel')
@login_required
@require_role('Master')
def exportar_excel():
    """
    Genera un workbook Excel con múltiples hojas:
      1. Libro Diario
      2. Balance General
      3. Estado de Resultados
      4. Registro de Compras
      5. Activos Fijos
      6. Cuadre Diario
    """
    from app.models.journal_entry import JournalEntry
    from app.models.journal_entry_line import JournalEntryLine
    from app.models.expense_record import ExpenseRecord
    from app.models.fixed_asset import FixedAsset
    from app.models.accounting_period import AccountingPeriod
    from sqlalchemy import func, extract
    import calendar

    year  = request.args.get('year',  type=int, default=date.today().year)
    month = request.args.get('month', type=int, default=date.today().month)
    _, last_day = calendar.monthrange(year, month)
    corte = date(year, month, last_day)

    info = _company_info()
    company = info['razon_social']
    ruc     = info['ruc']
    mes_label = f"{['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre'][month-1]} {year}"

    # ── Estilos comunes ───────────────────────────────────────────────────────
    H_FILL  = PatternFill('solid', fgColor='1F3864')   # azul oscuro
    H_FONT  = Font(color='FFFFFF', bold=True, size=10)
    T_FONT  = Font(bold=True, size=11)
    S_FONT  = Font(bold=True, size=9, color='444444')
    NUM_FMT = '#,##0.00'
    BORDER  = Border(
        bottom=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
    )

    def header_row(ws, cols, row=1):
        for c, (title, width) in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=title)
            cell.font      = H_FONT
            cell.fill      = H_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = width
        ws.row_dimensions[row].height = 22

    def title_block(ws, title, subtitle=''):
        ws['A1'] = company
        ws['A1'].font = T_FONT
        ws['A2'] = f'RUC: {ruc}'
        ws['A2'].font = Font(size=9, color='555555')
        ws['A3'] = title
        ws['A3'].font = Font(bold=True, size=12)
        ws['A4'] = subtitle
        ws['A4'].font = Font(italic=True, size=9, color='777777')
        ws.append([])   # row 5 blank

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 1 — Libro Diario
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.create_sheet('Libro Diario')
    title_block(ws1, 'LIBRO DIARIO', mes_label)

    cols1 = [
        ('Nro Asiento', 14), ('Fecha', 12), ('Tipo', 14), ('Descripción', 40),
        ('Cuenta', 10), ('Nombre Cuenta', 28), ('DEBE', 14), ('HABER', 14),
    ]
    header_row(ws1, cols1, row=6)

    entries = JournalEntry.query.filter(
        extract('year',  JournalEntry.entry_date) == year,
        extract('month', JournalEntry.entry_date) == month,
        JournalEntry.status == 'activo',
    ).order_by(JournalEntry.entry_date, JournalEntry.id).all()

    from app.models.accounting_account import AccountingAccount
    _catalog_names = {a.code: a.name for a in AccountingAccount.query.all()}

    r = 7
    total_debe = Decimal('0')
    total_haber = Decimal('0')
    alt_fill = PatternFill('solid', fgColor='EFF3FB')

    for entry in entries:
        lines = JournalEntryLine.query.filter_by(journal_entry_id=entry.id).all()
        for i, line in enumerate(lines):
            row_fill = alt_fill if (entry.id % 2 == 0) else None
            cells = [
                entry.entry_number or '',
                entry.entry_date,
                entry.entry_type or '',
                entry.description[:80] if entry.description else '',
                line.account_code,
                _catalog_names.get(line.account_code, ''),
                float(line.debe or 0),
                float(line.haber or 0),
            ]
            for c, val in enumerate(cells, 1):
                cell = ws1.cell(row=r, column=c, value=val)
                if row_fill:
                    cell.fill = row_fill
                if c in (7, 8):
                    cell.number_format = NUM_FMT
                    cell.alignment = Alignment(horizontal='right')
                if c == 2 and isinstance(val, date):
                    cell.number_format = 'DD/MM/YYYY'
            total_debe  += Decimal(str(line.debe or 0))
            total_haber += Decimal(str(line.haber or 0))
            r += 1

    # Totales
    ws1.cell(row=r, column=6, value='TOTALES').font = Font(bold=True)
    ws1.cell(row=r, column=7, value=float(total_debe)).number_format  = NUM_FMT
    ws1.cell(row=r, column=8, value=float(total_haber)).number_format = NUM_FMT
    ws1.cell(row=r, column=7).font = Font(bold=True)
    ws1.cell(row=r, column=8).font = Font(bold=True)
    ws1.freeze_panes = 'A7'

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 2 — Balance General
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Balance General')
    title_block(ws2, 'BALANCE GENERAL (Estado de Situación Financiera)',
                f'Al {corte.strftime("%d/%m/%Y")}')

    def saldo_d(prefix): return _saldo_acumulado_hasta(prefix, corte, 'deudora')
    def saldo_a(prefix): return _saldo_acumulado_hasta(prefix, corte, 'acreedora')

    caja_mn     = saldo_d('1011')
    caja_me     = saldo_d('1012')
    bancos_pen  = sum(saldo_d(c) for c in ('1041','1048','1049','1051'))
    bancos_usd  = sum(saldo_d(c) for c in ('1044','1047','1050','1052'))
    ctas_cobrar = saldo_d('121')
    activo_corriente = caja_mn + caja_me + bancos_pen + bancos_usd + ctas_cobrar

    fixed_assets = FixedAsset.query.filter_by(status='activo').all()
    activos_netos = sum(fa.net_book_value for fa in fixed_assets) if fixed_assets else (
        sum(saldo_d(c) for c in ('3321','3351','3361','3362')) -
        sum(saldo_a(c) for c in ('3921','3951','3961','3962'))
    )
    total_activo = activo_corriente + activos_netos

    facturas_pagar = saldo_a('4211')
    otras_ctas_pag = saldo_a('4699')
    sueldos_pagar  = saldo_a('4111')
    ir_pagar       = saldo_a('4017')
    essalud_pagar  = saldo_a('4031')
    afp_pagar      = saldo_a('4032')
    total_pasivo   = facturas_pagar + otras_ctas_pag + sueldos_pagar + ir_pagar + essalud_pagar + afp_pagar

    capital         = saldo_a('501')
    utilidades_acc  = saldo_a('591')
    perdidas_acc    = saldo_d('592')
    resultado_ejercicio = saldo_a('77') - saldo_d('6')
    total_patrimonio = capital + utilidades_acc - perdidas_acc + resultado_ejercicio

    BG_SECTION = PatternFill('solid', fgColor='D9E1F2')

    bal_rows = [
        ('ACTIVO', '', True),
        ('ACTIVO CORRIENTE', '', False),
        ('1011 – Caja M/N', float(caja_mn), False),
        ('1012 – Caja M/E', float(caja_me), False),
        ('104x/105x – Bancos PEN', float(bancos_pen), False),
        ('104x/105x – Bancos USD', float(bancos_usd), False),
        ('121 – Cuentas por cobrar', float(ctas_cobrar), False),
        ('TOTAL ACTIVO CORRIENTE', float(activo_corriente), True),
        ('', '', False),
        ('ACTIVO NO CORRIENTE', '', False),
        ('33xx – Activos fijos (neto)', float(activos_netos), False),
        ('TOTAL ACTIVO NO CORRIENTE', float(activos_netos), True),
        ('', '', False),
        ('TOTAL ACTIVO', float(total_activo), True),
        ('', '', False),
        ('PASIVO', '', True),
        ('4211 – Facturas por pagar', float(facturas_pagar), False),
        ('4699 – Otras cuentas por pagar', float(otras_ctas_pag), False),
        ('4111 – Remuneraciones por pagar', float(sueldos_pagar), False),
        ('4031 – EsSalud por pagar', float(essalud_pagar), False),
        ('4032 – AFP por pagar', float(afp_pagar), False),
        ('4017 – IR pago a cuenta', float(ir_pagar), False),
        ('TOTAL PASIVO', float(total_pasivo), True),
        ('', '', False),
        ('PATRIMONIO', '', True),
        ('501 – Capital social', float(capital), False),
        ('591 – Utilidades acumuladas', float(utilidades_acc), False),
        ('592 – Pérdidas acumuladas', float(-perdidas_acc), False),
        ('Resultado del ejercicio', float(resultado_ejercicio), False),
        ('TOTAL PATRIMONIO', float(total_patrimonio), True),
        ('', '', False),
        ('TOTAL PASIVO + PATRIMONIO', float(total_pasivo + total_patrimonio), True),
    ]

    header_row(ws2, [('Concepto', 40), ('Importe (S/)', 18)], row=6)
    for i, (label, amount, bold) in enumerate(bal_rows, 7):
        c1 = ws2.cell(row=i, column=1, value=label)
        c2 = ws2.cell(row=i, column=2, value=amount if amount != '' else '')
        if bold:
            c1.font = Font(bold=True)
            c2.font = Font(bold=True)
            c1.fill = c2.fill = BG_SECTION
        if isinstance(amount, float):
            c2.number_format = NUM_FMT
            c2.alignment = Alignment(horizontal='right')

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 3 — Estado de Resultados
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('Estado de Resultados')
    title_block(ws3, 'ESTADO DE RESULTADOS', mes_label)

    # Ingresos por tipo de asiento (calce_netting = utilidad de operaciones)
    ingresos_op = db.session.query(func.sum(JournalEntryLine.haber)).join(
        JournalEntry, JournalEntryLine.journal_entry_id == JournalEntry.id
    ).filter(
        JournalEntryLine.account_code.like('77%'),
        extract('year',  JournalEntry.entry_date) == year,
        extract('month', JournalEntry.entry_date) == month,
        JournalEntry.status == 'activo',
    ).scalar() or Decimal('0')

    # Gastos por categoría
    gastos_q = db.session.query(
        func.sum(ExpenseRecord.amount_pen).label('total'),
    ).filter(
        extract('year',  ExpenseRecord.expense_date) == year,
        extract('month', ExpenseRecord.expense_date) == month,
    ).scalar() or Decimal('0')

    # Gastos de depreciación del mes
    deprec_mes = db.session.query(func.sum(JournalEntryLine.debe)).join(
        JournalEntry, JournalEntryLine.journal_entry_id == JournalEntry.id
    ).filter(
        JournalEntryLine.account_code == '6814',
        extract('year',  JournalEntry.entry_date) == year,
        extract('month', JournalEntry.entry_date) == month,
        JournalEntry.status == 'activo',
    ).scalar() or Decimal('0')

    utilidad_bruta = Decimal(str(ingresos_op)) - Decimal(str(gastos_q))

    ER_FILL = PatternFill('solid', fgColor='E2EFDA')
    er_rows = [
        ('INGRESOS', '', True),
        ('7711 – Utilidad por diferencia de cambio', float(ingresos_op), False),
        ('TOTAL INGRESOS', float(ingresos_op), True),
        ('', '', False),
        ('GASTOS OPERATIVOS', '', True),
        ('6xxx – Gastos del período', float(gastos_q), False),
        ('6814 – Depreciación', float(deprec_mes), False),
        ('TOTAL GASTOS', float(Decimal(str(gastos_q)) + Decimal(str(deprec_mes))), True),
        ('', '', False),
        ('RESULTADO NETO DEL PERÍODO', float(utilidad_bruta - Decimal(str(deprec_mes))), True),
    ]

    header_row(ws3, [('Concepto', 40), ('Importe (S/)', 18)], row=6)
    for i, (label, amount, bold) in enumerate(er_rows, 7):
        c1 = ws3.cell(row=i, column=1, value=label)
        c2 = ws3.cell(row=i, column=2, value=amount if amount != '' else '')
        if bold:
            c1.font = c2.font = Font(bold=True)
            c1.fill = c2.fill = ER_FILL
        if isinstance(amount, float):
            c2.number_format = NUM_FMT
            c2.alignment = Alignment(horizontal='right')

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 4 — Registro de Compras / Gastos
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet('Registro de Compras')
    title_block(ws4, 'REGISTRO DE COMPRAS / GASTOS', mes_label)

    cols4 = [
        ('Fecha', 12), ('Tipo Comp.', 12), ('Nro Comp.', 14), ('RUC Proveedor', 14),
        ('Proveedor', 28), ('Tipo Gasto', 14), ('Descripción', 36),
        ('Base (S/)', 13), ('IGV (S/)', 13), ('Total (S/)', 13),
    ]
    header_row(ws4, cols4, row=6)

    gastos = ExpenseRecord.query.filter(
        extract('year',  ExpenseRecord.expense_date) == year,
        extract('month', ExpenseRecord.expense_date) == month,
    ).order_by(ExpenseRecord.expense_date).all()

    g_total_base = g_total_igv = g_total = Decimal('0')
    for i, g in enumerate(gastos, 7):
        base = float(g.base_pen or g.amount_pen or 0)
        igv  = float(g.igv_pen or 0)
        total = float(g.amount_pen or 0)
        row_fill = alt_fill if i % 2 == 0 else None
        vals = [
            g.expense_date, g.voucher_type or '', g.voucher_number or '',
            g.supplier_ruc or '', g.supplier_name or '',
            g.expense_type or 'servicio', g.description[:60] if g.description else '',
            base, igv, total,
        ]
        for c, val in enumerate(vals, 1):
            cell = ws4.cell(row=i, column=c, value=val)
            if row_fill:
                cell.fill = row_fill
            if c in (8, 9, 10):
                cell.number_format = NUM_FMT
                cell.alignment = Alignment(horizontal='right')
            if c == 1 and isinstance(val, date):
                cell.number_format = 'DD/MM/YYYY'
        g_total_base += Decimal(str(base))
        g_total_igv  += Decimal(str(igv))
        g_total      += Decimal(str(total))

    r4 = len(gastos) + 7
    ws4.cell(row=r4, column=7, value='TOTALES').font = Font(bold=True)
    for c, val in [(8, float(g_total_base)), (9, float(g_total_igv)), (10, float(g_total))]:
        cell = ws4.cell(row=r4, column=c, value=val)
        cell.number_format = NUM_FMT
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='right')
    ws4.freeze_panes = 'A7'

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 5 — Control de Activos Fijos
    # ══════════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet('Activos Fijos')
    title_block(ws5, 'CONTROL DE ACTIVOS FIJOS', f'Al {corte.strftime("%d/%m/%Y")}')

    cols5 = [
        ('Código', 12), ('Descripción', 30), ('Categoría', 14),
        ('Cta. Activo', 12), ('Cta. Deprec.', 12), ('Fecha Adq.', 13),
        ('Costo (S/)', 13), ('Deprec./mes', 12), ('Deprec. Acum.', 14),
        ('Valor Neto', 13), ('Meses', 8), ('Vida Útil', 10), ('Estado', 10),
    ]
    header_row(ws5, cols5, row=6)

    all_assets = FixedAsset.query.order_by(FixedAsset.acquisition_date).all()
    for i, a in enumerate(all_assets, 7):
        row_fill = alt_fill if i % 2 == 0 else None
        vals = [
            a.asset_code, a.name, a.category,
            a.account_code, a.deprec_account,
            a.acquisition_date, float(a.cost_pen),
            float(a.monthly_depreciation),
            float(a.accumulated_depreciation or 0),
            float(a.net_book_value),
            a.months_depreciated or 0, a.useful_life_months, a.status,
        ]
        for c, val in enumerate(vals, 1):
            cell = ws5.cell(row=i, column=c, value=val)
            if row_fill:
                cell.fill = row_fill
            if c in (7, 8, 9, 10):
                cell.number_format = NUM_FMT
                cell.alignment = Alignment(horizontal='right')
            if c == 6 and isinstance(val, date):
                cell.number_format = 'DD/MM/YYYY'

    ws5.freeze_panes = 'A7'

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 6 — Cuadre Diario
    # ══════════════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet('Cuadre Diario')
    title_block(ws6, 'CUADRE DIARIO DE ASIENTOS', mes_label)

    cols6 = [('Fecha', 12), ('Nro Asientos', 14), ('Total DEBE', 14), ('Total HABER', 14), ('Diferencia', 14)]
    header_row(ws6, cols6, row=6)

    daily = db.session.query(
        JournalEntry.entry_date,
        func.count(JournalEntry.id).label('count'),
        func.sum(JournalEntry.total_debe).label('debe'),
        func.sum(JournalEntry.total_haber).label('haber'),
    ).filter(
        extract('year',  JournalEntry.entry_date) == year,
        extract('month', JournalEntry.entry_date) == month,
        JournalEntry.status == 'activo',
    ).group_by(JournalEntry.entry_date).order_by(JournalEntry.entry_date).all()

    for i, row in enumerate(daily, 7):
        debe  = float(row.debe  or 0)
        haber = float(row.haber or 0)
        diff  = debe - haber
        vals  = [row.entry_date, row.count, debe, haber, diff]
        for c, val in enumerate(vals, 1):
            cell = ws6.cell(row=i, column=c, value=val)
            if c in (3, 4, 5):
                cell.number_format = NUM_FMT
                cell.alignment = Alignment(horizontal='right')
            if c == 1:
                cell.number_format = 'DD/MM/YYYY'
            if c == 5 and abs(diff) > 0.01:
                cell.font = Font(color='FF0000', bold=True)

    ws6.freeze_panes = 'A7'

    # ── Enviar archivo ────────────────────────────────────────────────────────
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'QoriCash_Contabilidad_{year}{month:02d}.xlsx'
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# ── Amarres y Lotes de Neteo ───────────────────────────────────────────────────

@contabilidad_bp.route('/amarres')
@login_required
@require_role('Master')
def amarres():
    """Página de gestión de amarres y lotes de neteo."""
    return render_template('contabilidad/matches.html', user=current_user)


@contabilidad_bp.route('/amarres/api/operaciones-disponibles')
@login_required
@require_role('Master')
def amarres_operaciones_disponibles():
    """API: operaciones completadas con monto disponible para amarrar."""
    from app.services.accounting_service import AccountingService
    from sqlalchemy import inspect as sa_inspect
    try:
        # Verificar que las tablas de amarres existen; si no, crearlas ahora
        inspector = sa_inspect(db.engine)
        if 'accounting_matches' not in inspector.get_table_names():
            db.create_all()

        ops = AccountingService.get_available_operations(
            fecha_inicio=request.args.get('fecha_inicio'),
            fecha_fin=request.args.get('fecha_fin'),
            operation_type=request.args.get('operation_type'),
        )
        results = []
        for op in ops:
            try:
                available = float(AccountingService.get_available_amount_for_operation(op.id))
            except Exception:
                available = float(op.amount_usd)
            if available > 0:
                try:
                    matched = float(AccountingService.get_matched_amount_for_operation(op.id))
                except Exception:
                    matched = 0.0
                results.append({
                    'id': op.id,
                    'operation_id': op.operation_id,
                    'operation_type': op.operation_type,
                    'amount_usd': float(op.amount_usd),
                    'exchange_rate': float(op.exchange_rate),
                    'amount_pen': float(op.amount_pen),
                    'base_rate': float(op.base_rate) if op.base_rate else None,
                    'client_name': op.client.full_name if op.client else 'N/A',
                    'trader_name': op.user.username if op.user else None,
                    'completed_at': op.completed_at.isoformat() if op.completed_at else None,
                    'available_usd': available,
                    'matched_usd': matched,
                })
        return jsonify({'success': True, 'operations': results})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@contabilidad_bp.route('/amarres/crear', methods=['POST'])
@login_required
@require_role('Master')
def amarres_crear():
    """API: crear amarre entre compra y venta."""
    from app.services.accounting_service import AccountingService
    try:
        data = request.get_json() or {}
        buy_op_id  = data.get('buy_operation_id')
        sell_op_id = data.get('sell_operation_id')
        amount_usd = data.get('matched_amount_usd')
        notes      = data.get('notes', '')

        if not buy_op_id or not sell_op_id:
            return jsonify({'success': False, 'error': 'Debe seleccionar ambas operaciones'}), 400
        if not amount_usd or float(amount_usd) <= 0:
            return jsonify({'success': False, 'error': 'El monto debe ser mayor a cero'}), 400

        success, message, match = AccountingService.create_match(
            buy_operation_id=buy_op_id,
            sell_operation_id=sell_op_id,
            matched_amount_usd=amount_usd,
            user_id=current_user.id,
            notes=notes,
        )
        if success:
            return jsonify({'success': True, 'message': message, 'match': match.to_dict()})
        return jsonify({'success': False, 'error': message}), 400
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': f'Error interno: {str(e)}'}), 500


@contabilidad_bp.route('/amarres/<int:match_id>/anular', methods=['DELETE'])
@login_required
@require_role('Master')
def amarres_anular(match_id):
    """API: anular amarre."""
    from app.services.accounting_service import AccountingService
    try:
        success, message = AccountingService.delete_match(match_id, current_user.id)
        if success:
            return jsonify({'success': True, 'message': message})
        return jsonify({'success': False, 'error': message}), 400
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@contabilidad_bp.route('/amarres/lista')
@login_required
@require_role('Master')
def amarres_lista():
    """API: listar amarres con filtros opcionales."""
    from app.services.accounting_service import AccountingService
    try:
        matches = AccountingService.get_all_matches(
            fecha_inicio=request.args.get('fecha_inicio'),
            fecha_fin=request.args.get('fecha_fin'),
            batch_id=request.args.get('batch_id'),
            status=request.args.get('status', 'Activo'),
        )
        return jsonify({'success': True, 'matches': [m.to_dict() for m in matches]})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@contabilidad_bp.route('/lotes/crear', methods=['POST'])
@login_required
@require_role('Master')
def lotes_crear():
    """API: crear lote de neteo."""
    from app.services.accounting_service import AccountingService
    from datetime import datetime as dt
    try:
        data = request.get_json() or {}
        netting_date = data.get('netting_date') or dt.now().strftime('%Y-%m-%d')
        success, message, batch = AccountingService.create_batch(
            match_ids=data.get('match_ids', []),
            description=data.get('description', ''),
            netting_date=netting_date,
            user_id=current_user.id,
        )
        if success:
            return jsonify({'success': True, 'message': message, 'batch': batch.to_dict(include_matches=True)})
        return jsonify({'success': False, 'error': message}), 400
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@contabilidad_bp.route('/lotes/lista')
@login_required
@require_role('Master')
def lotes_lista():
    """API: listar lotes de neteo."""
    from app.services.accounting_service import AccountingService
    try:
        batches = AccountingService.get_all_batches(
            fecha_inicio=request.args.get('fecha_inicio'),
            fecha_fin=request.args.get('fecha_fin'),
            status=request.args.get('status'),
        )
        return jsonify({'success': True, 'batches': [b.to_dict() for b in batches]})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@contabilidad_bp.route('/lotes/<int:batch_id>')
@login_required
@require_role('Master')
def lotes_detalle(batch_id):
    """API: detalle de un lote."""
    from app.models import AccountingBatch
    try:
        batch = AccountingBatch.query.get(batch_id)
        if not batch:
            return jsonify({'success': False, 'error': 'Lote no encontrado'}), 404
        return jsonify({'success': True, 'batch': batch.to_dict(include_matches=True)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@contabilidad_bp.route('/lotes/<int:batch_id>/cerrar', methods=['POST'])
@login_required
@require_role('Master')
def lotes_cerrar(batch_id):
    """API: cerrar lote de neteo."""
    from app.services.accounting_service import AccountingService
    try:
        success, message = AccountingService.close_batch(batch_id, current_user.id)
        if success:
            return jsonify({'success': True, 'message': message})
        return jsonify({'success': False, 'error': message}), 400
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@contabilidad_bp.route('/amarres/export')
@login_required
@require_role('Master')
def amarres_export():
    """Exportar amarres y lotes en Excel multi-hoja (3 hojas)."""
    from app.services.accounting_service import AccountingService
    from app.models.user import User
    from collections import defaultdict

    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin    = request.args.get('fecha_fin')
    status_fil   = request.args.get('status', '')

    matches = AccountingService.get_all_matches(
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        status=status_fil if status_fil else None,
    )
    batches = AccountingService.get_all_batches(
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
    )

    co      = _company_info()
    now_str = date.today().strftime('%d/%m/%Y')

    H_FILL   = PatternFill('solid', fgColor='1B3A6B')
    H_FONT   = Font(bold=True, color='FFFFFF', size=10)
    T_FONT   = Font(bold=True, size=11)
    SUB_FILL = PatternFill('solid', fgColor='E8EDF4')
    TOT_FONT = Font(bold=True, size=10)
    NUM_FMT  = '#,##0.00'
    thin     = Side(style='thin', color='CCCCCC')
    BRD      = Border(left=thin, right=thin, top=thin, bottom=thin)
    EVEN     = PatternFill('solid', fgColor='F4F7FB')

    TIPO_MAP = {
        'client_to_client': 'C2C',
        'self_match':       'Self-match',
        'market_hedge':     'Hedge',
    }

    def title_block(ws, title):
        ws['A1'] = co['razon_social'];     ws['A1'].font = T_FONT
        ws['A2'] = f'RUC: {co["ruc"]}';   ws['A2'].font = Font(size=9, color='555555')
        ws['A3'] = title;                  ws['A3'].font = Font(bold=True, size=12)
        ws['A4'] = f'Generado: {now_str}'; ws['A4'].font = Font(italic=True, size=9, color='777777')

    def hrow(ws, row_n, headers_widths):
        for col, (h, w) in enumerate(headers_widths, 1):
            c = ws.cell(row=row_n, column=col, value=h)
            c.fill = H_FILL; c.font = H_FONT; c.border = BRD
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
        ws.row_dimensions[row_n].height = 28

    def wnum(ws, r, c, v):
        cell = ws.cell(row=r, column=c, value=round(float(v or 0), 2))
        cell.number_format = NUM_FMT; cell.border = BRD
        return cell

    def wtxt(ws, r, c, v, **kw):
        cell = ws.cell(row=r, column=c, value=v or '')
        cell.border = BRD
        for k, val in kw.items(): setattr(cell, k, val)
        return cell

    user_map = {u.id: u.username for u in User.query.all()}
    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════════════════════
    # HOJA 1 — AMARRES
    # ══════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Amarres'
    title_block(ws1, 'REPORTE DE AMARRES')
    ws1.freeze_panes = 'A7'

    h1 = [
        ('ID',              5), ('Tipo',          11), ('Fecha',          11),
        ('Op. Compra',     14), ('Cliente Compra', 24), ('Trader Compra',  18),
        ('Op. Venta',      14), ('Cliente Venta',  24), ('Trader Venta',   18),
        ('USD Amarrado',   13), ('TC Compra',      10), ('TC Venta',       10),
        ('Base Compra',    10), ('Base Venta',     10),
        ('Util. Total S/', 15),
        ('Util. Tdr. Compra S/', 18),
        ('Util. Tdr. Venta S/',  18),
        ('Util. QoriCash S/',    16),
        ('Lote',            9), ('Estado',          9),
    ]
    hrow(ws1, 6, h1)

    row = 7
    t_usd = t_util = t_ubuy = t_usell = t_uhouse = 0
    for m in matches:
        buy_op  = m.buy_operation
        sell_op = m.sell_operation
        bt = user_map.get(buy_op.user_id,  '—') if buy_op  else '—'
        st = user_map.get(sell_op.user_id, '—') if sell_op else '—'
        usd  = float(m.matched_amount_usd)
        util = float(m.profit_pen               or 0)
        ubuy = float(m.trader_buy_profit_pen    or 0)
        usel = float(m.trader_sell_profit_pen   or 0)
        uhou = float(m.house_profit_pen         or 0)
        t_usd += usd; t_util += util; t_ubuy += ubuy; t_usell += usel; t_uhouse += uhou
        fill = EVEN if row % 2 == 0 else None

        def wr(col, val, is_num=False, **kw):
            c = wnum(ws1, row, col, val) if is_num else wtxt(ws1, row, col, val, **kw)
            if fill: c.fill = fill
            return c

        wr(1,  m.id)
        wr(2,  TIPO_MAP.get(m.match_type, m.match_type or 'C2C'))
        wr(3,  m.created_at.strftime('%d/%m/%Y') if m.created_at else '')
        wr(4,  buy_op.operation_id            if buy_op  else '—')
        wr(5,  buy_op.client.full_name        if buy_op  and buy_op.client  else '—')
        wr(6,  bt)
        wr(7,  sell_op.operation_id           if sell_op else '—')
        wr(8,  sell_op.client.full_name       if sell_op and sell_op.client else '—')
        wr(9,  st)
        wr(10, usd,  True); wr(11, float(m.buy_exchange_rate),  True)
        wr(12, float(m.sell_exchange_rate), True)
        wr(13, float(m.buy_base_rate  or 0), True)
        wr(14, float(m.sell_base_rate or 0), True)
        wr(15, util, True); wr(16, ubuy, True); wr(17, usel, True); wr(18, uhou, True)
        wr(19, f'#{m.batch_id}' if m.batch_id else '—')
        ec = wr(20, m.status)
        ec.font = Font(bold=True, color='1D7A3A' if m.status == 'Activo' else '888888')
        row += 1

    # Totales hoja 1
    ws1.cell(row=row, column=9, value='TOTALES').font = TOT_FONT
    for col, val in [(10, t_usd), (15, t_util), (16, t_ubuy), (17, t_usell), (18, t_uhouse)]:
        c = wnum(ws1, row, col, val); c.fill = SUB_FILL; c.font = TOT_FONT

    # ══════════════════════════════════════════════════════════════════════
    # HOJA 2 — LOTES DE NETEO
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Lotes de Neteo')
    title_block(ws2, 'LOTES DE NETEO')
    ws2.freeze_panes = 'A7'

    h2 = [
        ('Código',        14), ('Descripción',    36), ('Fecha Neteo',  13),
        ('N° Amarres',    12), ('USD Total',       13), ('Util. Total S/', 16),
        ('Estado',         9),
    ]
    hrow(ws2, 6, h2)

    row2 = 7; bt_usd = bt_util = 0
    for b in batches:
        usd_b  = float(b.total_buys_usd or 0) + float(b.total_sells_usd or 0)
        util_b = float(b.total_profit_pen or 0)
        bt_usd += usd_b; bt_util += util_b
        fill = EVEN if row2 % 2 == 0 else None

        def wb2(col, val, is_num=False, **kw):
            c = wnum(ws2, row2, col, val) if is_num else wtxt(ws2, row2, col, val, **kw)
            if fill: c.fill = fill
            return c

        wb2(1, b.batch_code or str(b.id))
        wb2(2, b.description or '')
        wb2(3, b.netting_date.strftime('%d/%m/%Y') if b.netting_date else '')
        wb2(4, b.num_matches or 0)
        wb2(5, usd_b,  True); wb2(6, util_b, True)
        ec = wb2(7, b.status)
        ec.font = Font(bold=True, color='1D7A3A' if b.status == 'Abierto' else '888888')
        row2 += 1

    ws2.cell(row=row2, column=4, value='TOTALES').font = TOT_FONT
    for col, val in [(5, bt_usd), (6, bt_util)]:
        c = wnum(ws2, row2, col, val); c.fill = SUB_FILL; c.font = TOT_FONT

    # ══════════════════════════════════════════════════════════════════════
    # HOJA 3 — RESUMEN POR TRADER
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('Resumen por Trader')
    title_block(ws3, 'UTILIDAD POR TRADER')
    ws3.freeze_panes = 'A7'

    h3 = [
        ('Trader',           20), ('Amarres Compra', 14), ('Amarres Venta',  14),
        ('Total Amarres',    13), ('USD Operado',    13),
        ('Util. Comprador S/', 20), ('Util. Vendedor S/', 20),
        ('Util. Total Trader S/', 20),
    ]
    hrow(ws3, 6, h3)

    stats = defaultdict(lambda: {
        'bc': 0, 'sc': 0, 'usd': 0.0, 'ubuy': 0.0, 'usell': 0.0
    })
    for m in matches:
        if m.status != 'Activo':
            continue
        usd = float(m.matched_amount_usd)
        buy_op  = m.buy_operation
        sell_op = m.sell_operation
        if buy_op and buy_op.user_id:
            t = user_map.get(buy_op.user_id, f'ID {buy_op.user_id}')
            stats[t]['bc']   += 1
            stats[t]['usd']  += usd
            stats[t]['ubuy'] += float(m.trader_buy_profit_pen or 0)
        if sell_op and sell_op.user_id:
            t = user_map.get(sell_op.user_id, f'ID {sell_op.user_id}')
            stats[t]['sc']    += 1
            # Evitar doble suma USD en self-match
            if not (buy_op and buy_op.user_id == sell_op.user_id):
                stats[t]['usd'] += usd
            stats[t]['usell'] += float(m.trader_sell_profit_pen or 0)

    row3 = 7; gt_usd = gt_ubuy = gt_usell = 0
    for tname in sorted(stats):
        s = stats[tname]
        total = s['ubuy'] + s['usell']
        fill  = EVEN if row3 % 2 == 0 else None

        def wt3(col, val, is_num=False, bold=False):
            c = wnum(ws3, row3, col, val) if is_num else wtxt(ws3, row3, col, val)
            if fill:  c.fill = fill
            if bold:  c.font = Font(bold=True)
            return c

        wt3(1, tname); wt3(2, s['bc']); wt3(3, s['sc'])
        wt3(4, s['bc'] + s['sc'])
        wt3(5, s['usd'],   True)
        wt3(6, s['ubuy'],  True)
        wt3(7, s['usell'], True)
        wt3(8, total,      True, bold=True)
        gt_usd += s['usd']; gt_ubuy += s['ubuy']; gt_usell += s['usell']
        row3 += 1

    ws3.cell(row=row3, column=4, value='TOTALES').font = TOT_FONT
    for col, val in [(5, gt_usd), (6, gt_ubuy), (7, gt_usell), (8, gt_ubuy + gt_usell)]:
        c = wnum(ws3, row3, col, val); c.fill = SUB_FILL; c.font = TOT_FONT

    # ── Generar archivo ────────────────────────────────────────────────────
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    suffix   = f'{(fecha_inicio or "todo").replace("-", "")}_{(fecha_fin or "hoy").replace("-", "")}'
    filename = f'amarres_qoricash_{suffix}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )
